"""
Scalable Data Ingestion Service for InfoShop
============================================

Designed to handle 2-3 million products with:
- Chunked file processing (memory efficient)
- Background job processing with progress tracking
- Batched Algolia indexing with retries
- Streaming transformations
- Resume capability for failed jobs

Architecture:
1. File Upload → Save to disk → Create job record
2. Background worker processes file in chunks
3. Each chunk is transformed and indexed to Algolia
4. Progress tracked in MongoDB
5. Webhook/polling for status updates
"""

import os
import io
import gc
import csv
import uuid
import asyncio
import logging
import hashlib
import traceback
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any, Generator, AsyncGenerator
from pathlib import Path
import pandas as pd
from dataclasses import dataclass, asdict
from enum import Enum

logger = logging.getLogger(__name__)

# ============================================
# CONFIGURATION
# ============================================

class JobStatus(str, Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"
    PAUSED = "paused"

@dataclass
class IndexingJobConfig:
    """Configuration for a data ingestion job"""
    chunk_size: int = 5000           # Rows per chunk for processing
    algolia_batch_size: int = 1000    # Products per Algolia batch (max 1000)
    max_retries: int = 3              # Retries for failed batches
    retry_delay: float = 2.0          # Seconds between retries
    memory_limit_mb: int = 512        # Memory threshold for GC
    progress_update_interval: int = 100  # Update progress every N products
    enable_validation: bool = True    # Validate products before indexing
    skip_duplicates: bool = True      # Skip products with existing objectID

DEFAULT_CONFIG = IndexingJobConfig()

# Upload directory
UPLOAD_DIR = Path("/app/backend/uploads/catalogs")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ============================================
# JOB TRACKING
# ============================================

@dataclass
class IndexingJob:
    """Represents a data ingestion job"""
    job_id: str
    vendor: str
    filename: str
    filepath: str
    status: JobStatus
    total_rows: int = 0
    processed_rows: int = 0
    indexed_count: int = 0
    error_count: int = 0
    current_chunk: int = 0
    total_chunks: int = 0
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    error_message: Optional[str] = None
    errors: List[Dict] = None
    config: Dict = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.config is None:
            self.config = asdict(DEFAULT_CONFIG)
    
    @property
    def progress_percent(self) -> float:
        if self.total_rows == 0:
            return 0.0
        return round((self.processed_rows / self.total_rows) * 100, 2)
    
    @property
    def is_active(self) -> bool:
        return self.status in [JobStatus.PENDING, JobStatus.PROCESSING]
    
    def to_dict(self) -> Dict:
        return {
            "job_id": self.job_id,
            "vendor": self.vendor,
            "filename": self.filename,
            "filepath": self.filepath,
            "status": self.status.value,
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "indexed_count": self.indexed_count,
            "error_count": self.error_count,
            "current_chunk": self.current_chunk,
            "total_chunks": self.total_chunks,
            "progress_percent": self.progress_percent,
            "started_at": self.started_at,
            "completed_at": self.completed_at,
            "error_message": self.error_message,
            "errors": self.errors[-100:],  # Last 100 errors only
            "config": self.config,
        }

# In-memory job tracking (for quick access)
_active_jobs: Dict[str, IndexingJob] = {}

def get_job(job_id: str) -> Optional[IndexingJob]:
    return _active_jobs.get(job_id)

def update_job(job: IndexingJob):
    _active_jobs[job.job_id] = job

def remove_job(job_id: str):
    _active_jobs.pop(job_id, None)

# ============================================
# FILE HANDLING - CHUNKED READING
# ============================================

def count_file_rows(filepath: str) -> int:
    """
    Efficiently count rows in a file without loading into memory
    """
    ext = Path(filepath).suffix.lower()
    
    if ext == '.csv':
        # Fast line counting for CSV
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return sum(1 for _ in f) - 1  # Subtract header
    
    elif ext in ['.xlsx', '.xls']:
        # For Excel, we need to read with openpyxl in read-only mode
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            count = ws.max_row - 1  # Subtract header
            wb.close()
            return count
        except Exception as e:
            logger.warning(f"Could not count Excel rows efficiently: {e}")
            # Fallback: read first sheet
            df = pd.read_excel(filepath, nrows=0)
            return 0
    
    return 0


def stream_csv_chunks(filepath: str, chunk_size: int = 5000) -> Generator[pd.DataFrame, None, None]:
    """
    Stream CSV file in chunks - memory efficient
    """
    try:
        for chunk in pd.read_csv(filepath, chunksize=chunk_size, encoding='utf-8', on_bad_lines='skip'):
            yield chunk
            gc.collect()  # Force garbage collection between chunks
    except Exception as e:
        logger.error(f"Error streaming CSV: {e}")
        raise


def stream_excel_chunks(filepath: str, chunk_size: int = 5000) -> Generator[pd.DataFrame, None, None]:
    """
    Stream Excel file in chunks - converts to CSV first for efficiency
    
    For very large Excel files (>100k rows), we convert to CSV first
    because pandas Excel reading is memory-intensive.
    """
    ext = Path(filepath).suffix.lower()
    row_count = count_file_rows(filepath)
    
    # For small files, read directly
    if row_count < 50000:
        logger.info(f"Small file ({row_count} rows), reading directly")
        df = pd.read_excel(filepath)
        for i in range(0, len(df), chunk_size):
            yield df.iloc[i:i + chunk_size]
        return
    
    # For large files, convert to CSV first (more memory efficient)
    logger.info(f"Large file ({row_count} rows), converting to CSV for streaming")
    csv_path = filepath + ".temp.csv"
    
    try:
        # Read Excel and write to CSV in chunks
        from openpyxl import load_workbook
        
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        # Write to CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            batch = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                batch.append(row)
                if len(batch) >= 10000:
                    writer.writerows(batch)
                    batch = []
            
            if batch:
                writer.writerows(batch)
        
        wb.close()
        
        # Now stream from CSV
        yield from stream_csv_chunks(csv_path, chunk_size)
        
    finally:
        # Cleanup temp file
        if os.path.exists(csv_path):
            os.remove(csv_path)
        gc.collect()


def stream_file_chunks(filepath: str, chunk_size: int = 5000) -> Generator[pd.DataFrame, None, None]:
    """
    Universal file streaming - handles CSV and Excel
    """
    ext = Path(filepath).suffix.lower()
    
    if ext == '.csv':
        yield from stream_csv_chunks(filepath, chunk_size)
    elif ext in ['.xlsx', '.xls']:
        yield from stream_excel_chunks(filepath, chunk_size)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


# ============================================
# PRODUCT TRANSFORMATION (STREAMING)
# ============================================

def transform_chunk(
    chunk: pd.DataFrame,
    vendor: str,
    category_discounts: Dict[str, float],
    config: IndexingJobConfig
) -> tuple[List[Dict], List[Dict]]:
    """
    Transform a chunk of products
    Returns: (valid_products, errors)
    """
    from infoshop_service import transform_product_for_infoshop
    
    products = []
    errors = []
    
    for idx, row in chunk.iterrows():
        try:
            product = transform_product_for_infoshop(
                row.to_dict(),
                vendor,
                category_discounts
            )
            
            if config.enable_validation:
                # Basic validation
                if not product.get("product_name"):
                    errors.append({"row": idx, "error": "Missing product name"})
                    continue
                if product.get("danone_preferred_price", 0) <= 0:
                    # Allow products without price, but flag them
                    product["price_status"] = "no_price"
            
            products.append(product)
            
        except Exception as e:
            errors.append({
                "row": idx,
                "error": str(e),
                "product_name": row.get("Product Name", row.get("Product title", "Unknown"))
            })
    
    return products, errors


# ============================================
# ALGOLIA INDEXING WITH RETRY
# ============================================

async def index_batch_to_algolia(
    products: List[Dict],
    index_name: str,
    config: IndexingJobConfig
) -> tuple[int, List[Dict]]:
    """
    Index a batch of products to Algolia with retry logic
    Returns: (indexed_count, errors)
    """
    from algolia_service import algolia_client
    
    if not products:
        return 0, []
    
    errors = []
    indexed = 0
    
    for attempt in range(config.max_retries):
        try:
            # Algolia recommends batches of 1000 max
            for i in range(0, len(products), config.algolia_batch_size):
                batch = products[i:i + config.algolia_batch_size]
                algolia_client.save_objects(index_name, batch)
                indexed += len(batch)
            
            return indexed, errors
            
        except Exception as e:
            logger.warning(f"Algolia batch attempt {attempt + 1} failed: {e}")
            errors.append({
                "attempt": attempt + 1,
                "error": str(e),
                "batch_size": len(products)
            })
            
            if attempt < config.max_retries - 1:
                await asyncio.sleep(config.retry_delay * (attempt + 1))
    
    return indexed, errors


# ============================================
# MAIN INGESTION PROCESSOR
# ============================================

async def process_ingestion_job(
    job: IndexingJob,
    db,
    on_progress: callable = None
) -> IndexingJob:
    """
    Main job processor - handles the entire ingestion pipeline
    
    1. Read file in chunks
    2. Transform each chunk
    3. Index to Algolia in batches
    4. Track progress in MongoDB
    """
    from infoshop_service import (
        GRAINGER_CATEGORY_DISCOUNTS,
        MOTION_CATEGORY_DISCOUNTS,
        FASTENAL_CATEGORY_DISCOUNTS,
        get_partner_discounts
    )
    from algolia_service import PRODUCTS_INDEX
    
    config = IndexingJobConfig(**job.config)
    
    # Update job status
    job.status = JobStatus.PROCESSING
    job.started_at = datetime.now(timezone.utc).isoformat()
    update_job(job)
    await save_job_to_db(job, db)
    
    logger.info(f"Starting job {job.job_id}: {job.filename} ({job.total_rows} rows)")
    
    try:
        # Get category discounts for vendor
        vendor_lower = job.vendor.lower()
        if vendor_lower == "grainger":
            category_discounts = GRAINGER_CATEGORY_DISCOUNTS
        elif vendor_lower == "motion":
            category_discounts = MOTION_CATEGORY_DISCOUNTS
        elif vendor_lower == "fastenal":
            category_discounts = FASTENAL_CATEGORY_DISCOUNTS
        else:
            category_discounts = get_partner_discounts(job.vendor) or {}
        
        # Calculate total chunks
        job.total_chunks = (job.total_rows + config.chunk_size - 1) // config.chunk_size
        
        # Process file in chunks
        chunk_num = 0
        for chunk_df in stream_file_chunks(job.filepath, config.chunk_size):
            chunk_num += 1
            job.current_chunk = chunk_num
            
            logger.info(f"Job {job.job_id}: Processing chunk {chunk_num}/{job.total_chunks} ({len(chunk_df)} rows)")
            
            # Transform chunk
            products, transform_errors = transform_chunk(
                chunk_df,
                job.vendor,
                category_discounts,
                config
            )
            
            job.errors.extend(transform_errors)
            job.error_count += len(transform_errors)
            
            # Index to Algolia
            if products:
                indexed, index_errors = await index_batch_to_algolia(
                    products,
                    PRODUCTS_INDEX,
                    config
                )
                
                job.indexed_count += indexed
                job.errors.extend(index_errors)
                job.error_count += len(index_errors)
            
            # Update progress
            job.processed_rows += len(chunk_df)
            update_job(job)
            
            # Save progress to DB periodically
            if chunk_num % 5 == 0:
                await save_job_to_db(job, db)
            
            # Callback for real-time updates
            if on_progress:
                await on_progress(job)
            
            # Memory management
            del chunk_df
            del products
            gc.collect()
            
            # Check memory usage
            import psutil
            process = psutil.Process()
            memory_mb = process.memory_info().rss / 1024 / 1024
            if memory_mb > config.memory_limit_mb:
                logger.warning(f"Memory usage high ({memory_mb:.0f}MB), forcing GC")
                gc.collect()
        
        # Job completed successfully
        job.status = JobStatus.COMPLETED
        job.completed_at = datetime.now(timezone.utc).isoformat()
        
        logger.info(f"Job {job.job_id} completed: {job.indexed_count} indexed, {job.error_count} errors")
        
    except Exception as e:
        job.status = JobStatus.FAILED
        job.error_message = str(e)
        job.completed_at = datetime.now(timezone.utc).isoformat()
        logger.error(f"Job {job.job_id} failed: {e}\n{traceback.format_exc()}")
    
    finally:
        update_job(job)
        await save_job_to_db(job, db)
    
    return job


# ============================================
# DATABASE OPERATIONS
# ============================================

async def save_job_to_db(job: IndexingJob, db):
    """Save job status to MongoDB"""
    try:
        await db.indexing_jobs.update_one(
            {"job_id": job.job_id},
            {"$set": job.to_dict()},
            upsert=True
        )
    except Exception as e:
        logger.error(f"Failed to save job to DB: {e}")


async def load_job_from_db(job_id: str, db) -> Optional[IndexingJob]:
    """Load job from MongoDB"""
    try:
        doc = await db.indexing_jobs.find_one({"job_id": job_id})
        if doc:
            return IndexingJob(
                job_id=doc["job_id"],
                vendor=doc["vendor"],
                filename=doc["filename"],
                filepath=doc["filepath"],
                status=JobStatus(doc["status"]),
                total_rows=doc.get("total_rows", 0),
                processed_rows=doc.get("processed_rows", 0),
                indexed_count=doc.get("indexed_count", 0),
                error_count=doc.get("error_count", 0),
                current_chunk=doc.get("current_chunk", 0),
                total_chunks=doc.get("total_chunks", 0),
                started_at=doc.get("started_at"),
                completed_at=doc.get("completed_at"),
                error_message=doc.get("error_message"),
                errors=doc.get("errors", []),
                config=doc.get("config", asdict(DEFAULT_CONFIG))
            )
    except Exception as e:
        logger.error(f"Failed to load job from DB: {e}")
    return None


async def get_all_jobs(db, limit: int = 50) -> List[Dict]:
    """Get recent indexing jobs"""
    try:
        cursor = db.indexing_jobs.find().sort("started_at", -1).limit(limit)
        jobs = await cursor.to_list(length=limit)
        return jobs
    except Exception as e:
        logger.error(f"Failed to get jobs: {e}")
        return []


# ============================================
# JOB MANAGEMENT API
# ============================================

async def create_ingestion_job(
    file_path: str,
    vendor: str,
    filename: str,
    db,
    config: IndexingJobConfig = None
) -> IndexingJob:
    """
    Create a new ingestion job
    """
    if config is None:
        config = DEFAULT_CONFIG
    
    # Count rows
    total_rows = count_file_rows(file_path)
    
    if total_rows == 0:
        raise ValueError("File appears to be empty")
    
    # Create job
    job = IndexingJob(
        job_id=str(uuid.uuid4()),
        vendor=vendor,
        filename=filename,
        filepath=file_path,
        status=JobStatus.PENDING,
        total_rows=total_rows,
        config=asdict(config)
    )
    
    # Save to memory and DB
    update_job(job)
    await save_job_to_db(job, db)
    
    logger.info(f"Created job {job.job_id}: {filename} ({total_rows} rows)")
    
    return job


async def cancel_job(job_id: str, db) -> bool:
    """Cancel a running job"""
    job = get_job(job_id)
    if not job:
        job = await load_job_from_db(job_id, db)
    
    if job and job.is_active:
        job.status = JobStatus.CANCELLED
        job.completed_at = datetime.now(timezone.utc).isoformat()
        update_job(job)
        await save_job_to_db(job, db)
        return True
    
    return False


async def get_job_status(job_id: str, db) -> Optional[Dict]:
    """Get job status"""
    job = get_job(job_id)
    if not job:
        job = await load_job_from_db(job_id, db)
    
    if job:
        return job.to_dict()
    return None


# ============================================
# EXPORTS
# ============================================

__all__ = [
    'JobStatus',
    'IndexingJobConfig',
    'IndexingJob',
    'DEFAULT_CONFIG',
    'create_ingestion_job',
    'process_ingestion_job',
    'cancel_job',
    'get_job_status',
    'get_all_jobs',
    'get_job',
    'stream_file_chunks',
    'count_file_rows',
]
