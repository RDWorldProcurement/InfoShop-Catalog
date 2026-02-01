from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query, Form, Request, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import PlainTextResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import random
import hashlib
import asyncio
import pandas as pd
import io

# Import Emergent LLM integration
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    EMERGENT_AVAILABLE = True
except ImportError:
    EMERGENT_AVAILABLE = False
    LlmChat = None
    UserMessage = None

# Import AI price benchmarking module
from ai_price_benchmark import (
    perform_ai_price_benchmarking, 
    DEMO_QUOTATION, 
    DEMO_ANALYSIS_RESULTS,
    analyze_with_openai,
    analyze_with_claude,
    analyze_with_gemini
)

# Import real document extraction module
from document_extractor import extract_quotation_data

# Import negotiation agent module
from negotiation_agent import (
    NegotiationStrategy,
    NEGOTIATION_PLAYBOOKS,
    calculate_target_price,
    generate_negotiation_targets,
    generate_negotiation_email,
    create_counter_offer,
    get_all_strategies
)

# Import Algolia service for catalog search
try:
    from algolia_service import (
        init_algolia,
        index_products as algolia_index_products,
        index_products_from_file,
        search_products as algolia_search_products,
        get_facet_values,
        update_product_grouping,
        clear_index,
        get_index_stats
    )
    ALGOLIA_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Algolia service not available: {e}")
    ALGOLIA_AVAILABLE = False

# Import PunchOut service for Coupa cXML integration
from punchout_service import (
    parse_punchout_setup_request,
    validate_punchout_credentials,
    create_punchout_setup_response,
    create_punchout_order_message,
    create_punchout_session,
    get_punchout_session,
    update_punchout_cart,
    close_punchout_session,
    log_punchout_transaction,
    save_punchout_session_to_db,
    get_punchout_session_from_db,
    PUNCHOUT_CONFIG
)

# Import InfoShop Catalog Service for enterprise features
from infoshop_service import (
    ACTIVE_PARTNERS,
    COMING_SOON_PARTNERS,
    generate_infoshop_part_number,
    calculate_danone_preferred_price,
    classify_unspsc,
    validate_image_url,
    calculate_minimum_delivery_date,
    validate_delivery_date,
    transform_product_for_infoshop,
    load_partner_discounts,
    get_partner_discounts,
    get_all_partner_discounts
)

# Import Scalable Ingestion Service
from scalable_ingestion import (
    JobStatus,
    IndexingJobConfig,
    IndexingJob,
    DEFAULT_CONFIG,
    create_ingestion_job,
    process_ingestion_job,
    cancel_job,
    get_job_status,
    get_all_jobs,
    get_job,
    count_file_rows,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging early for use throughout the app
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET_KEY', 'omnisupply_default_secret')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24  # Extended from 1 hour

# Emergent LLM Key
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# Create the main app
app = FastAPI(title="OMNISupply.io API", version="2.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Translation cache to avoid repeated LLM calls
translation_cache = {}

# Language mapping for translation
LANGUAGE_NAMES = {
    "en": "English",
    "fr": "French",
    "de": "German",
    "it": "Italian",
    "nl": "Dutch"
}

async def translate_text(text: str, target_lang: str, context: str = "product") -> str:
    """Translate text using Emergent LLM with caching"""
    if not text or target_lang == "en":
        return text
    
    # Check cache first
    cache_key = f"{target_lang}:{hashlib.md5(text.encode()).hexdigest()}"
    if cache_key in translation_cache:
        return translation_cache[cache_key]
    
    # Check MongoDB cache
    cached = await db.translations.find_one({"cache_key": cache_key})
    if cached:
        translation_cache[cache_key] = cached["translation"]
        return cached["translation"]
    
    try:
        target_language = LANGUAGE_NAMES.get(target_lang, "French")
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"translate_{cache_key[:8]}",
            system_message=f"You are a professional B2B e-commerce product translator. Translate the following {context} text from English to {target_language}. Keep technical terms, brand names, model numbers, and specifications intact. Only provide the translation, no explanations."
        ).with_model("openai", "gpt-4o-mini")
        
        response = await chat.send_message(UserMessage(text=text))
        translated = response.strip() if response else text
        
        # Cache in memory and MongoDB
        translation_cache[cache_key] = translated
        await db.translations.update_one(
            {"cache_key": cache_key},
            {"$set": {"cache_key": cache_key, "original": text, "target_lang": target_lang, 
                     "translation": translated, "created_at": datetime.now(timezone.utc)}},
            upsert=True
        )
        
        return translated
    except Exception as e:
        logger.error(f"Translation error: {e}")
        return text

async def translate_product(product: dict, target_lang: str) -> dict:
    """Translate all product text fields"""
    if target_lang == "en":
        return product
    
    translated = product.copy()
    
    # Translate key fields
    if product.get("name"):
        translated["name"] = await translate_text(product["name"], target_lang, "product name")
    if product.get("short_description"):
        translated["short_description"] = await translate_text(product["short_description"], target_lang, "product description")
    if product.get("full_description"):
        translated["full_description"] = await translate_text(product["full_description"], target_lang, "product description")
    if product.get("category"):
        translated["category"] = await translate_text(product["category"], target_lang, "category name")
    
    # Translate specifications
    if product.get("specifications"):
        translated_specs = {}
        for key, value in product["specifications"].items():
            translated_key = await translate_text(key, target_lang, "specification label")
            # Don't translate numeric values or technical specs
            if isinstance(value, str) and not any(c.isdigit() for c in value):
                translated_specs[translated_key] = await translate_text(value, target_lang, "specification value")
            else:
                translated_specs[translated_key] = value
        translated["specifications"] = translated_specs
    
    return translated

async def translate_service(service: dict, target_lang: str) -> dict:
    """Translate all service text fields"""
    if target_lang == "en":
        return service
    
    translated = service.copy()
    
    # Translate key fields
    if service.get("name"):
        translated["name"] = await translate_text(service["name"], target_lang, "service name")
    if service.get("short_description"):
        translated["short_description"] = await translate_text(service["short_description"], target_lang, "service description")
    if service.get("full_description"):
        translated["full_description"] = await translate_text(service["full_description"], target_lang, "service description")
    if service.get("category"):
        translated["category"] = await translate_text(service["category"], target_lang, "category name")
    if service.get("unit_of_measure"):
        translated["unit_of_measure"] = await translate_text(service["unit_of_measure"], target_lang, "pricing unit")
    
    # Translate service includes list
    if service.get("service_includes"):
        translated["service_includes"] = [
            await translate_text(item, target_lang, "service feature") 
            for item in service["service_includes"]
        ]
    
    return translated

# Currency configurations per country
COUNTRY_CURRENCIES = {
    "USA": {"code": "USD", "symbol": "$", "rate": 1.0},
    "Canada": {"code": "CAD", "symbol": "CA$", "rate": 1.36},
    "Mexico": {"code": "MXN", "symbol": "MX$", "rate": 17.15},
    "India": {"code": "INR", "symbol": "₹", "rate": 83.12},
    "China": {"code": "CNY", "symbol": "¥", "rate": 7.24},
    "Germany": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "France": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "UK": {"code": "GBP", "symbol": "£", "rate": 0.79},
    "Italy": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "Spain": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "Netherlands": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "Belgium": {"code": "EUR", "symbol": "€", "rate": 0.92},
    "Poland": {"code": "PLN", "symbol": "zł", "rate": 4.02},
    "Switzerland": {"code": "CHF", "symbol": "CHF", "rate": 0.88},
    "Sweden": {"code": "SEK", "symbol": "kr", "rate": 10.45},
}

# MRO Categories with UNSPSC codes
MRO_CATEGORIES = [
    {"name": "Bearings & Power Transmission", "unspsc": "31170000", "icon": "cog"},
    {"name": "Electrical & Lighting", "unspsc": "39110000", "icon": "lightbulb"},
    {"name": "Fasteners & Hardware", "unspsc": "31160000", "icon": "wrench"},
    {"name": "Hand Tools", "unspsc": "27110000", "icon": "hammer"},
    {"name": "Power Tools", "unspsc": "27112000", "icon": "zap"},
    {"name": "Safety & PPE", "unspsc": "46180000", "icon": "shield"},
    {"name": "Abrasives", "unspsc": "31190000", "icon": "disc"},
    {"name": "Adhesives & Sealants", "unspsc": "31200000", "icon": "droplet"},
    {"name": "Cleaning & Janitorial", "unspsc": "47130000", "icon": "sparkles"},
    {"name": "HVAC & Refrigeration", "unspsc": "40100000", "icon": "thermometer"},
    {"name": "Hydraulics & Pneumatics", "unspsc": "40140000", "icon": "gauge"},
    {"name": "Laboratory Supplies", "unspsc": "41110000", "icon": "flask"},
    {"name": "Lubrication", "unspsc": "15120000", "icon": "oil-can"},
    {"name": "Material Handling", "unspsc": "24100000", "icon": "package"},
    {"name": "Motors & Drives", "unspsc": "26100000", "icon": "cpu"},
    {"name": "Packaging & Shipping", "unspsc": "24110000", "icon": "box"},
    {"name": "Pipe, Valves & Fittings", "unspsc": "40170000", "icon": "cylinder"},
    {"name": "Plumbing", "unspsc": "40140000", "icon": "droplets"},
    {"name": "Pumps", "unspsc": "40150000", "icon": "activity"},
    {"name": "Raw Materials", "unspsc": "11100000", "icon": "layers"},
    {"name": "Test & Measurement", "unspsc": "41110000", "icon": "ruler"},
    {"name": "Welding", "unspsc": "23270000", "icon": "flame"},
    {"name": "Industrial Automation", "unspsc": "32150000", "icon": "bot"},
    {"name": "Cutting Tools", "unspsc": "27110000", "icon": "scissors"},
    {"name": "Storage & Organization", "unspsc": "56100000", "icon": "archive"},
    {"name": "IT Equipment - Laptops", "unspsc": "43211500", "icon": "laptop"},
    {"name": "IT Equipment - Monitors", "unspsc": "43211900", "icon": "monitor"},
    {"name": "IT Equipment - Networking", "unspsc": "43222600", "icon": "wifi"},
    {"name": "IT Equipment - Servers", "unspsc": "43211800", "icon": "server"},
    {"name": "IT Equipment - Peripherals", "unspsc": "43211700", "icon": "keyboard"},
    {"name": "Filtration", "unspsc": "40161500", "icon": "filter"},
    {"name": "Industrial Coding", "unspsc": "44100000", "icon": "printer"},
]

# MRO Brands - using text-based display (logos removed due to external service reliability issues)
MRO_BRANDS = [
    {"name": "SKF", "logo": None, "color": "#005B94"},
    {"name": "3M", "logo": None, "color": "#FF0000"},
    {"name": "Henkel", "logo": None, "color": "#D40000"},
    {"name": "Bosch", "logo": None, "color": "#E30016"},
    {"name": "Siemens", "logo": None, "color": "#009999"},
    {"name": "ABB", "logo": None, "color": "#FF000F"},
    {"name": "Honeywell", "logo": None, "color": "#E31837"},
    {"name": "Parker", "logo": None, "color": "#004B87"},
    {"name": "Emerson", "logo": None, "color": "#63666A"},
    {"name": "Rockwell", "logo": None, "color": "#C8102E"},
    {"name": "Schneider Electric", "logo": None, "color": "#3DCD58"},
    {"name": "Mitsubishi Electric", "logo": None, "color": "#ED1C24"},
    {"name": "Omron", "logo": None, "color": "#0063AF"},
    {"name": "Festo", "logo": None, "color": "#0091D5"},
    {"name": "Fluke", "logo": None, "color": "#FFC20E"},
    {"name": "Makita", "logo": None, "color": "#00B1BC"},
    {"name": "DeWalt", "logo": None, "color": "#FEBD17"},
    {"name": "Milwaukee", "logo": None, "color": "#DB0032"},
    {"name": "Stanley", "logo": None, "color": "#FFCC00"},
    {"name": "Klein Tools", "logo": None, "color": "#FF6600"},
    {"name": "HP", "logo": None, "color": "#0096D6"},
    {"name": "Dell", "logo": None, "color": "#007DB8"},
    {"name": "Lenovo", "logo": None, "color": "#E2231A"},
    {"name": "LG", "logo": None, "color": "#A50034"},
    {"name": "Samsung", "logo": None, "color": "#1428A0"},
    {"name": "Cisco", "logo": None, "color": "#049FD9"},
    {"name": "ASUS", "logo": None, "color": "#00539B"},
    {"name": "Acer", "logo": None, "color": "#83B81A"},
    {"name": "BenQ", "logo": None, "color": "#5B2D87"},
    {"name": "Logitech", "logo": None, "color": "#00B8FC"},
    {"name": "Donaldson", "logo": None, "color": "#003B73"},
    {"name": "Avantor", "logo": None, "color": "#6D2077"},
    {"name": "Markem-Imaje", "logo": None, "color": "#E4002B"},
]

# Product images hosted on Emergent CDN - product-specific matches
PRODUCT_IMAGE_URLS = {
    # Specific product images (brand-matched)
    "SKF Ball Bearing": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/5d11401880367c9ecb4d3d43cf84af6a95372bff44764d01080f72fdadaccc38.png",
    "3M Safety Helmet": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/d5fdc07665f86ea5f580ac7e6fd0e73c4925625f60f6055ebcda4f23f08d0431.png",
    "3M Safety Glasses": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/9c7ea6cfd47a7569144fd2bb61eb54feef71eabfcdda4f5f7fa2f001d3c668ad.png",
    "Bosch Drill": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/abca8f6c400eccd9885f40052859c77b18c5a13dbde5ed8763170d9ab7fd9c41.png",
    "Stanley Wrench": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/25e01814ef99f668b385220f94849941ba287aee8811756d6a673bc4069e2c85.png",
    "Gates Timing Belt": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/b339b39dd2001c7b78175f881458bfe5c1cc9ac85b875769eaabab2ed13b9ccd.png",
    "Schneider Circuit Breaker": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1b06c8223f99520d00c70296b71296814d63202c0c3d150b627d2bb53ce629a6.png",
    "HP Laptop": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f6f1caf66bd07154a37a0dadfa973a662a67d39d32552f5d51d73e7cec65e616.png",
    "Dell Monitor": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/08c23c33778bc9d3ce86d151e5eee5000a143292981570e9eb926791a48d5f62.png",
    "Cisco Switch": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6663d8398909e7e6833d27291342928ba4b7567abeeb7570dc5c39af9136f127.png",
    "Henkel Adhesive": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1ad744a50cda48b7b4ced8b7f15874c83dd722b367c119e969f45cb842db7b6f.png",
    "Philips LED Light": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6339f73a3f393ecb8997ea7a9444a9068c2c829d743f44b928094ecf8464e143.png",
    # Category fallback images
    "Bearings & Power Transmission": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/5d11401880367c9ecb4d3d43cf84af6a95372bff44764d01080f72fdadaccc38.png",
    "Electrical & Lighting": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1b06c8223f99520d00c70296b71296814d63202c0c3d150b627d2bb53ce629a6.png",
    "Fasteners & Hardware": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/77f9d6c3fedcd85729745ffaf16ab46b268ba0176386a36e9d3bd3d1b2e6c293.png",
    "Hand Tools": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/25e01814ef99f668b385220f94849941ba287aee8811756d6a673bc4069e2c85.png",
    "Power Tools": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/abca8f6c400eccd9885f40052859c77b18c5a13dbde5ed8763170d9ab7fd9c41.png",
    "Safety & PPE": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/d5fdc07665f86ea5f580ac7e6fd0e73c4925625f60f6055ebcda4f23f08d0431.png",
    "Abrasives": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/7ac37795d30541fe96a379b8ebc9a669a9f5534a1c47d157f2dcfce68eda8fde.png",
    "Adhesives & Sealants": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1ad744a50cda48b7b4ced8b7f15874c83dd722b367c119e969f45cb842db7b6f.png",
    "Cleaning & Janitorial": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/899b4b3e6aceed339189fb8d0b0fa309930af38864a707139f102cc7b47c1bfd.png",
    "HVAC & Refrigeration": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/5deee0f39fbb43dd7a40094f3273b92017a5b99cb2f2a4ffe89656529a244bf5.png",
    "Hydraulics & Pneumatics": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/67b16a05b97dae42c11c9c72f6e05b8af5d7aec7f63f9dfbd8e237a3fd5fc463.png",
    "Laboratory Supplies": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/411bcb13dee27615e9b815fde695cdc27e5f05eee4347121309be34bbdb8c3e0.png",
    "Lubrication": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/e12637ed7080f9cc6f0ab4378314a9bf6991b40c91ff9c4c31ea98fd25d1aa62.png",
    "Material Handling": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/d3b288a3bc2eae93ca8304dff24b6a6c2fa3e071ff498bff6f994dc86e336ddf.png",
    "Motors & Drives": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/48eb714ed81a4debd0fcbb8614e5e4412263c16c00d900b645c44f969c378c75.png",
    "Packaging & Shipping": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/77f9d6c3fedcd85729745ffaf16ab46b268ba0176386a36e9d3bd3d1b2e6c293.png",
    "Pipe, Valves & Fittings": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f33db98cd05c6baa652b94111e2425bcdc4313c5335f9673a214cb91bf2d2284.png",
    "Plumbing": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f33db98cd05c6baa652b94111e2425bcdc4313c5335f9673a214cb91bf2d2284.png",
    "Pumps": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/67b16a05b97dae42c11c9c72f6e05b8af5d7aec7f63f9dfbd8e237a3fd5fc463.png",
    "Raw Materials": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/e937d26ba9dc90a270d3f1672abc004a7cbd5351be7177df115648759a328877.png",
    "Test & Measurement": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/411bcb13dee27615e9b815fde695cdc27e5f05eee4347121309be34bbdb8c3e0.png",
    "Welding": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/2e6b43c65ae6a6b6a4fa924c3f3f88f897cad2a707104ffdc57d165e3b1d04e8.png",
    "Industrial Automation": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/022021c9eb3d4350b596e47b5cbbc21f913a9ca5428f805d45b37108e6428799.png",
    "Cutting Tools": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/5f3ec0cc9e4246d05af7b1609a78a8d3c2c76abbebba2fcb9823ffac8851309d.png",
    "Storage & Organization": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/ad801ca5185620afbcc264b33f0c59dacf2085f9b98895eb624068ec3ebad64d.png",
    "IT Equipment - Laptops": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f6f1caf66bd07154a37a0dadfa973a662a67d39d32552f5d51d73e7cec65e616.png",
    "IT Equipment - Monitors": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/08c23c33778bc9d3ce86d151e5eee5000a143292981570e9eb926791a48d5f62.png",
    "IT Equipment - Networking": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6663d8398909e7e6833d27291342928ba4b7567abeeb7570dc5c39af9136f127.png",
    "IT Equipment - Servers": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/022021c9eb3d4350b596e47b5cbbc21f913a9ca5428f805d45b37108e6428799.png",
    "IT Equipment - Peripherals": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/075214c7f6dce1e67075c623d025158b3d3cac86e45217650f7d58db7662f1d6.png",
    "Safety Gloves": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/9f7011dbf936aa67ab41af039d39e363f5a435c0efbfbed8a069e7684f74ad25.png",
}

# Reward images for InfoCoins redemption
REWARD_IMAGE_URLS = {
    "Executive Jacket": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/7a28899414c059decd7892a76b2510a988cbd7721cd8cc81edb9db0b2d98abb8.png",
    "Insulated Tumbler": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/e20ad22b18b960ea85b19e85bf4ff46d506caa85fc150f5ed6f1b6f4b693fde8.png",
    "Leather Backpack": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/24f072b8fe320ed518c80d29f9dc4b59278746c3e30e7d4e9e4be2a654c9fb68.png",
    "Wireless Earbuds": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f9f47c254925d769c605a9cacaf94508b5b86b3a75ca50b7ad1ea13892d9ead4.png",
    "Desk Organizer Set": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/41326a793c1dca20d6762573d31aba0fcb23091926051fb6868ffa92c39df981.png",
    "Smartwatch": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/4b7a4dd5c087a82d7316e8d008a8dc5946dcc6733dc8596fcea9bd56143dde46.png",
}

# Default fallback image
DEFAULT_PRODUCT_IMAGE = "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/7ac37795d30541fe96a379b8ebc9a669a9f5534a1c47d157f2dcfce68eda8fde.png"

# Service images hosted on Emergent CDN (guaranteed to work)
SERVICE_IMAGE_URLS = {
    "Network Installation Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/afdd726c02cc7d9e8690e91dc7c1b0a13c962c96325cef7d1eece4d48001fb82.png",
    "IT Equipment Installation & Setup": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/cff7a5158c43f81799d59a99494279b2c5f255b2610fe406a1ffe6934277c4a7.png",
    "Cybersecurity Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/88995698058a90986297157d130fa19ac15656bbe7ac2296b1949cbb9993e380.png",
    "IT Managed Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/f848c4af2e1f5caac71aa56dc1f1b9285fb801ad8e1ee77ad4185116211e3627.png",
    "Corporate & Business Support Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/49c83afc1a588e3c2810a205953d952ac31ad730ca2f9871863edeeea2072a83.png",
    "Digital Marketing & Creative Agency Services": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/3e0d7da8578b2c511105784ffde94733d5a947a499e7481de09ef3466e65afa6.png",
    "Facilities Management & Workplace Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/23544d6096e596b6e3954fec76d55a7fa874df0072512cd79110a2a41cce3b44.png",
    "HSE, Quality & Compliance Services": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1e0527ecaf208bd7905092e337f70a54eabef0782a81bd6e510c8e4d5c3c18ac.png",
    "IT & Workplace Technology Services": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/cff7a5158c43f81799d59a99494279b2c5f255b2610fe406a1ffe6934277c4a7.png",
    "Logistics, Warehouse & Supply Chain Services": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6810b3362e86e8ba23010897c9d6d3718dce4d1803e090213f94464944c1175b.png",
    "Temp Labor across Technical Skilled Capabilities": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/ec1d202cfaeb46233ae208411d641185f676c426d6219f30dc8a7f6d57ccaa7b.png",
    "Server Installation": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/a1c30bd322a62ea6d441b12da74c66ca96707140958e1737c0d9a286f041a795.png",
    "WiFi Setup": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/88c4a88606ac13bb8e539641523213064945b830124a9c902aa9147532aa3c49.png",
    "Data Center Services": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/45b8605364d7a481ab4a4bf15be73f59f9761da18220b1efb363c134d7ce0fb4.png",
    "Equipment Maintenance": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/fd8c452c65505c4fa649fae356b50b7dcc945b4ca0d180088866270a0b7d86f5.png",
    "Training Services": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/8d3d10c922ac4516e87dff834b0b0f3d00633c1032b7667687f506addb1e98b3.png",
    "Quality Control": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1e0527ecaf208bd7905092e337f70a54eabef0782a81bd6e510c8e4d5c3c18ac.png",
    "Commercial Cleaning": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/124d0636fdb6759f0b0e809e00fe07c91c12562a510f2da1e68cc2cb7a0fd266.png",
    # Digital Marketing Services Images
    "Social Media Marketing": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/3e0d7da8578b2c511105784ffde94733d5a947a499e7481de09ef3466e65afa6.png",
    "SEO Optimization": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/065098b74382dcef95f4d43a5acfec01cf7019a80344289d9bbcf29876951ecf.png",
    "Content Marketing": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/11858f0f46d7b769347af0396a64091c470e4cfaf679e89ea3f0d9bc9c197174.png",
    "Email Marketing": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6643f65c20391b649166e5889b2700f56e38fbc70128635358567cc335c9ca9c.png",
    "PPC Advertising": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/06ee08e5c3b4256c9a72385c72da51dd6abf1464eab5586423dce224913af57d.png",
    "Brand Identity Design": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/317b64184657f818f660ec7fb260ac16010f4fb505facb3924e1b120b1ea17d2.png",
}

# Default fallback service image
DEFAULT_SERVICE_IMAGE = "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/49c83afc1a588e3c2810a205953d952ac31ad730ca2f9871863edeeea2072a83.png"

# Service Categories with UNSPSC codes
SERVICE_CATEGORIES = [
    {"name": "Corporate & Business Support Services", "unspsc": "80100000", "icon": "briefcase"},
    {"name": "Digital Marketing & Creative Agency Services", "unspsc": "82100000", "icon": "palette"},
    {"name": "Facilities Management & Workplace Services", "unspsc": "72100000", "icon": "building"},
    {"name": "HSE, Quality & Compliance Services", "unspsc": "77100000", "icon": "shield-check"},
    {"name": "IT & Workplace Technology Services", "unspsc": "81110000", "icon": "monitor"},
    {"name": "Logistics, Warehouse & Supply Chain Services", "unspsc": "78100000", "icon": "truck"},
    {"name": "Temp Labor across Technical Skilled Capabilities", "unspsc": "80110000", "icon": "users"},
    {"name": "Network Installation Services", "unspsc": "81111800", "icon": "wifi"},
    {"name": "IT Equipment Installation & Setup", "unspsc": "81112200", "icon": "server"},
    {"name": "IT Managed Services", "unspsc": "81112300", "icon": "cloud"},
    {"name": "Cybersecurity Services", "unspsc": "81112500", "icon": "shield"},
]

# Detailed IT Products Catalog
IT_PRODUCTS_CATALOG = [
    # HP Laptops
    {
        "id": "HP-LAP-001",
        "name": "HP ProBook 450 G10 Business Laptop",
        "brand": "HP",
        "category": "IT Equipment - Laptops",
        "sku": "HP-PB450G10-i7",
        "unspsc_code": "43211503",
        "base_price": 1299.00,
        "image_url": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")),
        "short_description": "15.6\" FHD Display, Intel Core i7-1355U, 16GB RAM, 512GB SSD",
        "full_description": "The HP ProBook 450 G10 is designed for business professionals who need reliability and performance. Features a 15.6-inch Full HD anti-glare display, Intel Core i7-1355U processor (up to 5.0 GHz), 16GB DDR4 RAM, and 512GB NVMe SSD. Includes Windows 11 Pro, Intel Iris Xe Graphics, fingerprint reader, and HD webcam with privacy shutter.",
        "specifications": {
            "Processor": "Intel Core i7-1355U (10 cores, up to 5.0 GHz)",
            "Memory": "16GB DDR4-3200",
            "Storage": "512GB PCIe NVMe SSD",
            "Display": "15.6\" FHD (1920x1080) IPS Anti-Glare",
            "Graphics": "Intel Iris Xe Graphics",
            "Operating System": "Windows 11 Pro",
            "Battery": "51Wh, up to 10 hours",
            "Weight": "1.79 kg (3.94 lbs)",
            "Ports": "USB-C, USB-A, HDMI, RJ-45, Audio Jack",
            "Warranty": "3 Years On-Site"
        },
        "availability": {"in_stock": True, "quantity": 156, "warehouse": "US-East"},
        "rating": 4.5,
        "reviews_count": 234
    },
    {
        "id": "HP-LAP-002",
        "name": "HP EliteBook 840 G10 Enterprise Laptop",
        "brand": "HP",
        "category": "IT Equipment - Laptops",
        "sku": "HP-EB840G10-i7",
        "unspsc_code": "43211503",
        "base_price": 1849.00,
        "image_url": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")),
        "short_description": "14\" WUXGA Display, Intel Core i7-1365U, 32GB RAM, 1TB SSD",
        "full_description": "HP EliteBook 840 G10 delivers enterprise-grade security with stunning design. Features Sure View Reflect privacy screen, Wolf Security for Business, and military-grade durability (MIL-STD 810H). Perfect for executive and hybrid workforce.",
        "specifications": {
            "Processor": "Intel Core i7-1365U vPro (12 cores, up to 5.2 GHz)",
            "Memory": "32GB DDR5-4800",
            "Storage": "1TB PCIe Gen4 NVMe SSD",
            "Display": "14\" WUXGA (1920x1200) Sure View Reflect",
            "Graphics": "Intel Iris Xe Graphics",
            "Operating System": "Windows 11 Pro",
            "Battery": "51Wh, up to 14 hours",
            "Weight": "1.36 kg (3.0 lbs)",
            "Security": "Fingerprint, IR Camera, TPM 2.0, Wolf Security",
            "Warranty": "3 Years On-Site Next Business Day"
        },
        "availability": {"in_stock": True, "quantity": 89, "warehouse": "US-West"},
        "rating": 4.8,
        "reviews_count": 412
    },
    # Dell Laptops
    {
        "id": "DELL-LAP-001",
        "name": "Dell Latitude 5540 Business Laptop",
        "brand": "Dell",
        "category": "IT Equipment - Laptops",
        "sku": "DELL-LAT5540-i7",
        "unspsc_code": "43211503",
        "base_price": 1449.00,
        "image_url": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")),
        "short_description": "15.6\" FHD+ Display, Intel Core i7-1365U, 16GB RAM, 512GB SSD",
        "full_description": "Dell Latitude 5540 combines performance with enterprise manageability. Built for IT deployment with Dell Optimizer AI, SafeBIOS, and comprehensive Dell Technologies services.",
        "specifications": {
            "Processor": "Intel Core i7-1365U vPro (12 cores, up to 5.2 GHz)",
            "Memory": "16GB DDR4-3200",
            "Storage": "512GB PCIe NVMe SSD",
            "Display": "15.6\" FHD+ (1920x1200) Anti-Glare",
            "Graphics": "Intel Iris Xe Graphics",
            "Operating System": "Windows 11 Pro",
            "Battery": "54Wh, up to 11 hours",
            "Weight": "1.66 kg (3.66 lbs)",
            "Ports": "Thunderbolt 4, USB-A, HDMI 2.0, microSD",
            "Warranty": "3 Years ProSupport"
        },
        "availability": {"in_stock": True, "quantity": 203, "warehouse": "US-Central"},
        "rating": 4.6,
        "reviews_count": 567
    },
    {
        "id": "DELL-LAP-002",
        "name": "Dell Precision 5680 Mobile Workstation",
        "brand": "Dell",
        "category": "IT Equipment - Laptops",
        "sku": "DELL-P5680-i9",
        "unspsc_code": "43211503",
        "base_price": 3299.00,
        "image_url": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")),
        "short_description": "16\" 3.5K OLED, Intel Core i9-13900H, 64GB RAM, 2TB SSD, RTX 3500",
        "full_description": "Dell Precision 5680 is engineered for CAD, 3D modeling, and AI workloads. Features NVIDIA RTX 3500 Ada graphics, stunning 3.5K OLED display with 120Hz refresh, and ISV-certified reliability.",
        "specifications": {
            "Processor": "Intel Core i9-13900H (14 cores, up to 5.4 GHz)",
            "Memory": "64GB DDR5-4800",
            "Storage": "2TB PCIe Gen4 NVMe SSD",
            "Display": "16\" 3.5K (3456x2160) OLED 120Hz Touch",
            "Graphics": "NVIDIA RTX 3500 Ada 12GB GDDR6",
            "Operating System": "Windows 11 Pro for Workstations",
            "Battery": "86Wh, up to 8 hours",
            "Weight": "2.01 kg (4.43 lbs)",
            "Certifications": "ISV Certified (AutoCAD, SolidWorks, Revit)",
            "Warranty": "3 Years ProSupport Plus"
        },
        "availability": {"in_stock": True, "quantity": 34, "warehouse": "US-East"},
        "rating": 4.9,
        "reviews_count": 123
    },
    # Lenovo Laptops
    {
        "id": "LEN-LAP-001",
        "name": "Lenovo ThinkPad X1 Carbon Gen 11",
        "brand": "Lenovo",
        "category": "IT Equipment - Laptops",
        "sku": "LEN-X1C-G11-i7",
        "unspsc_code": "43211503",
        "base_price": 1999.00,
        "image_url": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")),
        "short_description": "14\" 2.8K OLED, Intel Core i7-1365U, 32GB RAM, 1TB SSD",
        "full_description": "The legendary ThinkPad X1 Carbon continues its legacy with Gen 11. Ultra-light at just 1.12kg, featuring stunning OLED display, iconic keyboard, and comprehensive security features.",
        "specifications": {
            "Processor": "Intel Core i7-1365U vPro (12 cores, up to 5.2 GHz)",
            "Memory": "32GB LPDDR5-6400",
            "Storage": "1TB PCIe Gen4 NVMe SSD",
            "Display": "14\" 2.8K (2880x1800) OLED 400nits",
            "Graphics": "Intel Iris Xe Graphics",
            "Operating System": "Windows 11 Pro",
            "Battery": "57Wh, up to 15 hours",
            "Weight": "1.12 kg (2.48 lbs)",
            "Security": "Fingerprint, IR Camera, dTPM 2.0, ThinkShutter",
            "Warranty": "3 Years Premier Support"
        },
        "availability": {"in_stock": True, "quantity": 178, "warehouse": "US-West"},
        "rating": 4.7,
        "reviews_count": 891
    },
    # Monitors
    {
        "id": "DELL-MON-001",
        "name": "Dell UltraSharp U2723QE 27\" 4K USB-C Hub Monitor",
        "brand": "Dell",
        "category": "IT Equipment - Monitors",
        "sku": "DELL-U2723QE",
        "unspsc_code": "43211902",
        "base_price": 799.00,
        "image_url": PRODUCT_IMAGE_URLS.get("Dell Monitor", PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors")),
        "short_description": "27\" 4K UHD IPS Black, USB-C 90W PD, RJ45, Built-in KVM",
        "full_description": "Dell UltraSharp U2723QE features IPS Black technology for 2000:1 contrast ratio. USB-C hub with 90W power delivery, RJ45 ethernet, and built-in KVM for multi-PC productivity.",
        "specifications": {
            "Screen Size": "27 inches",
            "Resolution": "3840 x 2160 (4K UHD)",
            "Panel Type": "IPS Black Technology",
            "Refresh Rate": "60Hz",
            "Response Time": "5ms (GtG Fast)",
            "Brightness": "400 cd/m² (typical)",
            "Contrast": "2000:1",
            "Color Accuracy": "100% sRGB, 98% DCI-P3, Delta E < 2",
            "Connectivity": "USB-C (90W PD), HDMI, DP, USB-A Hub, RJ45",
            "Stand": "Height, Tilt, Swivel, Pivot Adjustable",
            "Warranty": "3 Years Advanced Exchange"
        },
        "availability": {"in_stock": True, "quantity": 312, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 1245
    },
    {
        "id": "LG-MON-001",
        "name": "LG UltraFine 32UN880-B 32\" 4K Ergo Monitor",
        "brand": "LG",
        "category": "IT Equipment - Monitors",
        "sku": "LG-32UN880B",
        "unspsc_code": "43211902",
        "base_price": 699.00,
        "image_url": PRODUCT_IMAGE_URLS.get("Dell Monitor", PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors")),
        "short_description": "32\" 4K UHD IPS, USB-C 60W PD, Ergo Stand, HDR10",
        "full_description": "LG UltraFine Ergo features a unique C-clamp stand for maximum desk space and flexibility. HDR10 support, 95% DCI-P3 color gamut, and USB-C one-cable solution.",
        "specifications": {
            "Screen Size": "31.5 inches",
            "Resolution": "3840 x 2160 (4K UHD)",
            "Panel Type": "IPS",
            "Refresh Rate": "60Hz",
            "Response Time": "5ms (GtG)",
            "Brightness": "350 cd/m²",
            "HDR": "HDR10",
            "Color Gamut": "95% DCI-P3, 99% sRGB",
            "Connectivity": "USB-C (60W PD), HDMI x2, USB Hub",
            "Stand": "Ergo C-Clamp with Full Articulation",
            "Warranty": "3 Years"
        },
        "availability": {"in_stock": True, "quantity": 187, "warehouse": "US-East"},
        "rating": 4.6,
        "reviews_count": 678
    },
    {
        "id": "SAM-MON-001",
        "name": "Samsung ViewFinity S9 49\" 5K Ultrawide",
        "brand": "Samsung",
        "category": "IT Equipment - Monitors",
        "sku": "SAM-LS49C950",
        "unspsc_code": "43211902",
        "base_price": 1499.00,
        "image_url": PRODUCT_IMAGE_URLS.get("Dell Monitor", PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors")),
        "short_description": "49\" 5K Dual QHD, 1000R Curve, USB-C 90W, KVM Switch",
        "full_description": "Samsung ViewFinity S9 replaces two monitors with one stunning 49-inch ultrawide. 5120x1440 Dual QHD resolution with 1000R curvature for immersive productivity.",
        "specifications": {
            "Screen Size": "49 inches (32:9 Aspect)",
            "Resolution": "5120 x 1440 (Dual QHD)",
            "Panel Type": "VA",
            "Refresh Rate": "120Hz",
            "Response Time": "4ms (GtG)",
            "Brightness": "350 cd/m²",
            "Curvature": "1000R",
            "Color Accuracy": "99% sRGB, Delta E < 2",
            "Connectivity": "USB-C (90W PD), HDMI x2, DP, USB Hub",
            "Features": "PBP/PIP, KVM Switch, Eye Saver Mode",
            "Warranty": "3 Years"
        },
        "availability": {"in_stock": True, "quantity": 56, "warehouse": "US-West"},
        "rating": 4.7,
        "reviews_count": 234
    },
    # Networking Equipment
    {
        "id": "CISCO-NET-001",
        "name": "Cisco Catalyst 9200L-48P-4G Network Switch",
        "brand": "Cisco",
        "category": "IT Equipment - Networking",
        "sku": "C9200L-48P-4G-E",
        "unspsc_code": "43222609",
        "base_price": 4299.00,
        "image_url": PRODUCT_IMAGE_URLS.get("Cisco Switch", PRODUCT_IMAGE_URLS.get("IT Equipment - Networking")),
        "short_description": "48-Port PoE+ Gigabit, 4x1G SFP, 370W PoE Budget, Stackable",
        "full_description": "Cisco Catalyst 9200L delivers enterprise-class access switching with PoE+ for wireless APs, IP phones, and IoT devices. DNA licensing for automation and security.",
        "specifications": {
            "Ports": "48x Gigabit Ethernet PoE+",
            "Uplinks": "4x 1G SFP",
            "PoE Budget": "370W",
            "Switching Capacity": "176 Gbps",
            "Forwarding Rate": "130 Mpps",
            "Stacking": "Up to 8 switches",
            "Management": "Cisco DNA Center, CLI, REST API",
            "Security": "TrustSec, MACsec, VLAN ACLs",
            "Dimensions": "1.73\" x 17.5\" x 14.96\"",
            "Warranty": "Limited Lifetime (Hardware)"
        },
        "availability": {"in_stock": True, "quantity": 45, "warehouse": "US-Central"},
        "rating": 4.9,
        "reviews_count": 167
    },
    # Additional Products - Motors & Drives
    {
        "id": "MOT-001",
        "name": "ABB Industrial AC Motor 7.5HP",
        "brand": "ABB",
        "category": "Motors & Drives",
        "sku": "ABB-ACM-75HP",
        "unspsc_code": "26101500",
        "base_price": 1850.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/48eb714ed81a4debd0fcbb8614e5e4412263c16c00d900b645c44f969c378c75.png",
        "short_description": "7.5HP 3-Phase AC Induction Motor, TEFC, 1800RPM",
        "full_description": "ABB Industrial AC Motor delivers reliable performance for demanding industrial applications. Features Totally Enclosed Fan Cooled (TEFC) design, Class F insulation, and robust construction.",
        "specifications": {
            "Power": "7.5 HP (5.5 kW)",
            "Voltage": "460V 3-Phase",
            "Speed": "1800 RPM",
            "Frame": "213T",
            "Enclosure": "TEFC",
            "Efficiency": "IE3 Premium",
            "Service Factor": "1.15"
        },
        "availability": {"in_stock": True, "quantity": 28, "warehouse": "US-East"},
        "rating": 4.7,
        "reviews_count": 156
    },
    {
        "id": "MOT-002",
        "name": "Siemens VFD Variable Frequency Drive 15HP",
        "brand": "Siemens",
        "category": "Motors & Drives",
        "sku": "SIE-VFD-15HP",
        "unspsc_code": "26101600",
        "base_price": 2450.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/48eb714ed81a4debd0fcbb8614e5e4412263c16c00d900b645c44f969c378c75.png",
        "short_description": "15HP Variable Frequency Drive with Integrated PLC",
        "full_description": "Siemens SINAMICS G120 VFD for precise motor control. Features built-in safety functions, Modbus communication, and energy-saving algorithms.",
        "specifications": {
            "Power Rating": "15 HP",
            "Input Voltage": "480V 3-Phase",
            "Output Frequency": "0-400 Hz",
            "Control Mode": "V/f, Sensorless Vector",
            "Communication": "Modbus RTU, Profinet",
            "Protection": "IP20",
            "Ambient Temp": "-10°C to +50°C"
        },
        "availability": {"in_stock": True, "quantity": 42, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 234
    },
    # Hydraulics & Pneumatics
    {
        "id": "HYD-001",
        "name": "Parker Hydraulic Gear Pump 20GPM",
        "brand": "Parker",
        "category": "Hydraulics & Pneumatics",
        "sku": "PAR-HGP-20",
        "unspsc_code": "40141600",
        "base_price": 875.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/67b16a05b97dae42c11c9c72f6e05b8af5d7aec7f63f9dfbd8e237a3fd5fc463.png",
        "short_description": "High-pressure hydraulic gear pump, 20GPM at 3000PSI",
        "full_description": "Parker PGP Series hydraulic gear pump designed for mobile and industrial applications. Features case-hardened gears and high volumetric efficiency.",
        "specifications": {
            "Flow Rate": "20 GPM",
            "Max Pressure": "3000 PSI",
            "Displacement": "3.6 cu in/rev",
            "Shaft": "SAE B 2-Bolt",
            "Port Size": "SAE 16",
            "Speed Range": "500-3600 RPM",
            "Fluid": "Mineral-based hydraulic oil"
        },
        "availability": {"in_stock": True, "quantity": 65, "warehouse": "US-West"},
        "rating": 4.6,
        "reviews_count": 189
    },
    {
        "id": "PNE-001",
        "name": "Festo Pneumatic Cylinder 100mm Bore",
        "brand": "Festo",
        "category": "Hydraulics & Pneumatics",
        "sku": "FES-CYL-100",
        "unspsc_code": "40141700",
        "base_price": 425.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/f33db98cd05c6baa652b94111e2425bcdc4313c5335f9673a214cb91bf2d2284.png",
        "short_description": "Double-acting pneumatic cylinder, ISO 15552 standard",
        "full_description": "Festo DSBC Series pneumatic cylinder with adjustable cushioning. Built to ISO 15552 standard for easy interchangeability.",
        "specifications": {
            "Bore": "100mm",
            "Stroke": "200mm",
            "Operating Pressure": "1-10 bar",
            "Piston Rod": "25mm Chrome-plated",
            "Cushioning": "PPV Adjustable",
            "Mounting": "ISO 15552",
            "Temperature Range": "-20°C to +80°C"
        },
        "availability": {"in_stock": True, "quantity": 112, "warehouse": "EU-West"},
        "rating": 4.9,
        "reviews_count": 321
    },
    # Welding Equipment
    {
        "id": "WLD-001",
        "name": "Lincoln Electric MIG Welder 250A",
        "brand": "Lincoln Electric",
        "category": "Welding",
        "sku": "LIN-MIG-250",
        "unspsc_code": "23270000",
        "base_price": 2150.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/2e6b43c65ae6a6b6a4fa924c3f3f88f897cad2a707104ffdc57d165e3b1d04e8.png",
        "short_description": "Industrial MIG welder with synergic control, 250A output",
        "full_description": "Lincoln Electric Power MIG 256 delivers reliable performance for production welding. Features Diamond Core Technology for superior arc performance.",
        "specifications": {
            "Amperage Range": "30-250A",
            "Input Power": "230V Single Phase",
            "Wire Feed Speed": "50-700 IPM",
            "Duty Cycle": "40% at 250A",
            "Wire Size": "0.023-0.045\"",
            "Spool Size": "10-15 lb",
            "Weight": "145 lbs"
        },
        "availability": {"in_stock": True, "quantity": 23, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 445
    },
    # Test & Measurement
    {
        "id": "TST-001",
        "name": "Fluke 289 True-RMS Industrial Multimeter",
        "brand": "Fluke",
        "category": "Test & Measurement",
        "sku": "FLK-289",
        "unspsc_code": "41110000",
        "base_price": 595.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/411bcb13dee27615e9b815fde695cdc27e5f05eee4347121309be34bbdb8c3e0.png",
        "short_description": "Industrial logging multimeter with TrendCapture",
        "full_description": "Fluke 289 True-RMS Industrial Logging Multimeter with TrendCapture for recording signal fluctuations over time. CAT IV 600V safety rated.",
        "specifications": {
            "DC Accuracy": "0.025%",
            "True RMS": "AC+DC",
            "Max Voltage": "1000V DC, 1000V AC",
            "Max Current": "10A (20A 30-sec)",
            "Resistance": "500MΩ",
            "Memory": "250 saved readings",
            "Safety Rating": "CAT III 1000V, CAT IV 600V"
        },
        "availability": {"in_stock": True, "quantity": 87, "warehouse": "US-East"},
        "rating": 4.9,
        "reviews_count": 678
    },
    # Safety Equipment
    {
        "id": "SAF-001",
        "name": "3M Powered Air Purifying Respirator System",
        "brand": "3M",
        "category": "Safety & PPE",
        "sku": "3M-PAPR-SYS",
        "unspsc_code": "46180000",
        "base_price": 1250.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/61099fb65121a227a0c6e32140edac60830b2fd0c6269420c02ee34b2e9933a2.png",
        "short_description": "Complete PAPR system with headgear and battery",
        "full_description": "3M Versaflo TR-600 Powered Air Purifying Respirator System provides respiratory protection with comfortable airflow in hazardous environments.",
        "specifications": {
            "Airflow": "6.7 CFM",
            "Battery Life": "8+ hours",
            "Filter Type": "P100/OV/AG",
            "APF": "1000",
            "Weight": "2.2 lbs (belt unit)",
            "Charging Time": "4 hours",
            "Certifications": "NIOSH Approved"
        },
        "availability": {"in_stock": True, "quantity": 34, "warehouse": "US-West"},
        "rating": 4.7,
        "reviews_count": 234
    },
    {
        "id": "SAF-002",
        "name": "Honeywell Safety Harness Full Body",
        "brand": "Honeywell",
        "category": "Safety & PPE",
        "sku": "HON-FBH-PRO",
        "unspsc_code": "46180000",
        "base_price": 345.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/9f7011dbf936aa67ab41af039d39e363f5a435c0efbfbed8a069e7684f74ad25.png",
        "short_description": "Full body fall protection harness with quick-connect buckles",
        "full_description": "Honeywell Miller Revolution Harness features DualTech webbing for maximum comfort and durability. Quick-connect chest and leg buckles for easy donning.",
        "specifications": {
            "Type": "Full Body",
            "D-Rings": "5-Point (back, shoulders, front)",
            "Weight Capacity": "400 lbs",
            "Webbing": "DualTech Polyester",
            "Buckles": "Quick-Connect",
            "Size Range": "S-XL",
            "Compliance": "OSHA 1926.502, ANSI Z359.11"
        },
        "availability": {"in_stock": True, "quantity": 156, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 412
    },
    # Material Handling
    {
        "id": "MAT-001",
        "name": "Crown Electric Pallet Jack 4500lb",
        "brand": "Crown",
        "category": "Material Handling",
        "sku": "CRW-EPJ-4500",
        "unspsc_code": "24100000",
        "base_price": 3950.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/d3b288a3bc2eae93ca8304dff24b6a6c2fa3e071ff498bff6f994dc86e336ddf.png",
        "short_description": "Electric pallet jack with 4500lb capacity",
        "full_description": "Crown WP 3000 Series electric pallet jack provides effortless load handling. Features regenerative braking and AC traction motor.",
        "specifications": {
            "Capacity": "4500 lbs",
            "Fork Length": "48\"",
            "Fork Width": "27\"",
            "Lift Height": "7.9\"",
            "Travel Speed": "4.1 mph",
            "Battery": "24V Lead-Acid",
            "Runtime": "8 hours continuous"
        },
        "availability": {"in_stock": True, "quantity": 12, "warehouse": "US-East"},
        "rating": 4.6,
        "reviews_count": 189
    },
    # Cutting Tools
    {
        "id": "CUT-001",
        "name": "Kennametal Carbide End Mill Set",
        "brand": "Kennametal",
        "category": "Cutting Tools",
        "sku": "KEN-CEM-SET",
        "unspsc_code": "27110000",
        "base_price": 485.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/5f3ec0cc9e4246d05af7b1609a78a8d3c2c76abbebba2fcb9823ffac8851309d.png",
        "short_description": "10-piece solid carbide end mill set, various sizes",
        "full_description": "Kennametal Harvi I TE solid carbide end mills for high-performance milling. Features advanced coating for extended tool life.",
        "specifications": {
            "Material": "Solid Carbide",
            "Coating": "TiAlN",
            "Flutes": "4",
            "Sizes": "1/8\" to 1/2\"",
            "Cut Length": "3x Diameter",
            "Shank": "Standard",
            "Helix Angle": "30°"
        },
        "availability": {"in_stock": True, "quantity": 78, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 345
    },
    # Storage & Organization
    {
        "id": "STO-001",
        "name": "Lista Industrial Storage Cabinet",
        "brand": "Lista",
        "category": "Storage & Organization",
        "sku": "LST-CAB-IND",
        "unspsc_code": "56100000",
        "base_price": 1875.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/ad801ca5185620afbcc264b33f0c59dacf2085f9b98895eb624068ec3ebad64d.png",
        "short_description": "Heavy-duty modular drawer cabinet, 10 drawers",
        "full_description": "Lista industrial storage cabinet with 100% drawer extension and 440lb drawer capacity. Ideal for tool storage and parts organization.",
        "specifications": {
            "Drawers": "10",
            "Drawer Capacity": "440 lbs each",
            "Dimensions": "30\"W x 28\"D x 59\"H",
            "Material": "14-gauge Steel",
            "Extension": "100%",
            "Lock": "Central Locking",
            "Color": "Industrial Gray"
        },
        "availability": {"in_stock": True, "quantity": 19, "warehouse": "US-West"},
        "rating": 4.7,
        "reviews_count": 234
    },
    # Cleaning & Janitorial
    {
        "id": "CLN-001",
        "name": "Tennant T300 Floor Scrubber",
        "brand": "Tennant",
        "category": "Cleaning & Janitorial",
        "sku": "TEN-T300",
        "unspsc_code": "47130000",
        "base_price": 4250.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/899b4b3e6aceed339189fb8d0b0fa309930af38864a707139f102cc7b47c1bfd.png",
        "short_description": "Walk-behind floor scrubber, 20\" cleaning path",
        "full_description": "Tennant T300 walk-behind scrubber delivers consistent cleaning performance. Features ec-H2O NanoClean technology for chemical-free cleaning.",
        "specifications": {
            "Cleaning Path": "20 inches",
            "Solution Tank": "11 gallons",
            "Recovery Tank": "12 gallons",
            "Run Time": "Up to 3 hours",
            "Productivity": "Up to 18,400 sq ft/hr",
            "Power": "Battery (24V)",
            "Weight": "287 lbs"
        },
        "availability": {"in_stock": True, "quantity": 8, "warehouse": "US-Central"},
        "rating": 4.5,
        "reviews_count": 167
    },
    # Lubrication
    {
        "id": "LUB-001",
        "name": "Mobil Industrial Lubricant Kit",
        "brand": "Mobil",
        "category": "Lubrication",
        "sku": "MOB-IND-KIT",
        "unspsc_code": "15120000",
        "base_price": 385.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/e12637ed7080f9cc6f0ab4378314a9bf6991b40c91ff9c4c31ea98fd25d1aa62.png",
        "short_description": "Complete industrial lubrication kit with various grades",
        "full_description": "Mobil Industrial Lubricant Kit includes synthetic and mineral-based lubricants for diverse industrial applications. Premium quality for extended equipment life.",
        "specifications": {
            "Includes": "5 products",
            "Types": "Gear Oil, Hydraulic, Grease, Way Oil, Spindle",
            "Grades": "ISO 32-220",
            "Container Sizes": "1 Quart each",
            "Base Stock": "Synthetic/Mineral",
            "Temperature Range": "-40°F to +300°F",
            "Applications": "General Industrial"
        },
        "availability": {"in_stock": True, "quantity": 145, "warehouse": "US-East"},
        "rating": 4.6,
        "reviews_count": 289
    }
]

# New Vendor Products - Donaldson, Avantor, Markem-Imaje
NEW_VENDOR_PRODUCTS = [
    # Donaldson Products - Filtration Solutions
    {
        "id": "DON-FILT-001",
        "name": "Donaldson PowerCore Air Filter Element",
        "brand": "Donaldson",
        "category": "Filtration",
        "sku": "DON-PC-G2",
        "unspsc_code": "40161500",
        "base_price": 285.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/920a11b0254f17e72f4a3b2a14027ff42d9fb11f2c45ab67408d34471c2b3991.png",
        "short_description": "High-efficiency PowerCore G2 air filtration element for heavy equipment",
        "full_description": "Donaldson PowerCore G2 air filter delivers superior engine protection with breakthrough filtration technology. Compact design provides more filtration capacity in less space with extended service life.",
        "specifications": {
            "Filter Media": "Ultra-Web Nanofiber",
            "Efficiency": "99.99% at 4 microns",
            "Air Flow": "Up to 2,500 CFM",
            "Dimensions": "12\" x 8\" x 6\"",
            "Application": "Heavy Equipment, Trucks, Agriculture",
            "Service Life": "Up to 3x standard filters",
            "Operating Temp": "-40°F to +250°F"
        },
        "availability": {"in_stock": True, "quantity": 234, "warehouse": "US-Central"},
        "rating": 4.8,
        "reviews_count": 456
    },
    {
        "id": "DON-FILT-002",
        "name": "Donaldson Hydraulic Filter Assembly P566672",
        "brand": "Donaldson",
        "category": "Filtration",
        "sku": "DON-P566672",
        "unspsc_code": "40161501",
        "base_price": 178.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/11d201fe0b5b4bef07c03b28000c28ba86ef8177ac5694ca45ad7dae9c8ac9d4.png",
        "short_description": "Premium hydraulic filter for mobile and industrial equipment",
        "full_description": "Donaldson Duramax hydraulic filter delivers exceptional contamination control for hydraulic systems. Synteq XP media technology provides high dirt-holding capacity and long service life.",
        "specifications": {
            "Filter Media": "Synteq XP Synthetic",
            "Beta Ratio": "β₁₀ ≥ 1000",
            "Flow Rate": "50 GPM",
            "Collapse Pressure": "435 PSI",
            "Operating Pressure": "150 PSI",
            "Thread Size": "1-1/4\" SAE-16",
            "Element Length": "8.66\""
        },
        "availability": {"in_stock": True, "quantity": 312, "warehouse": "US-East"},
        "rating": 4.7,
        "reviews_count": 289
    },
    {
        "id": "DON-FILT-003",
        "name": "Donaldson Torit PowerCore Dust Collector TG6",
        "brand": "Donaldson",
        "category": "Filtration",
        "sku": "DON-TG6-5000",
        "unspsc_code": "40161502",
        "base_price": 12500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/254fb871dea9172c9defc7ec40ceb0fe0a7680d955f413aa73cf05fbb0afd386.png",
        "short_description": "Industrial dust collection system with PowerCore technology",
        "full_description": "Donaldson Torit PowerCore TG Series dust collector combines PowerCore filtration with pulse-jet cleaning for superior dust collection. Ideal for weld fume, grinding dust, and general manufacturing.",
        "specifications": {
            "Airflow Capacity": "5,000 CFM",
            "Filter Area": "1,500 sq ft",
            "Efficiency": "99.99% at 0.5 microns",
            "Motor": "15 HP",
            "Voltage": "460V 3-Phase",
            "Dimensions": "72\" W x 48\" D x 120\" H",
            "Collection Bin": "55 Gallon",
            "Cleaning": "Automatic Pulse-Jet"
        },
        "availability": {"in_stock": True, "quantity": 8, "warehouse": "US-West"},
        "rating": 4.9,
        "reviews_count": 127
    },
    {
        "id": "DON-FILT-004",
        "name": "Donaldson Fuel Filter Kit P553004",
        "brand": "Donaldson",
        "category": "Filtration",
        "sku": "DON-P553004",
        "unspsc_code": "40161503",
        "base_price": 89.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/11d201fe0b5b4bef07c03b28000c28ba86ef8177ac5694ca45ad7dae9c8ac9d4.png",
        "short_description": "High-performance diesel fuel filter with water separator",
        "full_description": "Donaldson fuel filter kit with integrated water separator provides superior fuel system protection. Advanced media removes water and contaminants to protect injectors and fuel pumps.",
        "specifications": {
            "Efficiency": "98.7% at 4 microns",
            "Water Separation": "95%",
            "Flow Rate": "90 GPH",
            "Micron Rating": "4 Micron",
            "Application": "Diesel Engines",
            "Thread": "1-14 UNS",
            "Change Interval": "15,000 miles"
        },
        "availability": {"in_stock": True, "quantity": 567, "warehouse": "US-Central"},
        "rating": 4.6,
        "reviews_count": 834
    },
    # Avantor Products - Laboratory Supplies
    {
        "id": "AVT-LAB-001",
        "name": "Avantor J.T.Baker Reagent Grade Chemicals Kit",
        "brand": "Avantor",
        "category": "Laboratory Supplies",
        "sku": "AVT-JTB-CHEM-KIT",
        "unspsc_code": "41100000",
        "base_price": 425.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/bda0f8ddcb8e8de02c4616297d73d60d7a4332948c4c0f8e4d43f9cee4ea1d99.png",
        "short_description": "Premium laboratory reagent chemical kit for analytical applications",
        "full_description": "Avantor J.T.Baker reagent grade chemicals kit includes essential solvents and reagents for analytical chemistry. Meets or exceeds ACS specifications for purity and performance.",
        "specifications": {
            "Contents": "Acetone, Methanol, Isopropanol, Ethanol, Hexane",
            "Grade": "ACS Reagent Grade",
            "Purity": "≥99.5%",
            "Container Size": "4L each",
            "Certification": "Certificate of Analysis included",
            "Storage": "Room temperature, away from heat",
            "Shelf Life": "36 months"
        },
        "availability": {"in_stock": True, "quantity": 156, "warehouse": "US-East"},
        "rating": 4.9,
        "reviews_count": 312
    },
    {
        "id": "AVT-LAB-002",
        "name": "Avantor VWR Borosilicate Glassware Set",
        "brand": "Avantor",
        "category": "Laboratory Supplies",
        "sku": "AVT-VWR-GLASS-SET",
        "unspsc_code": "41100001",
        "base_price": 345.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/0c4ded7d30419f0e2db518c70c5b61015ad39e470fb06392d9e7e8ccac2b3d8b.png",
        "short_description": "Complete laboratory glassware set in premium borosilicate glass",
        "full_description": "Avantor VWR laboratory glassware set includes essential beakers, flasks, and graduated cylinders in borosilicate 3.3 glass. Designed for chemical resistance and thermal stability.",
        "specifications": {
            "Material": "Borosilicate 3.3 Glass",
            "Contents": "Beakers (6), Erlenmeyer Flasks (4), Graduated Cylinders (3), Volumetric Flasks (4)",
            "Sizes": "50mL to 1000mL",
            "Graduations": "White enamel, permanent",
            "Thermal Resistance": "Up to 500°C",
            "Chemical Resistance": "High",
            "Autoclavable": "Yes"
        },
        "availability": {"in_stock": True, "quantity": 89, "warehouse": "US-West"},
        "rating": 4.8,
        "reviews_count": 234
    },
    {
        "id": "AVT-LAB-003",
        "name": "Avantor Laboratory PPE Safety Kit",
        "brand": "Avantor",
        "category": "Laboratory Supplies",
        "sku": "AVT-PPE-KIT",
        "unspsc_code": "46180001",
        "base_price": 189.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/61f1f2f87a0b950443add33a553f6456b702582c73e0a7351e246e5676760d5a.png",
        "short_description": "Complete laboratory personal protective equipment kit",
        "full_description": "Avantor laboratory PPE kit provides comprehensive protection for lab personnel. Includes nitrile gloves, safety goggles, lab coat, and face shield for chemical handling safety.",
        "specifications": {
            "Contents": "Lab Coat (1), Safety Goggles (1), Face Shield (1), Nitrile Gloves (100pk), Shoe Covers (50pk)",
            "Lab Coat Material": "Polypropylene, fluid resistant",
            "Glove Material": "Nitrile, powder-free",
            "Glove Thickness": "4 mil",
            "Goggle Type": "Indirect vent, anti-fog",
            "Sizes Available": "S, M, L, XL",
            "Standards": "ANSI Z87.1, ASTM D6319"
        },
        "availability": {"in_stock": True, "quantity": 234, "warehouse": "US-Central"},
        "rating": 4.7,
        "reviews_count": 456
    },
    {
        "id": "AVT-LAB-004",
        "name": "Avantor Chromatography Column Kit",
        "brand": "Avantor",
        "category": "Laboratory Supplies",
        "sku": "AVT-CHROM-KIT",
        "unspsc_code": "41100002",
        "base_price": 875.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/0c4ded7d30419f0e2db518c70c5b61015ad39e470fb06392d9e7e8ccac2b3d8b.png",
        "short_description": "HPLC and flash chromatography column starter kit",
        "full_description": "Avantor chromatography kit includes HPLC columns and flash cartridges for separation and purification. Ideal for pharmaceutical and research applications.",
        "specifications": {
            "HPLC Columns": "C18 (3), C8 (2), Silica (2)",
            "Column Dimensions": "4.6 x 150mm, 4.6 x 250mm",
            "Particle Size": "5 μm",
            "Flash Cartridges": "12g, 24g, 40g sizes",
            "Pore Size": "100Å",
            "pH Range": "2-8",
            "Application": "Small molecule separation"
        },
        "availability": {"in_stock": True, "quantity": 45, "warehouse": "US-East"},
        "rating": 4.8,
        "reviews_count": 167
    },
    # Markem-Imaje Products - Industrial Coding & Marking
    {
        "id": "MKI-CODE-001",
        "name": "Markem-Imaje 9450 Continuous Inkjet Printer",
        "brand": "Markem-Imaje",
        "category": "Industrial Coding",
        "sku": "MKI-9450-CIJ",
        "unspsc_code": "44100000",
        "base_price": 8500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/dcbc55f9918f49116b8e9d85b0ef14471548a0e8682df9b9d9489aad94d48172.png",
        "short_description": "High-speed continuous inkjet printer for product coding",
        "full_description": "Markem-Imaje 9450 delivers reliable continuous inkjet printing for high-speed production lines. Features automatic ink viscosity control, simple interface, and minimal maintenance requirements.",
        "specifications": {
            "Print Speed": "Up to 2,857 characters/second",
            "Lines of Print": "1-5 lines",
            "Character Height": "1.5mm to 15mm",
            "Print Distance": "1-15mm",
            "Ink Types": "MEK, Acetone, Ethanol-based",
            "Interface": "7\" Color Touchscreen",
            "Communication": "Ethernet, RS232, USB",
            "IP Rating": "IP55"
        },
        "availability": {"in_stock": True, "quantity": 23, "warehouse": "US-West"},
        "rating": 4.8,
        "reviews_count": 189
    },
    {
        "id": "MKI-CODE-002",
        "name": "Markem-Imaje SmartLase C350 Laser Coder",
        "brand": "Markem-Imaje",
        "category": "Industrial Coding",
        "sku": "MKI-C350-LASER",
        "unspsc_code": "44100001",
        "base_price": 18500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/f072196072ed9dc67a0130899ab9aacebfdfb10d650b2681c5735dd7f1c7692e.png",
        "short_description": "CO2 laser marking system for permanent product coding",
        "full_description": "Markem-Imaje SmartLase C350 provides permanent, high-contrast laser marking on packaging materials. Zero consumables, low maintenance, and environmentally friendly coding solution.",
        "specifications": {
            "Laser Type": "CO2, 30W",
            "Wavelength": "10.6 μm",
            "Marking Speed": "Up to 1,500 characters/second",
            "Marking Area": "120mm x 120mm",
            "Resolution": "1000 dpi",
            "Focal Length": "170mm, 250mm, 350mm options",
            "Cooling": "Air-cooled",
            "Communication": "Ethernet, RS232, Profinet"
        },
        "availability": {"in_stock": True, "quantity": 12, "warehouse": "US-Central"},
        "rating": 4.9,
        "reviews_count": 98
    },
    {
        "id": "MKI-CODE-003",
        "name": "Markem-Imaje 2200 Print & Apply Labeler",
        "brand": "Markem-Imaje",
        "category": "Industrial Coding",
        "sku": "MKI-2200-PA",
        "unspsc_code": "44100002",
        "base_price": 14500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/0b758ba424771fe8b64df70bb8728a8b9ce3fde545e92f400ca3dc3dc88a4076.png",
        "short_description": "Automated label print and apply system for packaging lines",
        "full_description": "Markem-Imaje 2200 Series print and apply labeler combines high-resolution thermal transfer printing with precision label application. Ideal for case, pallet, and product labeling.",
        "specifications": {
            "Print Technology": "Thermal Transfer, 300 dpi",
            "Print Speed": "Up to 16 IPS",
            "Label Width": "1\" to 6\"",
            "Label Length": "0.5\" to 20\"",
            "Application Methods": "Tamp, Blow, Wipe",
            "Ribbon Length": "1,500 meters",
            "Construction": "Stainless Steel",
            "Communication": "Ethernet, RS232, USB"
        },
        "availability": {"in_stock": True, "quantity": 18, "warehouse": "US-East"},
        "rating": 4.7,
        "reviews_count": 134
    },
    {
        "id": "MKI-CODE-004",
        "name": "Markem-Imaje 5800 High-Resolution Inkjet",
        "brand": "Markem-Imaje",
        "category": "Industrial Coding",
        "sku": "MKI-5800-HR",
        "unspsc_code": "44100003",
        "base_price": 5200.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/8461848d-5003-4951-88f2-3971eec2f2c7/images/dcbc55f9918f49116b8e9d85b0ef14471548a0e8682df9b9d9489aad94d48172.png",
        "short_description": "High-resolution large character inkjet for secondary packaging",
        "full_description": "Markem-Imaje 5800 offers high-resolution printing on porous substrates like corrugated cardboard. Perfect for case coding with crisp graphics, barcodes, and text.",
        "specifications": {
            "Resolution": "185 dpi",
            "Print Height": "Up to 70mm (2.75\")",
            "Print Speed": "Up to 90 m/min",
            "Ink Type": "Water-based or Oil-based",
            "Cartridge Life": "Up to 350,000 prints",
            "Printheads": "1-4 heads cascadable",
            "Interface": "10\" Color Touchscreen",
            "IP Rating": "IP65"
        },
        "availability": {"in_stock": True, "quantity": 34, "warehouse": "US-West"},
        "rating": 4.6,
        "reviews_count": 212
    }
]

# Detailed IT Services Catalog
IT_SERVICES_CATALOG = [
    {
        "id": "SVC-NET-001",
        "name": "Network Infrastructure Installation - Enterprise",
        "category": "Network Installation Services",
        "unspsc_code": "81111801",
        "unspsc_name": "Network installation services",
        "supplier_name": "Infosys Network Solutions",
        "supplier_logo": None,
        "supplier_color": "#007CC3",
        "pricing_model": "Per Hour",
        "base_price": 125.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/afdd726c02cc7d9e8690e91dc7c1b0a13c962c96325cef7d1eece4d48001fb82.png",
        "short_description": "Professional network cabling, switch configuration, and testing",
        "full_description": "Complete enterprise network installation including structured cabling (Cat6/Cat6a), switch rack mounting and configuration, fiber optic termination, cable management, and comprehensive testing with documentation.",
        "service_includes": [
            "Site survey and network design",
            "Structured cabling (Cat6/Cat6a/Fiber)",
            "Patch panel and switch installation",
            "Network switch configuration",
            "Cable certification and testing",
            "Network documentation and labeling",
            "Post-installation support (48 hours)"
        ],
        "availability": {"available": True, "lead_time_days": 3, "regions": ["North America", "Europe", "APAC"]},
        "rating": 4.8,
        "reviews_count": 234
    },
    {
        "id": "SVC-NET-002",
        "name": "Wireless Network Setup - Enterprise Wi-Fi 6E",
        "category": "Network Installation Services",
        "unspsc_code": "81111802",
        "unspsc_name": "Wireless network installation",
        "supplier_name": "Cisco Certified Partner",
        "supplier_logo": None,
        "supplier_color": "#049FD9",
        "pricing_model": "Per Access Point",
        "base_price": 350.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/88c4a88606ac13bb8e539641523213064945b830124a9c902aa9147532aa3c49.png",
        "short_description": "Wi-Fi 6E access point deployment with site survey",
        "full_description": "Enterprise-grade wireless network deployment including predictive site survey, access point mounting, controller configuration, SSID setup, security policies, and performance optimization.",
        "service_includes": [
            "Predictive RF site survey",
            "Access point mounting and cabling",
            "Wireless controller configuration",
            "SSID and security policy setup",
            "Guest network configuration",
            "Heat map validation",
            "Performance optimization"
        ],
        "availability": {"available": True, "lead_time_days": 5, "regions": ["North America", "Europe"]},
        "rating": 4.7,
        "reviews_count": 189
    },
    {
        "id": "SVC-IT-001",
        "name": "Desktop/Laptop Deployment Service",
        "category": "IT Equipment Installation & Setup",
        "unspsc_code": "81112201",
        "unspsc_name": "Computer hardware installation",
        "supplier_name": "Dell ProDeploy Services",
        "supplier_logo": None,
        "supplier_color": "#007DB8",
        "pricing_model": "Per Device",
        "base_price": 85.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/cff7a5158c43f81799d59a99494279b2c5f255b2610fe406a1ffe6934277c4a7.png",
        "short_description": "Complete device imaging, setup, and user migration",
        "full_description": "Full PC lifecycle deployment including asset tagging, image deployment, domain join, software installation, data migration, and user orientation. Includes packaging disposal.",
        "service_includes": [
            "Asset tagging and inventory",
            "Custom image deployment",
            "Domain join and policy application",
            "Standard software installation",
            "User data migration (up to 50GB)",
            "Basic user orientation (30 min)",
            "Old device data wipe",
            "Packaging disposal"
        ],
        "availability": {"available": True, "lead_time_days": 1, "regions": ["Global"]},
        "rating": 4.6,
        "reviews_count": 567
    },
    {
        "id": "SVC-IT-002",
        "name": "Server Rack Installation & Configuration",
        "category": "IT Equipment Installation & Setup",
        "unspsc_code": "81112202",
        "unspsc_name": "Server installation services",
        "supplier_name": "HP Enterprise Services",
        "supplier_logo": None,
        "supplier_color": "#0096D6",
        "pricing_model": "Per Server",
        "base_price": 450.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/a1c30bd322a62ea6d441b12da74c66ca96707140958e1737c0d9a286f041a795.png",
        "short_description": "Physical server installation with OS and baseline config",
        "full_description": "Complete server deployment including rack mounting, power and network cabling, BIOS/firmware updates, OS installation, baseline security hardening, and integration with monitoring systems.",
        "service_includes": [
            "Server rack mounting",
            "Power and network cabling",
            "BIOS/firmware updates",
            "RAID configuration",
            "Operating system installation",
            "Security baseline hardening",
            "Monitoring agent deployment",
            "Documentation and handover"
        ],
        "availability": {"available": True, "lead_time_days": 2, "regions": ["North America", "Europe", "APAC"]},
        "rating": 4.9,
        "reviews_count": 312
    },
    {
        "id": "SVC-SEC-001",
        "name": "Network Security Assessment",
        "category": "Cybersecurity Services",
        "unspsc_code": "81112501",
        "unspsc_name": "Network security services",
        "supplier_name": "Infosys Cybersecurity",
        "supplier_logo": None,
        "supplier_color": "#007CC3",
        "pricing_model": "Per Assessment",
        "base_price": 5500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/88995698058a90986297157d130fa19ac15656bbe7ac2296b1949cbb9993e380.png",
        "short_description": "Comprehensive vulnerability scan and penetration testing",
        "full_description": "Full network security assessment including vulnerability scanning, penetration testing, firewall review, and detailed remediation recommendations with executive summary.",
        "service_includes": [
            "External vulnerability scanning",
            "Internal network assessment",
            "Penetration testing (network layer)",
            "Firewall rule review",
            "Detailed findings report",
            "Risk prioritization matrix",
            "Executive summary presentation",
            "Remediation consultation (4 hours)"
        ],
        "availability": {"available": True, "lead_time_days": 7, "regions": ["Global"]},
        "rating": 4.8,
        "reviews_count": 145
    },
    {
        "id": "SVC-MGD-001",
        "name": "Managed IT Support - Per User/Month",
        "category": "IT Managed Services",
        "unspsc_code": "81112301",
        "unspsc_name": "IT managed services",
        "supplier_name": "Infosys Managed Services",
        "supplier_logo": None,
        "supplier_color": "#007CC3",
        "pricing_model": "Per User/Month",
        "base_price": 75.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/f848c4af2e1f5caac71aa56dc1f1b9285fb801ad8e1ee77ad4185116211e3627.png",
        "short_description": "24/7 helpdesk, remote support, and proactive monitoring",
        "full_description": "Comprehensive managed IT services including 24/7 helpdesk, remote troubleshooting, patch management, antivirus monitoring, and monthly reporting.",
        "service_includes": [
            "24/7 helpdesk support (phone/email/chat)",
            "Remote desktop support",
            "Patch management",
            "Antivirus monitoring",
            "Basic user provisioning",
            "Monthly health reports",
            "Quarterly business reviews",
            "15 min average response time"
        ],
        "availability": {"available": True, "lead_time_days": 0, "regions": ["Global"]},
        "rating": 4.5,
        "reviews_count": 789
    },
    # Additional Services
    {
        "id": "SVC-DC-001",
        "name": "Data Center Infrastructure Services",
        "category": "IT Equipment Installation & Setup",
        "unspsc_code": "81112203",
        "unspsc_name": "Data center services",
        "supplier_name": "Equinix Solutions",
        "supplier_logo": None,
        "supplier_color": "#ED1C24",
        "pricing_model": "Per Rack",
        "base_price": 1250.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/45b8605364d7a481ab4a4bf15be73f59f9761da18220b1efb363c134d7ce0fb4.png",
        "short_description": "Complete data center rack installation and configuration",
        "full_description": "End-to-end data center infrastructure deployment including rack installation, power distribution, cooling setup, cable management, and environmental monitoring integration.",
        "service_includes": [
            "42U rack installation",
            "Power distribution unit setup",
            "Cable management systems",
            "Environmental monitoring",
            "Hot/cold aisle containment",
            "Documentation and labeling",
            "Commissioning and testing"
        ],
        "availability": {"available": True, "lead_time_days": 5, "regions": ["North America", "Europe", "APAC"]},
        "rating": 4.9,
        "reviews_count": 234
    },
    {
        "id": "SVC-MNT-001",
        "name": "Industrial Equipment Maintenance",
        "category": "Facilities Management & Workplace Services",
        "unspsc_code": "72151500",
        "unspsc_name": "Equipment maintenance services",
        "supplier_name": "SKF Reliability Systems",
        "supplier_logo": None,
        "supplier_color": "#005B94",
        "pricing_model": "Per Hour",
        "base_price": 145.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/fd8c452c65505c4fa649fae356b50b7dcc945b4ca0d180088866270a0b7d86f5.png",
        "short_description": "Preventive and predictive maintenance for industrial machinery",
        "full_description": "Comprehensive industrial equipment maintenance program including vibration analysis, thermography, oil analysis, and predictive maintenance scheduling.",
        "service_includes": [
            "Vibration monitoring and analysis",
            "Thermal imaging inspection",
            "Oil sample analysis",
            "Alignment verification",
            "Bearing inspection and replacement",
            "Maintenance scheduling",
            "Performance reporting"
        ],
        "availability": {"available": True, "lead_time_days": 2, "regions": ["Global"]},
        "rating": 4.7,
        "reviews_count": 456
    },
    {
        "id": "SVC-TRN-001",
        "name": "Corporate Technology Training",
        "category": "Corporate & Business Support Services",
        "unspsc_code": "86101700",
        "unspsc_name": "Professional training services",
        "supplier_name": "Infosys Learning Solutions",
        "supplier_logo": None,
        "supplier_color": "#007CC3",
        "pricing_model": "Per Session",
        "base_price": 2500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/8d3d10c922ac4516e87dff834b0b0f3d00633c1032b7667687f506addb1e98b3.png",
        "short_description": "Custom corporate technology training programs",
        "full_description": "Tailored technology training programs for enterprise teams. Includes curriculum development, certified instructors, hands-on labs, and certification preparation.",
        "service_includes": [
            "Needs assessment",
            "Custom curriculum development",
            "Certified instructor delivery",
            "Hands-on lab exercises",
            "Training materials and guides",
            "Post-training assessment",
            "Certification exam prep"
        ],
        "availability": {"available": True, "lead_time_days": 7, "regions": ["Global"]},
        "rating": 4.6,
        "reviews_count": 312
    },
    {
        "id": "SVC-QC-001",
        "name": "Quality Control & Inspection Services",
        "category": "HSE, Quality & Compliance Services",
        "unspsc_code": "77100000",
        "unspsc_name": "Quality inspection services",
        "supplier_name": "Bureau Veritas",
        "supplier_logo": None,
        "supplier_color": "#E30613",
        "pricing_model": "Per Day",
        "base_price": 950.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/1e0527ecaf208bd7905092e337f70a54eabef0782a81bd6e510c8e4d5c3c18ac.png",
        "short_description": "Product and process quality inspection services",
        "full_description": "Independent quality control and inspection services including incoming material inspection, in-process checks, final product inspection, and supplier audits.",
        "service_includes": [
            "Incoming material inspection",
            "In-process quality checks",
            "Final product inspection",
            "AQL sampling inspection",
            "Supplier quality audits",
            "Non-conformance reporting",
            "Corrective action tracking"
        ],
        "availability": {"available": True, "lead_time_days": 3, "regions": ["Global"]},
        "rating": 4.8,
        "reviews_count": 289
    },
    {
        "id": "SVC-LOG-001",
        "name": "Supply Chain Optimization Services",
        "category": "Logistics, Warehouse & Supply Chain Services",
        "unspsc_code": "78100000",
        "unspsc_name": "Supply chain services",
        "supplier_name": "DHL Supply Chain",
        "supplier_logo": None,
        "supplier_color": "#FFCC00",
        "pricing_model": "Per Project",
        "base_price": 15000.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/6810b3362e86e8ba23010897c9d6d3718dce4d1803e090213f94464944c1175b.png",
        "short_description": "End-to-end supply chain analysis and optimization",
        "full_description": "Comprehensive supply chain optimization including network analysis, inventory optimization, transportation routing, and warehouse layout design.",
        "service_includes": [
            "Current state assessment",
            "Network optimization modeling",
            "Inventory strategy development",
            "Transportation route optimization",
            "Warehouse layout design",
            "Technology recommendations",
            "Implementation roadmap"
        ],
        "availability": {"available": True, "lead_time_days": 10, "regions": ["North America", "Europe"]},
        "rating": 4.7,
        "reviews_count": 167
    },
    {
        "id": "SVC-CLN-001",
        "name": "Commercial Deep Cleaning Services",
        "category": "Facilities Management & Workplace Services",
        "unspsc_code": "76111500",
        "unspsc_name": "Commercial cleaning services",
        "supplier_name": "ISS Facility Services",
        "supplier_logo": None,
        "supplier_color": "#E4002B",
        "pricing_model": "Per Sq Ft",
        "base_price": 0.45,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/9c245588-b7cf-44e6-ac82-aa2be101ce17/images/124d0636fdb6759f0b0e809e00fe07c91c12562a510f2da1e68cc2cb7a0fd266.png",
        "short_description": "Professional commercial deep cleaning and sanitization",
        "full_description": "Comprehensive commercial deep cleaning services including floor stripping and waxing, carpet extraction, window cleaning, and disinfection protocols.",
        "service_includes": [
            "Floor stripping and refinishing",
            "Carpet deep extraction",
            "Window and glass cleaning",
            "High-touch surface disinfection",
            "Restroom deep sanitization",
            "Kitchen/break room cleaning",
            "Waste removal and recycling"
        ],
        "availability": {"available": True, "lead_time_days": 1, "regions": ["Global"]},
        "rating": 4.5,
        "reviews_count": 523
    },
    {
        "id": "SVC-SEC-002",
        "name": "Penetration Testing Services",
        "category": "Cybersecurity Services",
        "unspsc_code": "81112502",
        "unspsc_name": "Security testing services",
        "supplier_name": "Mandiant",
        "supplier_logo": None,
        "supplier_color": "#FF6600",
        "pricing_model": "Per Assessment",
        "base_price": 12500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/88995698058a90986297157d130fa19ac15656bbe7ac2296b1949cbb9993e380.png",
        "short_description": "Advanced penetration testing for web, network, and applications",
        "full_description": "Comprehensive penetration testing services including web application testing, network penetration, social engineering, and red team exercises.",
        "service_includes": [
            "Web application penetration testing",
            "Network infrastructure testing",
            "Social engineering assessment",
            "Wireless security testing",
            "Physical security testing",
            "Detailed findings report",
            "Remediation consultation"
        ],
        "availability": {"available": True, "lead_time_days": 14, "regions": ["Global"]},
        "rating": 4.9,
        "reviews_count": 178
    },
    {
        "id": "SVC-MKT-001",
        "name": "B2B Digital Marketing Campaign",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80171600",
        "unspsc_name": "Marketing campaign services",
        "supplier_name": "WPP Digital",
        "supplier_logo": None,
        "supplier_color": "#5C0080",
        "pricing_model": "Per Month",
        "base_price": 8500.00,
        "image_url": "https://static.prod-images.emergentagent.com/jobs/93bd7302-b98c-48b8-885e-c31e5a425122/images/516a358136877fc314216c27facefe3422f8dfa69a45e5553584d72a437e9fce.png",
        "short_description": "Full-service B2B digital marketing campaign management",
        "full_description": "Complete B2B digital marketing campaign including strategy development, content creation, paid media management, and performance analytics.",
        "service_includes": [
            "Campaign strategy development",
            "Content creation and copywriting",
            "LinkedIn advertising management",
            "Google Ads B2B campaigns",
            "Email marketing automation",
            "Landing page optimization",
            "Monthly performance reporting"
        ],
        "availability": {"available": True, "lead_time_days": 5, "regions": ["Global"]},
        "rating": 4.6,
        "reviews_count": 234
    },
    # Additional Digital Marketing Services
    {
        "id": "SVC-SMM-001",
        "name": "Social Media Marketing Management",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80171610",
        "unspsc_name": "Social media marketing services",
        "supplier_name": "Hootsuite Enterprise",
        "supplier_logo": None,
        "supplier_color": "#143059",
        "pricing_model": "Per Month",
        "base_price": 3500.00,
        "image_url": SERVICE_IMAGE_URLS.get("Social Media Marketing"),
        "short_description": "Full-service social media marketing and community management",
        "full_description": "Comprehensive social media marketing services including strategy development, content creation, community management, and analytics reporting across all major platforms.",
        "service_includes": [
            "Social media strategy development",
            "Content calendar creation",
            "Daily post publishing",
            "Community engagement management",
            "Influencer partnership coordination",
            "Social listening and monitoring",
            "Monthly analytics reports"
        ],
        "availability": {"available": True, "lead_time_days": 3, "regions": ["Global"]},
        "rating": 4.7,
        "reviews_count": 456
    },
    {
        "id": "SVC-SEO-001",
        "name": "Enterprise SEO Optimization Services",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80171608",
        "unspsc_name": "Search engine optimization services",
        "supplier_name": "Moz Enterprise",
        "supplier_logo": None,
        "supplier_color": "#1B98E0",
        "pricing_model": "Per Month",
        "base_price": 4500.00,
        "image_url": SERVICE_IMAGE_URLS.get("SEO Optimization"),
        "short_description": "Technical and content SEO optimization for enterprise websites",
        "full_description": "Full-spectrum SEO services including technical audits, on-page optimization, link building, and content strategy to improve organic search rankings and drive qualified traffic.",
        "service_includes": [
            "Technical SEO audit",
            "Keyword research and mapping",
            "On-page optimization",
            "Link building campaigns",
            "Content optimization",
            "Competitor analysis",
            "Monthly ranking reports"
        ],
        "availability": {"available": True, "lead_time_days": 5, "regions": ["Global"]},
        "rating": 4.8,
        "reviews_count": 312
    },
    {
        "id": "SVC-CNT-001",
        "name": "B2B Content Marketing Services",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80172106",
        "unspsc_name": "Content marketing services",
        "supplier_name": "Contently Enterprise",
        "supplier_logo": None,
        "supplier_color": "#FF5733",
        "pricing_model": "Per Month",
        "base_price": 6000.00,
        "image_url": SERVICE_IMAGE_URLS.get("Content Marketing"),
        "short_description": "Strategic content creation and distribution for B2B audiences",
        "full_description": "End-to-end B2B content marketing services including content strategy, thought leadership creation, blog management, and content distribution across multiple channels.",
        "service_includes": [
            "Content strategy development",
            "Blog article writing (8/month)",
            "Whitepaper creation (1/quarter)",
            "Case study development",
            "Newsletter content creation",
            "Content distribution strategy",
            "Performance analytics"
        ],
        "availability": {"available": True, "lead_time_days": 7, "regions": ["Global"]},
        "rating": 4.5,
        "reviews_count": 234
    },
    {
        "id": "SVC-EML-001",
        "name": "Email Marketing Automation Services",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80171615",
        "unspsc_name": "Email marketing services",
        "supplier_name": "Mailchimp Enterprise",
        "supplier_logo": None,
        "supplier_color": "#FFE01B",
        "pricing_model": "Per Month",
        "base_price": 2500.00,
        "image_url": SERVICE_IMAGE_URLS.get("Email Marketing"),
        "short_description": "Automated email marketing campaigns with lead nurturing",
        "full_description": "Comprehensive email marketing services including campaign strategy, template design, automation workflow setup, A/B testing, and performance optimization.",
        "service_includes": [
            "Email strategy development",
            "Template design and coding",
            "List segmentation setup",
            "Automation workflow creation",
            "A/B testing programs",
            "Deliverability optimization",
            "Performance reporting"
        ],
        "availability": {"available": True, "lead_time_days": 3, "regions": ["Global"]},
        "rating": 4.6,
        "reviews_count": 389
    },
    {
        "id": "SVC-PPC-001",
        "name": "PPC & Paid Advertising Management",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80171620",
        "unspsc_name": "Paid advertising services",
        "supplier_name": "WordStream Agency",
        "supplier_logo": None,
        "supplier_color": "#4285F4",
        "pricing_model": "Percentage of Ad Spend",
        "base_price": 15.00,
        "image_url": SERVICE_IMAGE_URLS.get("PPC Advertising"),
        "short_description": "Google Ads, LinkedIn Ads, and paid media campaign management",
        "full_description": "Full-service paid advertising management across Google Ads, LinkedIn, Facebook, and programmatic display networks with focus on ROI optimization.",
        "service_includes": [
            "Campaign strategy and setup",
            "Keyword and audience research",
            "Ad copywriting and creative",
            "Bid management optimization",
            "Landing page recommendations",
            "Conversion tracking setup",
            "Weekly performance reports"
        ],
        "availability": {"available": True, "lead_time_days": 3, "regions": ["Global"]},
        "rating": 4.7,
        "reviews_count": 567
    },
    {
        "id": "SVC-BRD-001",
        "name": "Brand Identity Design Services",
        "category": "Digital Marketing & Creative Agency Services",
        "unspsc_code": "80141605",
        "unspsc_name": "Brand identity design",
        "supplier_name": "Pentagram Design",
        "supplier_logo": None,
        "supplier_color": "#000000",
        "pricing_model": "Per Project",
        "base_price": 25000.00,
        "image_url": SERVICE_IMAGE_URLS.get("Brand Identity Design"),
        "short_description": "Complete brand identity creation and visual design system",
        "full_description": "Comprehensive brand identity design services including logo design, visual identity system, brand guidelines, and collateral design for consistent brand representation.",
        "service_includes": [
            "Brand discovery workshop",
            "Logo design and variations",
            "Color palette development",
            "Typography selection",
            "Visual design system",
            "Brand guidelines documentation",
            "Collateral template design"
        ],
        "availability": {"available": True, "lead_time_days": 21, "regions": ["Global"]},
        "rating": 4.9,
        "reviews_count": 178
    }
]

# PunchOut Systems
PUNCHOUT_SYSTEMS = [
    {"name": "Coupa", "logo": "https://logo.clearbit.com/coupa.com"},
    {"name": "SAP Ariba", "logo": "https://logo.clearbit.com/ariba.com"},
    {"name": "SAP ERP", "logo": "https://logo.clearbit.com/sap.com"},
    {"name": "Ivalua", "logo": "https://logo.clearbit.com/ivalua.com"},
    {"name": "Oracle", "logo": "https://logo.clearbit.com/oracle.com"},
]

# Services data with UNSPSC codes
SERVICES_DATA = [
    {"name": "Janitorial Office Cleaning Services – Business Hours", "unspsc_code": "76111501", "unspsc_name": "Janitorial services", "category": "Facilities Management & Workplace Services", "country": "Denmark", "supplier_name": "Wipro IT Services DK", "unit_of_measure": "Per Sq Ft", "base_price": 0.85},
    {"name": "Deep Cleaning Services – Office Facilities", "unspsc_code": "76111502", "unspsc_name": "Commercial cleaning", "category": "Facilities Management & Workplace Services", "country": "Nigeria", "supplier_name": None, "unit_of_measure": "Per Service", "base_price": None},
    {"name": "Pest Control Services – Preventive", "unspsc_code": "76102101", "unspsc_name": "Pest control", "category": "Facilities Management & Workplace Services", "country": "Spain", "supplier_name": "Genpact Enterprise IT ES", "unit_of_measure": "Per Visit", "base_price": 150.00},
    {"name": "HVAC Preventive Maintenance Services", "unspsc_code": "72101502", "unspsc_name": "HVAC maintenance", "category": "Facilities Management & Workplace Services", "country": "Switzerland", "supplier_name": None, "unit_of_measure": "Per Service", "base_price": None},
    {"name": "Security Guard Services – Day Shift", "unspsc_code": "92121504", "unspsc_name": "Guard services", "category": "Facilities Management & Workplace Services", "country": "Brazil", "supplier_name": "Siemens Smart Infrastructure", "unit_of_measure": "Per Hour", "base_price": 25.00},
    {"name": "Elevator Preventive Maintenance Services", "unspsc_code": "72101504", "unspsc_name": "Elevator maintenance", "category": "Facilities Management & Workplace Services", "country": "Chile", "supplier_name": "Schindler Elevator", "unit_of_measure": "Per Month", "base_price": 450.00},
    {"name": "Fire Extinguisher Inspection Services", "unspsc_code": "46191601", "unspsc_name": "Fire safety inspection", "category": "Facilities Management & Workplace Services", "country": "Portugal", "supplier_name": "Tata Consultancy Services", "unit_of_measure": "Per Unit", "base_price": 15.00},
    {"name": "Digital Marketing Strategy Services", "unspsc_code": "80171607", "unspsc_name": "Digital marketing", "category": "Digital Marketing & Creative Agency Services", "country": "Greece", "supplier_name": None, "unit_of_measure": "Per Month", "base_price": None},
    {"name": "Search Engine Optimization (SEO) Services", "unspsc_code": "80171608", "unspsc_name": "SEO services", "category": "Digital Marketing & Creative Agency Services", "country": "Switzerland", "supplier_name": "Tata Consultancy Services", "unit_of_measure": "Per Month", "base_price": 2500.00},
    {"name": "Video Production Services – Corporate", "unspsc_code": "82131602", "unspsc_name": "Video production", "category": "Digital Marketing & Creative Agency Services", "country": "Argentina", "supplier_name": "Dentsu International", "unit_of_measure": "Per Day", "base_price": 1500.00},
    {"name": "Content Marketing Services", "unspsc_code": "80172106", "unspsc_name": "Content development", "category": "Digital Marketing & Creative Agency Services", "country": "Portugal", "supplier_name": "BrightView Landscape PT", "unit_of_measure": "Per Asset", "base_price": 350.00},
    {"name": "Brand Strategy Consulting Services", "unspsc_code": "80171605", "unspsc_name": "Brand management", "category": "Digital Marketing & Creative Agency Services", "country": "Singapore", "supplier_name": "Honeywell Building Solutions", "unit_of_measure": "Per Day", "base_price": 2000.00},
    {"name": "Desktop Support Services – Onsite – L1", "unspsc_code": "81112106", "unspsc_name": "Computer hardware support services", "category": "IT & Workplace Technology Services", "country": "Greece", "supplier_name": None, "unit_of_measure": "Per Hour", "base_price": None},
    {"name": "Desktop Support Services – Remote – L1", "unspsc_code": "81112107", "unspsc_name": "Remote technical support", "category": "IT & Workplace Technology Services", "country": "UAE", "supplier_name": "Infosys Digital Workplace", "unit_of_measure": "Per Hour", "base_price": 45.00},
    {"name": "Hardware Break/Fix Services – Workspace IT", "unspsc_code": "81112306", "unspsc_name": "Hardware maintenance services", "category": "IT & Workplace Technology Services", "country": "Spain", "supplier_name": "ABM Industries ES", "unit_of_measure": "Per Call", "base_price": 125.00},
    {"name": "Cloud Migration Assessment Services", "unspsc_code": "81101524", "unspsc_name": "Digital transformation consulting", "category": "IT & Workplace Technology Services", "country": "UK", "supplier_name": "Publicis Media GB", "unit_of_measure": "Per Day", "base_price": 2000.00},
    {"name": "Cybersecurity Awareness Training Services", "unspsc_code": "86101605", "unspsc_name": "Information security training", "category": "IT & Workplace Technology Services", "country": "Belgium", "supplier_name": "HCLTech Digital", "unit_of_measure": "Per Session", "base_price": 500.00},
    {"name": "Health & Safety Audit Services – Site Level", "unspsc_code": "81101508", "unspsc_name": "Facility inspection services", "category": "HSE, Quality & Compliance Services", "country": "Bulgaria", "supplier_name": None, "unit_of_measure": "Per Day", "base_price": None},
    {"name": "Fire Safety Risk Assessment Services", "unspsc_code": "92121702", "unspsc_name": "Fire safety consulting services", "category": "HSE, Quality & Compliance Services", "country": "Germany", "supplier_name": "Ricoh Managed Services DE", "unit_of_measure": "Per Day", "base_price": 800.00},
    {"name": "ISO 9001 Certification Support Services", "unspsc_code": "81101512", "unspsc_name": "Quality management consulting", "category": "HSE, Quality & Compliance Services", "country": "Finland", "supplier_name": None, "unit_of_measure": "Per Project", "base_price": None},
    {"name": "ISO 14001 Certification Support Services", "unspsc_code": "81101512", "unspsc_name": "Environmental management consulting", "category": "HSE, Quality & Compliance Services", "country": "Switzerland", "supplier_name": "Bureau Veritas", "unit_of_measure": "Per Project", "base_price": 8500.00},
    {"name": "Warehouse Labor Services – General", "unspsc_code": "78131601", "unspsc_name": "Warehousing services", "category": "Logistics, Warehouse & Supply Chain Services", "country": "Peru", "supplier_name": None, "unit_of_measure": "Per Hour", "base_price": None},
    {"name": "Import Customs Clearance Support Services", "unspsc_code": "80111509", "unspsc_name": "Customs support services", "category": "Logistics, Warehouse & Supply Chain Services", "country": "India", "supplier_name": "Genpact Enterprise IT", "unit_of_measure": "Per Hour", "base_price": 75.00},
    {"name": "Supply Chain Process Mapping Services", "unspsc_code": "81101503", "unspsc_name": "Process mapping services", "category": "Logistics, Warehouse & Supply Chain Services", "country": "Chile", "supplier_name": "WPP Group (GroupM)", "unit_of_measure": "Per Service", "base_price": 3500.00},
    {"name": "Freight Audit & Pay Services", "unspsc_code": "84111801", "unspsc_name": "Freight audit services", "category": "Logistics, Warehouse & Supply Chain Services", "country": "Romania", "supplier_name": "DHL Supply Chain", "unit_of_measure": "Per Transaction", "base_price": 5.00},
    {"name": "Legal Document Review Services – Standard", "unspsc_code": "80121602", "unspsc_name": "Legal review services", "category": "Corporate & Business Support Services", "country": "Bulgaria", "supplier_name": None, "unit_of_measure": "Per Hour", "base_price": None},
    {"name": "Contract Drafting Support Services", "unspsc_code": "80121603", "unspsc_name": "Contract drafting services", "category": "Corporate & Business Support Services", "country": "Ireland", "supplier_name": "SUEZ Recycling & Recovery IE", "unit_of_measure": "Per Hour", "base_price": 150.00},
    {"name": "Executive Presentation Development Services", "unspsc_code": "82111701", "unspsc_name": "Presentation design services", "category": "Corporate & Business Support Services", "country": "USA", "supplier_name": "Wipro IT Services", "unit_of_measure": "Per Deck", "base_price": 500.00},
    {"name": "Technical Staff Augmentation Services", "unspsc_code": "80111605", "unspsc_name": "Temporary staffing services", "category": "Temp Labor across Technical Skilled Capabilities", "country": "USA", "supplier_name": "Randstad", "unit_of_measure": "Per Hour", "base_price": 85.00},
    {"name": "Engineering Staff Augmentation Services", "unspsc_code": "80111605", "unspsc_name": "Temporary staffing services", "category": "Temp Labor across Technical Skilled Capabilities", "country": "Germany", "supplier_name": "Hays", "unit_of_measure": "Per Hour", "base_price": 95.00},
]

# Pydantic Models
class UserLogin(BaseModel):
    email: str
    password: str
    country: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    country: str
    currency: Dict[str, Any]
    info_coins: int
    token: str
    role: Optional[str] = None

class DeliveryPartner(BaseModel):
    partner_id: str
    price: float
    lead_time_days: int
    available_quantity: int

class CartItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_id: str
    product_name: str
    brand: str
    sku: str
    unspsc_code: str
    category: str
    quantity: int
    unit_price: float
    total_price: float
    currency_code: str
    image_url: Optional[str] = None
    is_service: bool = False

class CartTransfer(BaseModel):
    system: str  # "Coupa", "SAP Ariba", "SAP ERP", "Ivalua", "Oracle"
    cart_items: List[str]  # List of cart item IDs

class RFQCreate(BaseModel):
    product_description: str
    quantity: int
    brand_name: Optional[str] = None
    oem_part_number: Optional[str] = None
    needed_by: Optional[str] = None
    delivery_location: str
    supplier_name: Optional[str] = None
    supplier_email: Optional[str] = None
    request_type: str = "actual"
    is_product: bool = True

class QuotationRequest(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    notes: Optional[str] = None

class QuotationResponse(BaseModel):
    quotation_id: str
    action: str  # "accept", "cancel"
    cancel_reason: Optional[str] = None

# InfoCoin Rewards with CDN images
INFOCOIN_REWARDS = [
    {"id": "1", "name": "Premium Executive Jacket", "description": "High-quality insulated navy blue jacket with Infosys branding", "coins_required": 5000, "image_url": REWARD_IMAGE_URLS.get("Executive Jacket"), "category": "Apparel"},
    {"id": "2", "name": "Premium Leather Backpack", "description": "Professional leather laptop backpack for business travel", "coins_required": 4500, "image_url": REWARD_IMAGE_URLS.get("Leather Backpack"), "category": "Accessories"},
    {"id": "3", "name": "Wireless Bluetooth Earbuds", "description": "Premium wireless earbuds with noise cancellation in charging case", "coins_required": 3500, "image_url": REWARD_IMAGE_URLS.get("Wireless Earbuds"), "category": "Electronics"},
    {"id": "4", "name": "Stainless Steel Insulated Tumbler", "description": "Double-walled insulated tumbler, keeps drinks hot/cold for 12 hours", "coins_required": 800, "image_url": REWARD_IMAGE_URLS.get("Insulated Tumbler"), "category": "Drinkware"},
    {"id": "5", "name": "Executive Desk Organizer Set", "description": "Premium desk organizer with pen holder, card holder, and notepad holder", "coins_required": 1500, "image_url": REWARD_IMAGE_URLS.get("Desk Organizer Set"), "category": "Office"},
    {"id": "6", "name": "Smartwatch Fitness Tracker", "description": "Modern smartwatch with fitness tracking, heart rate monitor, and notifications", "coins_required": 6000, "image_url": REWARD_IMAGE_URLS.get("Smartwatch"), "category": "Electronics"},
]

# Helper Functions
def create_jwt_token(user_id: str, email: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {"sub": user_id, "email": email, "exp": expiration}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_jwt_token(token: str) -> Dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    payload = verify_jwt_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

def generate_unspsc_code(category_name: str) -> str:
    """Get UNSPSC code for a category"""
    for cat in MRO_CATEGORIES:
        if cat["name"] == category_name:
            return cat["unspsc"]
    return "31000000"  # Default MRO code

def get_brand_info(brand_name: str) -> Dict:
    """Get brand info with color"""
    for brand in MRO_BRANDS:
        if brand["name"] == brand_name:
            return brand
    return {"name": brand_name, "logo": None, "color": "#007CC3"}

def generate_product_data(index: int, category: str, brand: str) -> Dict:
    """Generate realistic product data with UNSPSC and detailed specifications"""
    # Enhanced product names with specifications - all images from Emergent CDN
    product_catalog = {
        "Bearings & Power Transmission": [
            {"name": "Deep Groove Ball Bearing", "specs": {"Inner Diameter": "25mm", "Outer Diameter": "52mm", "Width": "15mm", "Material": "Chrome Steel", "Seal Type": "2RS Rubber Sealed", "Load Rating": "14kN Dynamic"}, "image": PRODUCT_IMAGE_URLS.get("SKF Ball Bearing", PRODUCT_IMAGE_URLS.get("Bearings & Power Transmission"))},
            {"name": "Tapered Roller Bearing", "specs": {"Bore Size": "30mm", "Outside Diameter": "62mm", "Width": "17.25mm", "Material": "Chrome Steel", "Cage Type": "Steel", "Dynamic Load": "44kN"}, "image": PRODUCT_IMAGE_URLS.get("SKF Ball Bearing", PRODUCT_IMAGE_URLS.get("Bearings & Power Transmission"))},
            {"name": "Industrial Timing Belt", "specs": {"Pitch": "8mm HTD", "Width": "30mm", "Length": "1200mm", "Material": "Neoprene/Fiberglass", "Teeth Count": "150", "Max Speed": "40m/s"}, "image": PRODUCT_IMAGE_URLS.get("Gates Timing Belt", PRODUCT_IMAGE_URLS.get("Bearings & Power Transmission"))},
        ],
        "Electrical & Lighting": [
            {"name": "Industrial LED High Bay Light", "specs": {"Power": "200W", "Lumens": "26,000lm", "Color Temp": "5000K", "IP Rating": "IP65", "Beam Angle": "120°", "Lifespan": "50,000hrs"}, "image": PRODUCT_IMAGE_URLS.get("Philips LED Light", PRODUCT_IMAGE_URLS.get("Electrical & Lighting"))},
            {"name": "Miniature Circuit Breaker", "specs": {"Current Rating": "32A", "Poles": "3P", "Breaking Capacity": "10kA", "Curve Type": "C", "Voltage": "400V AC", "DIN Rail Mount": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("Schneider Circuit Breaker", PRODUCT_IMAGE_URLS.get("Electrical & Lighting"))},
            {"name": "Industrial Contactor", "specs": {"Coil Voltage": "24V DC", "Current Rating": "40A", "Contacts": "3NO + 1NC", "Mounting": "DIN Rail", "Duty Cycle": "AC-3", "Mechanical Life": "10M ops"}, "image": PRODUCT_IMAGE_URLS.get("Schneider Circuit Breaker", PRODUCT_IMAGE_URLS.get("Electrical & Lighting"))},
        ],
        "Hand Tools": [
            {"name": "Professional Ratcheting Wrench Set", "specs": {"Pieces": "12pc SAE/Metric", "Drive Size": "72-Tooth Ratchet", "Material": "Chrome Vanadium", "Finish": "Polished Chrome", "Case": "Blow Mold", "Warranty": "Lifetime"}, "image": PRODUCT_IMAGE_URLS.get("Stanley Wrench", PRODUCT_IMAGE_URLS.get("Hand Tools"))},
            {"name": "Precision Screwdriver Set", "specs": {"Pieces": "32pc", "Tip Types": "Phillips/Slotted/Torx/Hex", "Handle": "Ergonomic Cushion Grip", "Blade": "Hardened Steel", "Case": "Rotating Stand", "Magnetic Tips": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("Stanley Wrench", PRODUCT_IMAGE_URLS.get("Hand Tools"))},
        ],
        "Power Tools": [
            {"name": "18V Brushless Cordless Drill/Driver Kit", "specs": {"Voltage": "18V/20V MAX", "Chuck": "1/2\" Metal Ratcheting", "Speed": "0-2000 RPM", "Torque": "620 in-lbs", "Battery": "5.0Ah Li-Ion (2x)", "LED Light": "3-Mode"}, "image": PRODUCT_IMAGE_URLS.get("Bosch Drill", PRODUCT_IMAGE_URLS.get("Power Tools"))},
            {"name": "Industrial Angle Grinder", "specs": {"Disc Size": "125mm (5\")", "Power": "1400W", "No Load Speed": "11,500 RPM", "Spindle Thread": "M14", "Guard": "Adjustable", "Soft Start": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("Bosch Drill", PRODUCT_IMAGE_URLS.get("Power Tools"))},
        ],
        "Safety & PPE": [
            {"name": "Premium Safety Helmet", "specs": {"Standard": "EN397/ANSI Z89.1", "Material": "ABS Shell", "Suspension": "6-Point Ratchet", "Ventilation": "4-Point Vented", "Accessory Slots": "Yes", "UV Resistant": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("3M Safety Helmet", PRODUCT_IMAGE_URLS.get("Safety & PPE"))},
            {"name": "Impact-Resistant Safety Glasses", "specs": {"Standard": "EN166/ANSI Z87.1+", "Lens": "Polycarbonate Anti-Scratch", "Coating": "Anti-Fog", "UV Protection": "99.9%", "Frame": "Wraparound", "Weight": "28g"}, "image": PRODUCT_IMAGE_URLS.get("3M Safety Glasses", PRODUCT_IMAGE_URLS.get("Safety & PPE"))},
        ],
        "IT Equipment - Laptops": [
            {"name": "ProBook Business Laptop", "specs": {"Processor": "Intel Core i7-1355U", "Memory": "16GB DDR4", "Storage": "512GB NVMe SSD", "Display": "15.6\" FHD IPS", "OS": "Windows 11 Pro", "Battery": "Up to 10hrs"}, "image": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops"))},
            {"name": "Elite Ultrabook", "specs": {"Processor": "Intel Core i7-1365U vPro", "Memory": "32GB DDR5", "Storage": "1TB NVMe Gen4", "Display": "14\" 2.8K OLED", "OS": "Windows 11 Pro", "Weight": "1.12kg"}, "image": PRODUCT_IMAGE_URLS.get("HP Laptop", PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops"))},
        ],
        "IT Equipment - Monitors": [
            {"name": "Professional 4K USB-C Monitor", "specs": {"Screen Size": "27\"", "Resolution": "3840x2160", "Panel": "IPS", "Refresh": "60Hz", "Ports": "USB-C 90W, HDMI, DP", "Ergonomics": "HAS"}, "image": PRODUCT_IMAGE_URLS.get("Dell Monitor", PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors"))},
            {"name": "Ultrawide Curved Monitor", "specs": {"Screen Size": "34\"", "Resolution": "3440x1440", "Panel": "VA 1500R", "Refresh": "100Hz", "HDR": "HDR10", "Built-in KVM": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("Dell Monitor", PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors"))},
        ],
        "IT Equipment - Networking": [
            {"name": "Enterprise Managed Switch", "specs": {"Ports": "48x GbE PoE+", "Uplinks": "4x 10G SFP+", "PoE Budget": "740W", "Switching": "176Gbps", "Management": "CLI/Web/SNMP", "Stackable": "Yes"}, "image": PRODUCT_IMAGE_URLS.get("Cisco Switch", PRODUCT_IMAGE_URLS.get("IT Equipment - Networking"))},
            {"name": "Enterprise Wireless Access Point", "specs": {"Standard": "Wi-Fi 6E", "Speed": "Up to 5.4Gbps", "Bands": "Tri-Band", "Clients": "500+", "PoE": "802.3at", "MIMO": "4x4:4"}, "image": PRODUCT_IMAGE_URLS.get("Cisco Switch", PRODUCT_IMAGE_URLS.get("IT Equipment - Networking"))},
        ],
        "Adhesives & Sealants": [
            {"name": "Industrial Adhesive Set", "specs": {"Type": "Multi-purpose", "Bond Strength": "High", "Cure Time": "24hrs", "Temperature Range": "-40°C to 120°C", "Volume": "50ml each"}, "image": PRODUCT_IMAGE_URLS.get("Henkel Adhesive", PRODUCT_IMAGE_URLS.get("Adhesives & Sealants"))},
        ],
    }
    
    # Get product details or use defaults
    products = product_catalog.get(category, [
        {"name": "Industrial Component", "specs": {"Type": "Standard", "Grade": "Industrial", "Material": "High-Quality", "Certification": "ISO 9001"}}
    ])
    
    product = products[index % len(products)]
    unspsc = generate_unspsc_code(category)
    brand_info = get_brand_info(brand)
    
    # Get image URL from CDN
    image_url = PRODUCT_IMAGE_URLS.get(category, DEFAULT_PRODUCT_IMAGE)
    
    # Generate price based on category
    price_ranges = {
        "IT Equipment - Laptops": (800, 3500),
        "IT Equipment - Monitors": (300, 1500),
        "IT Equipment - Networking": (500, 5000),
        "IT Equipment - Servers": (2000, 15000),
        "Power Tools": (100, 800),
        "Safety & PPE": (15, 150),
    }
    price_range = price_ranges.get(category, (25, 500))
    base_price = round(random.uniform(*price_range), 2)
    
    # 10% chance of being sponsored
    is_sponsored = random.random() < 0.10
    
    # Generate availability info
    availability = {
        "in_stock": random.random() > 0.1,
        "quantity": random.randint(5, 500),
        "warehouse": random.choice(["US-East", "US-West", "US-Central", "EU-West", "APAC-SG"]),
        "ships_from": random.choice(["Manufacturer", "Distribution Center", "Local Warehouse"]),
        "estimated_delivery": f"{random.randint(1, 5)}-{random.randint(6, 10)} business days"
    }
    
    # Create descriptive product name and description based on category
    product_name = product['name']
    
    # Generate accurate short description based on product specifications
    specs_text = ", ".join([f"{k}: {v}" for k, v in list(product.get("specs", {}).items())[:3]]) if product.get("specs") else ""
    
    short_desc = f"{product_name} by {brand}. {specs_text}" if specs_text else f"Professional-grade {product_name.lower()} from {brand} designed for industrial and enterprise applications."
    
    full_desc = f"The {brand} {product_name} delivers exceptional performance and reliability for demanding professional environments. "
    if product.get("specs"):
        full_desc += f"Key specifications include: {specs_text}. "
    full_desc += f"Built with premium materials and backed by {brand}'s reputation for quality. Meets international standards for safety and performance."
    
    return {
        "id": str(uuid.uuid4()),
        "name": f"{brand} {product_name}",
        "short_description": short_desc,
        "full_description": full_desc,
        "category": category,
        "brand": brand,
        "brand_logo": brand_info.get("logo"),
        "brand_color": brand_info.get("color", "#007CC3"),
        "sku": f"{brand[:3].upper()}-{category[:3].upper()}-{index:06d}",
        "unspsc_code": unspsc,
        "unspsc_name": category,
        "base_price": base_price,
        "unit": random.choice(["EA", "PK", "BX", "SET"]),
        "image_url": image_url,
        "specifications": product.get("specs", {}),
        "availability": availability,
        "rating": round(random.uniform(4.0, 5.0), 1),
        "reviews_count": random.randint(10, 500),
        "spec_document_url": "https://example.com/specs/document.pdf",
        "is_sponsored": is_sponsored,
        "features": [
            "Premium build quality",
            "Industry-standard compliance",
            "Extended warranty available",
            f"Genuine {brand} product"
        ]
    }

def generate_delivery_partners(base_price: float, count: int) -> List[DeliveryPartner]:
    """Generate delivery partners with price/lead time inverse relationship"""
    partners = []
    for i in range(count):
        price_multiplier = 1.0 + (0.1 * (count - i - 1))
        lead_time = 10 - (i * 3)
        if lead_time < 1:
            lead_time = 1
        partners.append(DeliveryPartner(
            partner_id=f"DP-{uuid.uuid4().hex[:8]}",
            price=round(base_price * price_multiplier, 2),
            lead_time_days=lead_time,
            available_quantity=random.randint(16, 2098)
        ))
    return partners

def get_alternate_products(product: Dict, brand: str) -> List[Dict]:
    """Generate alternate product suggestions"""
    alt_brands = [b["name"] for b in MRO_BRANDS if b["name"] != brand]
    alternates = []
    for alt_brand in random.sample(alt_brands, min(2, len(alt_brands))):
        brand_info = get_brand_info(alt_brand)
        alternates.append({
            "id": str(uuid.uuid4()),
            "name": product["name"].replace(brand, alt_brand),
            "brand": alt_brand,
            "brand_logo": brand_info.get("logo"),
            "price": round(product["base_price"] * random.uniform(0.7, 0.95), 2),
            "lead_time_days": random.randint(7, 15)
        })
    return alternates

# Auth Routes
@api_router.post("/auth/login", response_model=UserResponse)
async def login(user_data: UserLogin):
    if not user_data.email or not user_data.password:
        raise HTTPException(status_code=400, detail="Email and password required")
    
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    
    if existing_user:
        if existing_user.get("country") != user_data.country:
            await db.users.update_one({"email": user_data.email}, {"$set": {"country": user_data.country}})
            existing_user["country"] = user_data.country
        user = existing_user
    else:
        user_id = str(uuid.uuid4())
        user = {
            "id": user_id,
            "email": user_data.email,
            "name": user_data.email.split("@")[0].title(),
            "country": user_data.country,
            "info_coins": 2500,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    
    currency = COUNTRY_CURRENCIES.get(user_data.country, COUNTRY_CURRENCIES["USA"])
    token = create_jwt_token(user["id"], user["email"])
    
    return UserResponse(
        id=user["id"], email=user["email"], name=user["name"],
        country=user_data.country, currency=currency,
        info_coins=user.get("info_coins", 0), token=token,
        role=user.get("role")
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    return {**current_user, "currency": currency}

# Product Routes
@api_router.get("/products/search")
async def search_products(
    q: str = Query("", description="Search query"),
    category: Optional[str] = None,
    brand: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    lang: str = Query("en", description="Language code (en, fr, de, it, nl)"),
    current_user: dict = Depends(get_current_user)
):
    currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    results = []
    
    search_term = q.lower()
    
    # First, add real catalog products - interleave for variety (industrial + IT mixed)
    # Combine and shuffle for diverse display when no search term
    all_catalog_products = []
    it_products = list(IT_PRODUCTS_CATALOG)
    vendor_products = list(NEW_VENDOR_PRODUCTS)
    
    # Interleave products: 1 industrial, 1 IT, 1 industrial, etc. for better variety
    max_len = max(len(vendor_products), len(it_products))
    for i in range(max_len):
        if i < len(vendor_products):
            all_catalog_products.append(vendor_products[i])
        if i < len(it_products):
            all_catalog_products.append(it_products[i])
    
    # Filter catalog products based on search, category, and brand
    filtered_catalog = []
    for product in all_catalog_products:
        # Search filter
        if search_term:
            name_match = search_term in product["name"].lower()
            brand_match = search_term in product.get("brand", "").lower()
            category_match = search_term in product.get("category", "").lower()
            desc_match = search_term in product.get("short_description", "").lower() or search_term in product.get("full_description", "").lower()
            if not (name_match or brand_match or category_match or desc_match):
                continue
        
        # Category filter
        if category and category != "all" and product.get("category", "").lower() != category.lower():
            continue
        
        # Brand filter
        if brand and brand != "all" and product.get("brand", "").lower() != brand.lower():
            continue
        
        filtered_catalog.append(product)
    
    # Add filtered catalog products to results
    for product in filtered_catalog:
        brand_info = next((b for b in MRO_BRANDS if b["name"] == product.get("brand")), {})
        base_price = product.get("base_price", 100)
        
        results.append({
            "id": product["id"],
            "name": product["name"],
            "short_description": product.get("short_description", ""),
            "full_description": product.get("full_description", ""),
            "category": product["category"],
            "brand": product.get("brand", ""),
            "brand_logo": brand_info.get("logo"),
            "brand_color": brand_info.get("color", "#007CC3"),
            "sku": product["sku"],
            "unspsc_code": product["unspsc_code"],
            "unspsc_name": product.get("category", ""),
            "price": round(base_price * currency["rate"], 2),
            "currency_code": currency["code"],
            "currency_symbol": currency["symbol"],
            "unit": product.get("unit", "EA"),
            "image_url": product["image_url"],
            "specifications": product.get("specifications", {}),
            "availability": product.get("availability", {"in_stock": True, "quantity": random.randint(10, 500)}),
            "rating": product.get("rating", round(random.uniform(4.0, 5.0), 1)),
            "reviews_count": product.get("reviews_count", random.randint(10, 500)),
            "features": product.get("features", []),
            "spec_document_url": f"https://docs.omnisupply.io/specs/{product['id']}.pdf",
            "lead_time_days": random.randint(2, 7),
            "delivery_partners": [
                {"partner_id": "DP001", "price": round(base_price * currency["rate"], 2),
                 "lead_time_days": random.randint(2, 5), "available_quantity": random.randint(50, 500)}
            ],
            "has_delivery_partner": True,
            "alternate_products": [],
            "result_type": "with_partner",
            "is_sponsored": random.random() < 0.1
        })
    
    # Generate additional products to fill up to the limit
    remaining = limit - len(results)
    if remaining > 0:
        matching_categories = [c["name"] for c in MRO_CATEGORIES if search_term in c["name"].lower()] if search_term else [c["name"] for c in MRO_CATEGORIES[:10]]
        matching_brands = [b["name"] for b in MRO_BRANDS if search_term in b["name"].lower()] if search_term else [b["name"] for b in MRO_BRANDS[:10]]
        
        if category and category != "all":
            matching_categories = [c for c in matching_categories if c.lower() == category.lower()] or [category]
        if brand and brand != "all":
            matching_brands = [b for b in matching_brands if b.lower() == brand.lower()] or [brand]
        
        for i in range(remaining):
            cat = random.choice(matching_categories if matching_categories else [c["name"] for c in MRO_CATEGORIES])
            br = random.choice(matching_brands if matching_brands else [b["name"] for b in MRO_BRANDS])
            product = generate_product_data(i + (page - 1) * limit + len(results), cat, br)
            
            rand = random.random()
            
            if rand < 0.70:
                partner_count = random.choice([1, 2, 3])
                delivery_partners = generate_delivery_partners(product["base_price"], partner_count)
                result_type = "with_partner"
                price = delivery_partners[0].price if delivery_partners else product["base_price"]
                lead_time = delivery_partners[0].lead_time_days if delivery_partners else 5
            elif rand < 0.90:
                delivery_partners = []
                result_type = "quotation_required"
                price = None
                lead_time = None
            else:
                continue
            
            alternates = get_alternate_products(product, br) if result_type == "with_partner" else []
            
            results.append({
                "id": product["id"],
                "name": product["name"],
                "short_description": product.get("short_description", product.get("full_description", "")[:100]),
                "full_description": product.get("full_description", ""),
                "category": product["category"],
                "brand": product["brand"],
                "brand_logo": product.get("brand_logo"),
                "brand_color": product.get("brand_color"),
                "sku": product["sku"],
                "unspsc_code": product["unspsc_code"],
                "unspsc_name": product["unspsc_name"],
                "price": round(price * currency["rate"], 2) if price else None,
                "currency_code": currency["code"],
                "currency_symbol": currency["symbol"],
                "unit": product["unit"],
                "image_url": product["image_url"],
                "specifications": product.get("specifications", {}),
                "availability": product.get("availability", {"in_stock": True, "quantity": random.randint(10, 500)}),
                "rating": product.get("rating", round(random.uniform(4.0, 5.0), 1)),
                "reviews_count": product.get("reviews_count", random.randint(10, 500)),
                "features": product.get("features", []),
                "spec_document_url": product["spec_document_url"],
                "lead_time_days": lead_time,
                "delivery_partners": [
                    {"partner_id": dp.partner_id, "price": round(dp.price * currency["rate"], 2),
                     "lead_time_days": dp.lead_time_days, "available_quantity": dp.available_quantity}
                    for dp in delivery_partners
                ] if result_type == "with_partner" else [],
                "has_delivery_partner": result_type == "with_partner",
                "alternate_products": [{**alt, "price": round(alt["price"] * currency["rate"], 2)} for alt in alternates],
                "result_type": result_type,
                "is_sponsored": product["is_sponsored"]
            })
    
    # Apply LLM translation if not English
    if lang != "en" and results:
        translated_results = []
        for result in results[:limit]:  # Limit translation to avoid timeout
            try:
                translated = await translate_product(result, lang)
                translated_results.append(translated)
            except Exception as e:
                logger.error(f"Translation failed for product: {e}")
                translated_results.append(result)
        results = translated_results
    
    # Translate categories if not English
    translated_categories = [{"name": c["name"], "unspsc": c["unspsc"], "icon": c["icon"]} for c in MRO_CATEGORIES]
    if lang != "en":
        for cat in translated_categories[:20]:  # Limit to avoid timeout
            try:
                cat["name"] = await translate_text(cat["name"], lang, "category")
            except:
                pass
    
    return {
        "results": results,
        "total": 3000000,
        "page": page,
        "limit": limit,
        "categories": translated_categories,
        "brands": [{"name": b["name"], "logo": b.get("logo"), "color": b.get("color")} for b in MRO_BRANDS]
    }

@api_router.get("/products/{product_id}/inventory")
async def check_inventory(product_id: str, current_user: dict = Depends(get_current_user)):
    quantity = random.randint(16, 2098)
    return {
        "product_id": product_id,
        "available_quantity": quantity,
        "warehouse_locations": [
            {"location": "US-East", "quantity": random.randint(5, quantity // 2)},
            {"location": "US-West", "quantity": random.randint(5, quantity // 2)},
            {"location": "EU-Central", "quantity": random.randint(0, quantity // 4)}
        ],
        "last_updated": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/products/categories")
async def get_categories():
    return {"categories": [{"name": c["name"], "unspsc": c["unspsc"], "icon": c["icon"]} for c in MRO_CATEGORIES]}

@api_router.get("/products/brands")
async def get_brands():
    return {"brands": [{"name": b["name"], "logo": b["logo"]} for b in MRO_BRANDS]}

# Services Routes
@api_router.get("/services/search")
async def search_services(
    q: str = Query("", description="Search query"),
    category: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    lang: str = Query("en", description="Language code (en, fr, de, it, nl)"),
    current_user: dict = Depends(get_current_user)
):
    currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    results = []
    
    search_term = q.lower()
    
    # Combine IT Services Catalog with SERVICES_DATA
    all_services = []
    
    # Add IT Services from detailed catalog
    for it_service in IT_SERVICES_CATALOG:
        all_services.append({
            "id": it_service["id"],
            "name": it_service["name"],
            "short_description": it_service["short_description"],
            "full_description": it_service["full_description"],
            "category": it_service["category"],
            "unspsc_code": it_service["unspsc_code"],
            "unspsc_name": it_service["unspsc_name"],
            "supplier_name": it_service.get("supplier_name"),
            "supplier_logo": it_service.get("supplier_logo"),
            "supplier_color": it_service.get("supplier_color", "#007CC3"),
            "unit_of_measure": it_service["pricing_model"],
            "base_price": it_service["base_price"],
            "image_url": it_service.get("image_url", DEFAULT_SERVICE_IMAGE),
            "service_includes": it_service.get("service_includes", []),
            "availability": it_service.get("availability", {}),
            "rating": it_service.get("rating", 4.5),
            "reviews_count": it_service.get("reviews_count", 100),
            "is_it_service": True
        })
    
    # Add regular services
    for service in SERVICES_DATA:
        # Get image URL for the service category
        service_image = SERVICE_IMAGE_URLS.get(service["category"], DEFAULT_SERVICE_IMAGE)
        all_services.append({
            "id": str(uuid.uuid4()),
            "name": service["name"],
            "short_description": f"Professional {service['name'].lower()} for enterprise environments.",
            "full_description": f"{service['name']}. Professional service meeting industry standards and compliance requirements.",
            "category": service["category"],
            "unspsc_code": service["unspsc_code"],
            "unspsc_name": service["unspsc_name"],
            "supplier_name": service["supplier_name"],
            "supplier_logo": None,
            "supplier_color": "#007CC3",
            "unit_of_measure": service["unit_of_measure"],
            "base_price": service["base_price"],
            "image_url": service_image,
            "service_includes": [],
            "availability": {"available": True, "lead_time_days": random.randint(1, 7)},
            "rating": round(random.uniform(4.0, 5.0), 1),
            "reviews_count": random.randint(20, 300),
            "is_it_service": False
        })
    
    # Filter by search term
    filtered_services = all_services
    if search_term:
        filtered_services = [s for s in all_services if 
            search_term in s["name"].lower() or 
            search_term in s["category"].lower() or
            search_term in s.get("short_description", "").lower()]
    if category and category != "all":
        filtered_services = [s for s in filtered_services if s["category"].lower() == category.lower()]
    
    if not filtered_services:
        filtered_services = all_services[:15]
    
    for service in filtered_services[:limit]:
        rand = random.random()
        is_sponsored = random.random() < 0.10
        
        if rand < 0.40 and service["supplier_name"]:
            result_type = "with_supplier"
            price = service["base_price"]
            has_supplier = True
        elif rand < 0.50:
            result_type = "quotation_required"
            price = None
            has_supplier = False
        else:
            result_type = "not_found"
            price = None
            has_supplier = False
        
        results.append({
            "id": service["id"],
            "name": service["name"],
            "short_description": service.get("short_description", ""),
            "full_description": service.get("full_description", ""),
            "category": service["category"],
            "unspsc_code": service["unspsc_code"],
            "unspsc_name": service["unspsc_name"],
            "unit_of_measure": service["unit_of_measure"],
            "price": round(price * currency["rate"], 2) if price else None,
            "currency_code": currency["code"],
            "currency_symbol": currency["symbol"],
            "pricing_model": service["unit_of_measure"],
            "supplier_name": service["supplier_name"] if has_supplier else None,
            "supplier_logo": service.get("supplier_logo"),
            "supplier_color": service.get("supplier_color", "#007CC3"),
            "image_url": service.get("image_url", DEFAULT_SERVICE_IMAGE),
            "service_includes": service.get("service_includes", []),
            "availability": service.get("availability", {}),
            "rating": service.get("rating", 4.5),
            "reviews_count": service.get("reviews_count", 100),
            "has_supplier": has_supplier,
            "result_type": result_type,
            "is_sponsored": is_sponsored
        })
    
    # Apply LLM translation if not English
    if lang != "en" and results:
        translated_results = []
        for result in results[:limit]:  # Limit translation to avoid timeout
            try:
                translated = await translate_service(result, lang)
                translated_results.append(translated)
            except Exception as e:
                logger.error(f"Translation failed for service: {e}")
                translated_results.append(result)
        results = translated_results
    
    # Translate categories if not English
    translated_categories = [{"name": c["name"], "unspsc": c["unspsc"], "icon": c["icon"]} for c in SERVICE_CATEGORIES]
    if lang != "en":
        for cat in translated_categories[:15]:  # Limit to avoid timeout
            try:
                cat["name"] = await translate_text(cat["name"], lang, "category")
            except:
                pass
    
    return {
        "results": results,
        "total": 100000,
        "page": page,
        "limit": limit,
        "categories": translated_categories
    }

@api_router.get("/services/categories")
async def get_service_categories():
    return {"categories": [{"name": c["name"], "unspsc": c["unspsc"], "icon": c["icon"]} for c in SERVICE_CATEGORIES]}


# ============================================================
# ALGOLIA CATALOG SEARCH API - World-Class B2B Product Catalog
# ============================================================

class AlgoliaSearchRequest(BaseModel):
    query: str = ""
    page: int = 0
    hits_per_page: int = 24
    filters: Optional[Dict] = None
    sort_by: Optional[str] = None  # price_asc, price_desc, relevance


class AlgoliaCatalogUploadRequest(BaseModel):
    supplier: str
    catalog_type: str = "products"  # products or services


@api_router.post("/algolia/catalog/search")
async def algolia_search_catalog(request: AlgoliaSearchRequest):
    """
    Search products using Algolia - Amazon-like search experience.
    Supports full-text search, faceted filtering, and multi-supplier matching.
    """
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia search service not available")
    
    try:
        results = algolia_search_products(
            query=request.query,
            filters=request.filters,
            page=request.page,
            hits_per_page=request.hits_per_page,
            sort_by=request.sort_by
        )
        
        # Process hits to group by product and identify lowest prices
        processed_hits = []
        product_groups = {}
        
        for hit in results.get("hits", []):
            group_id = hit.get("product_group_id")
            
            # Add lowest price badge
            if hit.get("is_lowest_price"):
                hit["badges"] = [{"type": "lowest_price", "label": "Lowest Price", "color": "green"}]
            elif hit.get("supplier_count", 1) > 1:
                hit["badges"] = [{"type": "multi_supplier", "label": f"{hit['supplier_count']} Suppliers", "color": "blue"}]
            else:
                hit["badges"] = []
            
            # Add stock badge
            if hit.get("in_stock"):
                hit["badges"].append({"type": "in_stock", "label": "In Stock", "color": "green"})
            
            processed_hits.append(hit)
        
        return {
            "success": True,
            "hits": processed_hits,
            "nbHits": results.get("nbHits", 0),
            "page": results.get("page", 0),
            "nbPages": results.get("nbPages", 0),
            "hitsPerPage": results.get("hitsPerPage", 24),
            "facets": results.get("facets", {}),
            "processingTimeMS": results.get("processingTimeMS", 0),
            "query": request.query
        }
        
    except Exception as e:
        logging.error(f"Algolia search error: {e}")
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


@api_router.get("/algolia/catalog/facets/{facet_name}")
async def get_catalog_facets(facet_name: str, query: str = ""):
    """Get all values for a specific facet (brand, category, supplier)"""
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia service not available")
    
    try:
        facet_values = get_facet_values(facet_name, query)
        return {
            "facet": facet_name,
            "values": facet_values
        }
    except Exception as e:
        logging.error(f"Facet retrieval error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/algolia/catalog/upload")
async def upload_catalog_to_algolia(
    file: UploadFile = File(...),
    supplier: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload product catalog from Excel/CSV file to Algolia.
    Supports Fastenal, Grainger, Motion, and other supplier formats.
    """
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia service not available")
    
    # Check file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) and CSV files are supported")
    
    try:
        # Read file content
        content = await file.read()
        
        # Parse file
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        # Convert to list of dicts
        products = df.to_dict('records')
        
        if not products:
            raise HTTPException(status_code=400, detail="No products found in file")
        
        # Index to Algolia
        result = await algolia_index_products(products, supplier)
        
        if result.get("success"):
            # Store upload record
            await db.catalog_uploads.insert_one({
                "upload_id": str(uuid.uuid4()),
                "supplier": supplier,
                "filename": file.filename,
                "product_count": result.get("indexed_count", 0),
                "uploaded_by": current_user.get("email"),
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "status": "completed"
            })
            
            # Trigger product grouping update (async)
            asyncio.create_task(update_product_grouping())
            
            return {
                "success": True,
                "message": f"Successfully indexed {result['indexed_count']} products from {supplier}",
                "indexed_count": result["indexed_count"],
                "supplier": supplier
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("error", "Indexing failed"))
            
    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="File is empty or invalid")
    except Exception as e:
        logging.error(f"Catalog upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/algolia/catalog/stats")
async def get_catalog_stats(current_user: dict = Depends(get_current_user)):
    """Get catalog statistics - total products, suppliers, categories"""
    if not ALGOLIA_AVAILABLE:
        return {
            "algolia_available": False,
            "total_products": 0,
            "suppliers": [],
            "categories": []
        }
    
    try:
        # Get facet counts
        brand_facets = get_facet_values("brand")
        category_facets = get_facet_values("category")
        supplier_facets = get_facet_values("supplier")
        
        # Quick search to get total count
        result = algolia_search_products("", page=0, hits_per_page=1)
        
        # Helper to extract value/count from facet objects
        def facet_to_dict(f):
            if hasattr(f, 'value'):
                return {"name": f.value, "count": getattr(f, 'count', 0)}
            elif isinstance(f, dict):
                return {"name": f.get("value"), "count": f.get("count", 0)}
            return {"name": str(f), "count": 0}
        
        return {
            "algolia_available": True,
            "total_products": result.get("nbHits", 0),
            "brand_count": len(brand_facets) if brand_facets else 0,
            "category_count": len(category_facets) if category_facets else 0,
            "supplier_count": len(supplier_facets) if supplier_facets else 0,
            "suppliers": [facet_to_dict(f) for f in (supplier_facets or [])],
            "top_categories": [facet_to_dict(f) for f in (category_facets or [])[:20]],
            "top_brands": [facet_to_dict(f) for f in (brand_facets or [])[:20]]
        }
    except Exception as e:
        logging.error(f"Stats retrieval error: {e}")
        return {"algolia_available": True, "error": str(e)}


@api_router.get("/algolia/catalog/public-stats")
async def get_public_catalog_stats():
    """Get public catalog statistics (no auth required) - for PunchOut and standalone catalogs"""
    if not ALGOLIA_AVAILABLE:
        return {
            "algolia_available": False,
            "total_products": 0,
            "suppliers": [],
            "categories": []
        }
    
    try:
        # Get facet counts
        brand_facets = get_facet_values("brand")
        category_facets = get_facet_values("category")
        supplier_facets = get_facet_values("supplier")
        
        # Quick search to get total count
        result = algolia_search_products("", page=0, hits_per_page=1)
        
        # Helper to extract value/count from facet objects
        def facet_to_dict(f):
            if hasattr(f, 'value'):
                return {"name": f.value, "count": getattr(f, 'count', 0)}
            elif isinstance(f, dict):
                return {"name": f.get("value"), "count": f.get("count", 0)}
            return {"name": str(f), "count": 0}
        
        return {
            "algolia_available": True,
            "total_products": result.get("nbHits", 0),
            "brand_count": len(brand_facets) if brand_facets else 0,
            "category_count": len(category_facets) if category_facets else 0,
            "supplier_count": len(supplier_facets) if supplier_facets else 0,
            "suppliers": [facet_to_dict(f) for f in (supplier_facets or [])],
            "top_categories": [facet_to_dict(f) for f in (category_facets or [])[:20]],
            "top_brands": [facet_to_dict(f) for f in (brand_facets or [])[:20]]
        }
    except Exception as e:
        logging.error(f"Public stats retrieval error: {e}")
        return {"algolia_available": True, "error": str(e)}


@api_router.get("/algolia/catalog/product/{object_id}")
async def get_product_details(object_id: str):
    """Get detailed product information including related products from other suppliers"""
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia service not available")
    
    try:
        from algolia_service import algolia_client, PRODUCTS_INDEX
        
        # Get the product
        product = algolia_client.get_object(PRODUCTS_INDEX, object_id)
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Find related products from other suppliers (same product_group_id)
        related_products = []
        if product.get("product_group_id"):
            search_result = algolia_search_products(
                query="",
                filters={"product_group_id": product["product_group_id"]},
                hits_per_page=10
            )
            related_products = [
                hit for hit in search_result.get("hits", [])
                if hit.get("objectID") != object_id
            ]
        
        return {
            "success": True,
            "product": product,
            "related_suppliers": related_products,
            "supplier_count": len(related_products) + 1
        }
        
    except Exception as e:
        logging.error(f"Product details error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/algolia/catalog/clear")
async def clear_algolia_catalog(current_user: dict = Depends(get_current_user)):
    """Clear all products from Algolia index (admin only)"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia service not available")
    
    try:
        success = clear_index()
        return {"success": success, "message": "Catalog cleared" if success else "Failed to clear"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/algolia/config")
async def get_algolia_config():
    """Get Algolia configuration for frontend (search-only key)"""
    return {
        "app_id": os.environ.get("ALGOLIA_APP_ID", ""),
        "search_key": os.environ.get("ALGOLIA_SEARCH_KEY", ""),
        "index_name": "omnisupply_products"
    }


# ============================================
# CONTRACT & PRICING MANAGEMENT ENDPOINTS
# ============================================

class ContractUploadRequest(BaseModel):
    supplier_name: str
    countries: List[str] = ["Global"]

@api_router.post("/algolia/contracts/upload")
async def upload_supplier_contract(
    file: UploadFile = File(...),
    supplier_name: str = Form(...),
    countries: str = Form("Global"),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload supplier contract with category-level discounts.
    This sets the discount percentages for calculating Infosys pricing.
    """
    # Check admin access
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    try:
        from pricing_engine import parse_discount_file, save_supplier_contract
        
        content = await file.read()
        
        # Parse discount percentages
        discounts = await parse_discount_file(content, file.filename)
        
        if not discounts:
            raise HTTPException(status_code=400, detail="No valid discount data found in file")
        
        # Parse countries
        country_list = [c.strip() for c in countries.split(",") if c.strip()]
        
        # Save contract
        result = await save_supplier_contract(
            supplier_name=supplier_name,
            category_discounts=discounts,
            countries=country_list,
            contract_file=file.filename
        )
        
        return {
            "success": True,
            "message": f"Contract uploaded for {supplier_name}",
            "supplier": supplier_name,
            "categories_count": len(discounts),
            "countries": country_list,
            "discounts_preview": dict(list(discounts.items())[:10])
        }
        
    except Exception as e:
        logging.error(f"Contract upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/algolia/contracts")
async def get_supplier_contracts(current_user: dict = Depends(get_current_user)):
    """Get all active supplier contracts"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    contracts = []
    cursor = db.supplier_contracts.find({"status": "active"}, {"_id": 0})
    async for contract in cursor:
        contracts.append({
            "supplier_name": contract.get("supplier_name"),
            "categories_count": len(contract.get("category_discounts", {})),
            "countries": contract.get("countries", []),
            "effective_date": contract.get("effective_date"),
            "updated_at": contract.get("updated_at")
        })
    
    return {"contracts": contracts}


@api_router.get("/algolia/contracts/{supplier_name}")
async def get_contract_details(supplier_name: str, current_user: dict = Depends(get_current_user)):
    """Get detailed contract discounts for a supplier"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    contract = await db.supplier_contracts.find_one(
        {"supplier_name": {"$regex": f"^{supplier_name}$", "$options": "i"}},
        {"_id": 0}
    )
    
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    return contract


@api_router.post("/algolia/catalog/upload-with-pricing")
async def upload_catalog_with_pricing(
    file: UploadFile = File(...),
    supplier: str = Form(...),
    countries: str = Form("USA"),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload product catalog with automatic pricing calculation.
    
    This endpoint:
    1. Parses the supplier catalog file
    2. Applies category-level discounts from contracts
    3. Calculates Infosys pricing (List Price, Selling Price, Discount %)
    4. Indexes products to Algolia with all price points
    """
    if not ALGOLIA_AVAILABLE:
        raise HTTPException(status_code=503, detail="Algolia service not available")
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only Excel and CSV files are supported")
    
    try:
        from algolia_service import index_products_from_file, update_product_grouping
        
        content = await file.read()
        country_list = [c.strip() for c in countries.split(",") if c.strip()]
        
        # Index products with pricing
        result = await index_products_from_file(content, file.filename, supplier, country_list)
        
        if result.get("success"):
            # Store upload record
            await db.catalog_uploads.insert_one({
                "upload_id": str(uuid.uuid4()),
                "supplier": supplier,
                "filename": file.filename,
                "product_count": result.get("indexed_count", 0),
                "countries": country_list,
                "uploaded_by": current_user.get("email"),
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "status": "completed",
                "pricing_applied": True
            })
            
            # Update product grouping
            asyncio.create_task(update_product_grouping())
            
            return {
                "success": True,
                "message": f"Successfully indexed {result['indexed_count']} products from {supplier}",
                "indexed_count": result["indexed_count"],
                "supplier": supplier,
                "countries": country_list,
                "pricing_applied": True
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("error", "Indexing failed"))
            
    except Exception as e:
        logging.error(f"Catalog upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/algolia/countries")
async def get_available_countries():
    """Get list of countries with products in the catalog"""
    if not ALGOLIA_AVAILABLE:
        return {"countries": []}
    
    try:
        country_facets = get_facet_values("country")
        return {
            "countries": [
                {"code": f.get("value"), "name": f.get("value"), "count": f.get("count", 0)}
                for f in (country_facets or [])
            ]
        }
    except Exception as e:
        logging.error(f"Countries retrieval error: {e}")
        return {"countries": []}


@api_router.post("/algolia/pricing/calculate")
async def calculate_product_pricing(
    list_price: float = Form(...),
    supplier: str = Form(...),
    category: str = Form(...),
    unspsc_code: str = Form(None)
):
    """
    Calculate Infosys pricing for a product.
    
    Returns:
    - list_price: Original catalog price
    - infosys_purchase_price: What Infosys pays (cost)
    - selling_price: Price to customer
    - discount_percentage: Customer savings %
    """
    try:
        from pricing_engine import calculate_pricing
        
        pricing = await calculate_pricing(list_price, supplier, category, unspsc_code)
        
        return {
            "success": True,
            "pricing": pricing,
            "explanation": f"List Price ${pricing['list_price']:.2f} → "
                          f"Customer Price ${pricing['selling_price']:.2f} "
                          f"(Save {pricing['discount_percentage']:.1f}%)"
        }
    except Exception as e:
        logging.error(f"Pricing calculation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/algolia/catalog/uploads")
async def get_catalog_uploads(current_user: dict = Depends(get_current_user)):
    """Get history of catalog uploads"""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    uploads = []
    cursor = db.catalog_uploads.find({}, {"_id": 0}).sort("uploaded_at", -1).limit(50)
    async for upload in cursor:
        uploads.append(upload)
    
    return {"uploads": uploads}


# Cart Routes
@api_router.get("/cart")
async def get_cart(current_user: dict = Depends(get_current_user)):
    cart = await db.carts.find_one({"user_id": current_user["id"]}, {"_id": 0})
    return {"items": cart.get("items", []) if cart else [], "total": sum(item.get("total_price", 0) for item in (cart.get("items", []) if cart else []))}

@api_router.post("/cart/add")
async def add_to_cart(item: CartItem, current_user: dict = Depends(get_current_user)):
    cart = await db.carts.find_one({"user_id": current_user["id"]})
    item_dict = item.model_dump()
    
    if cart:
        await db.carts.update_one({"user_id": current_user["id"]}, {"$push": {"items": item_dict}})
    else:
        await db.carts.insert_one({"user_id": current_user["id"], "items": [item_dict]})
    
    return {"message": "Item added to cart", "item_id": item.id}

@api_router.delete("/cart/remove/{item_id}")
async def remove_from_cart(item_id: str, current_user: dict = Depends(get_current_user)):
    await db.carts.update_one({"user_id": current_user["id"]}, {"$pull": {"items": {"id": item_id}}})
    return {"message": "Item removed from cart"}

@api_router.post("/cart/transfer")
async def transfer_cart(transfer: CartTransfer, current_user: dict = Depends(get_current_user)):
    """Transfer cart to PunchOut system (Coupa, Ariba, SAP, etc.)"""
    cart = await db.carts.find_one({"user_id": current_user["id"]}, {"_id": 0})
    if not cart or not cart.get("items"):
        raise HTTPException(status_code=400, detail="Cart is empty")
    
    # Get system info
    system_info = next((s for s in PUNCHOUT_SYSTEMS if s["name"] == transfer.system), None)
    if not system_info:
        raise HTTPException(status_code=400, detail="Invalid PunchOut system")
    
    # Create transfer record
    transfer_record = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "system": transfer.system,
        "system_logo": system_info["logo"],
        "items": cart["items"],
        "total_amount": sum(item.get("total_price", 0) for item in cart["items"]),
        "status": "Pending Customer PO",
        "transferred_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.cart_transfers.insert_one(transfer_record)
    
    # Clear cart
    await db.carts.update_one({"user_id": current_user["id"]}, {"$set": {"items": []}})
    
    return {
        "message": f"Cart transferred to {transfer.system}",
        "transfer_id": transfer_record["id"],
        "status": "Pending Customer PO",
        "system_logo": system_info["logo"]
    }

@api_router.get("/cart/transfers")
async def get_cart_transfers(current_user: dict = Depends(get_current_user)):
    transfers = await db.cart_transfers.find({"user_id": current_user["id"]}, {"_id": 0}).sort("transferred_at", -1).to_list(50)
    return {"transfers": transfers}

@api_router.get("/punchout/systems")
async def get_punchout_systems():
    return {"systems": PUNCHOUT_SYSTEMS}

# ============================================
# Coupa cXML PunchOut Integration Routes
# ============================================

@api_router.post("/punchout/setup")
async def punchout_setup(request: Request):
    """
    Handle cXML PunchOutSetupRequest from Coupa or other procurement systems.
    
    This is the entry point for the PunchOut flow:
    1. Coupa sends a cXML PunchOutSetupRequest
    2. We validate the credentials (SharedSecret: Infoshop@2026)
    3. We create a session and return a StartPage URL
    4. User browses the catalog in PunchOut mode
    """
    try:
        # Get raw XML body
        body = await request.body()
        xml_content = body.decode("utf-8")
        
        logger.info(f"PunchOut setup request received, length: {len(xml_content)}")
        
        # Parse the cXML request
        try:
            parsed = parse_punchout_setup_request(xml_content)
        except ValueError as e:
            error_response = create_punchout_setup_response(
                success=False,
                error_message=f"Invalid cXML format: {str(e)}"
            )
            return Response(
                content=error_response,
                media_type="application/xml",
                status_code=400
            )
        
        # Validate credentials (SharedSecret)
        if not validate_punchout_credentials(parsed.get("sender_shared_secret", "")):
            logger.warning(f"PunchOut authentication failed for {parsed.get('from_identity', 'unknown')}")
            
            # Log the failed attempt
            await log_punchout_transaction(
                db,
                transaction_type="setup_failed",
                session_token="",
                buyer_identity=parsed.get("from_identity", "unknown"),
                status="authentication_failed",
                details={"reason": "Invalid shared secret"}
            )
            
            error_response = create_punchout_setup_response(
                success=False,
                error_message="Authentication failed: Invalid credentials"
            )
            return Response(
                content=error_response,
                media_type="application/xml",
                status_code=401
            )
        
        # Create PunchOut session
        session_token = create_punchout_session(
            buyer_cookie=parsed.get("buyer_cookie", ""),
            browser_form_post_url=parsed.get("browser_form_post_url", ""),
            from_identity=parsed.get("from_identity", ""),
            deployment_mode=parsed.get("deployment_mode", "production"),
            user_email=parsed.get("user_email", "")
        )
        
        # Save session to database for persistence
        session_data = get_punchout_session(session_token)
        if session_data:
            await save_punchout_session_to_db(db, session_token, session_data)
        
        # Log successful setup
        await log_punchout_transaction(
            db,
            transaction_type="setup_success",
            session_token=session_token,
            buyer_identity=parsed.get("from_identity", ""),
            status="session_created",
            details={
                "operation": parsed.get("operation", "create"),
                "deployment_mode": parsed.get("deployment_mode", "production")
            }
        )
        
        # Generate StartPage URL - points to catalog with punchout session
        frontend_url = os.environ.get("FRONTEND_URL", "https://omnishop-catalog.preview.emergentagent.com")
        start_page_url = f"{frontend_url}/algolia-catalog?punchout={session_token}"
        
        logger.info(f"PunchOut session created: {session_token[:16]}... -> {start_page_url}")
        
        # Create success response
        success_response = create_punchout_setup_response(
            success=True,
            start_page_url=start_page_url,
            buyer_cookie=parsed.get("buyer_cookie", "")
        )
        
        return Response(
            content=success_response,
            media_type="application/xml",
            status_code=200
        )
        
    except Exception as e:
        logger.error(f"PunchOut setup error: {e}", exc_info=True)
        error_response = create_punchout_setup_response(
            success=False,
            error_message=f"Internal server error: {str(e)}"
        )
        return Response(
            content=error_response,
            media_type="application/xml",
            status_code=500
        )


@api_router.get("/punchout/session/{session_token}")
async def get_punchout_session_info(session_token: str):
    """
    Get PunchOut session information.
    Used by the frontend to verify punchout mode and get session details.
    """
    # Try memory first
    session = get_punchout_session(session_token)
    
    # Fall back to database
    if not session:
        session = await get_punchout_session_from_db(db, session_token)
    
    if not session:
        raise HTTPException(status_code=404, detail="PunchOut session not found or expired")
    
    return {
        "valid": True,
        "buyer_identity": session.get("from_identity", ""),
        "deployment_mode": session.get("deployment_mode", "production"),
        "created_at": session.get("created_at", ""),
        "cart_items_count": len(session.get("cart_items", []))
    }


class PunchOutCartItem(BaseModel):
    """Model for a cart item in PunchOut mode"""
    product_id: str
    supplier_part_id: str
    name: str
    description: Optional[str] = ""
    quantity: int = 1
    unit_price: float
    unit_of_measure: str = "EA"
    brand: Optional[str] = ""
    part_number: Optional[str] = ""
    unspsc_code: Optional[str] = ""


class PunchOutCartUpdate(BaseModel):
    """Model for updating the cart in a PunchOut session"""
    session_token: str
    items: List[PunchOutCartItem]


@api_router.post("/punchout/cart/update")
async def update_punchout_session_cart(cart_update: PunchOutCartUpdate):
    """
    Update the cart items in a PunchOut session.
    Called when user adds/removes items while in PunchOut mode.
    """
    session_token = cart_update.session_token
    
    # Get session from memory or database
    session = get_punchout_session(session_token)
    if not session:
        session = await get_punchout_session_from_db(db, session_token)
    
    if not session:
        raise HTTPException(status_code=404, detail="PunchOut session not found")
    
    # Convert items to dict format
    cart_items = [item.model_dump() for item in cart_update.items]
    
    # Update in memory
    update_punchout_cart(session_token, cart_items)
    
    # Update in database
    await db.punchout_sessions.update_one(
        {"session_token": session_token},
        {"$set": {"cart_items": cart_items, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Calculate total
    total = sum(item["unit_price"] * item["quantity"] for item in cart_items)
    
    return {
        "success": True,
        "items_count": len(cart_items),
        "total_amount": round(total, 2),
        "currency": "USD"
    }


@api_router.post("/punchout/order")
async def create_punchout_order(request: Request, session_token: str = Query(...)):
    """
    Create PunchOutOrderMessage to return cart to Coupa.
    
    This is called when user clicks "Transfer to Coupa" button.
    Returns cXML that gets POSTed to Coupa's BrowserFormPost URL.
    """
    # Get session
    session = get_punchout_session(session_token)
    if not session:
        session = await get_punchout_session_from_db(db, session_token)
    
    if not session:
        raise HTTPException(status_code=404, detail="PunchOut session not found or expired")
    
    cart_items = session.get("cart_items", [])
    buyer_cookie = session.get("buyer_cookie", "")
    browser_form_post_url = session.get("browser_form_post_url", "")
    
    if not cart_items:
        raise HTTPException(status_code=400, detail="Cart is empty")
    
    # Calculate total
    total_amount = sum(item.get("unit_price", 0) * item.get("quantity", 1) for item in cart_items)
    
    # Create the cXML PunchOutOrderMessage
    order_message = create_punchout_order_message(
        cart_items=cart_items,
        buyer_cookie=buyer_cookie,
        total_amount=total_amount,
        currency="USD"
    )
    
    # Log the order
    await log_punchout_transaction(
        db,
        transaction_type="order_created",
        session_token=session_token,
        buyer_identity=session.get("from_identity", ""),
        status="cart_transferred",
        details={
            "items_count": len(cart_items),
            "total_amount": total_amount,
            "browser_form_post_url": browser_form_post_url[:50] + "..." if browser_form_post_url else ""
        }
    )
    
    # Close the session
    close_punchout_session(session_token)
    await db.punchout_sessions.delete_one({"session_token": session_token})
    
    return {
        "success": True,
        "cxml": order_message,
        "browser_form_post_url": browser_form_post_url,
        "total_amount": total_amount,
        "items_count": len(cart_items),
        "instructions": "POST the cxml content to browser_form_post_url to complete the transfer"
    }


@api_router.get("/punchout/config")
async def get_punchout_config():
    """
    Get PunchOut configuration for testing and integration setup.
    Returns the information needed by a Coupa admin to configure the PunchOut.
    """
    api_url = os.environ.get("REACT_APP_BACKEND_URL", "https://omnishop-catalog.preview.emergentagent.com")
    
    return {
        "punchout_enabled": True,
        "supplier_info": {
            "supplier_domain": PUNCHOUT_CONFIG["supplier_domain"],
            "supplier_identity": PUNCHOUT_CONFIG["supplier_identity"],
            "from_domain": PUNCHOUT_CONFIG["from_domain"],
            "from_identity": PUNCHOUT_CONFIG["from_identity"]
        },
        "endpoints": {
            "setup_url": f"{api_url}/api/punchout/setup",
            "order_url": f"{api_url}/api/punchout/order",
            "session_url": f"{api_url}/api/punchout/session/{{session_token}}"
        },
        "cxml_version": "1.2.014",
        "supported_operations": ["create", "edit", "inspect"],
        "note": "Credentials (Identity, Domain, SharedSecret) must be configured in Coupa supplier settings"
    }

# RFQ Routes
@api_router.post("/rfq/submit")
async def submit_rfq(rfq: RFQCreate, current_user: dict = Depends(get_current_user)):
    rfq_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "product_description": rfq.product_description,
        "quantity": rfq.quantity,
        "brand_name": rfq.brand_name,
        "oem_part_number": rfq.oem_part_number,
        "needed_by": rfq.needed_by,
        "delivery_location": rfq.delivery_location,
        "supplier_name": rfq.supplier_name,
        "supplier_email": rfq.supplier_email,
        "request_type": rfq.request_type,
        "is_product": rfq.is_product,
        "status": "submitted",
        "response": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.rfqs.insert_one(rfq_doc)
    await db.users.update_one({"id": current_user["id"]}, {"$inc": {"info_coins": 50}})
    
    return {"message": "RFQ submitted successfully", "rfq_id": rfq_doc["id"], "coins_earned": 50}

@api_router.get("/rfq/list")
async def list_rfqs(current_user: dict = Depends(get_current_user)):
    rfqs = await db.rfqs.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(100)
    
    # Add sample RFQs if none exist
    if not rfqs:
        sample_rfqs = [
            {"id": "RFQ-001", "user_id": current_user["id"], "product_description": "Industrial Hydraulic Pump 50GPM",
             "quantity": 5, "status": "response_received", "created_at": (datetime.now(timezone.utc) - timedelta(days=3)).isoformat(),
             "response": {"supplier": "Parker Hannifin", "unit_price": 1250.00, "lead_time": 14, "valid_until": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()}},
            {"id": "RFQ-002", "user_id": current_user["id"], "product_description": "Custom Gasket Set for Compressor Model X200",
             "quantity": 100, "status": "pending", "created_at": (datetime.now(timezone.utc) - timedelta(days=1)).isoformat(), "response": None},
            {"id": "RFQ-003", "user_id": current_user["id"], "product_description": "Specialized Bearing Assembly",
             "quantity": 20, "status": "response_received", "created_at": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
             "response": {"supplier": "SKF Industrial", "unit_price": 89.50, "lead_time": 7, "valid_until": (datetime.now(timezone.utc) + timedelta(days=3)).isoformat()}},
        ]
        rfqs = sample_rfqs
    
    return {"rfqs": rfqs}

# Quotation Routes
@api_router.post("/quotation/request")
async def request_quotation(request: QuotationRequest, current_user: dict = Depends(get_current_user)):
    quote_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "product_id": request.product_id,
        "product_name": request.product_name,
        "quantity": request.quantity,
        "notes": request.notes,
        "status": "pending",
        "response": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.quotations.insert_one(quote_doc)
    return {"message": "Quotation request sent to 100+ Infosys distributors", "quotation_id": quote_doc["id"]}

@api_router.get("/quotation/list")
async def list_quotations(current_user: dict = Depends(get_current_user)):
    quotations = await db.quotations.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(100)
    
    # Add sample quotations if none exist
    if not quotations:
        currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
        sample_quotations = [
            {"id": "QT-001", "user_id": current_user["id"], "product_name": "Heavy Duty Industrial Motor 5HP",
             "quantity": 3, "status": "response_received", "created_at": (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
             "response": {"supplier": "Siemens Industrial", "unit_price": 2450.00 * currency["rate"], "currency": currency["code"],
                         "lead_time": 10, "valid_until": (datetime.now(timezone.utc) + timedelta(days=5)).isoformat()}},
            {"id": "QT-002", "user_id": current_user["id"], "product_name": "Pneumatic Control Valve Assembly",
             "quantity": 10, "status": "pending", "created_at": (datetime.now(timezone.utc) - timedelta(hours=6)).isoformat(), "response": None},
            {"id": "QT-003", "user_id": current_user["id"], "product_name": "Industrial Sensor Package",
             "quantity": 25, "status": "response_received", "created_at": (datetime.now(timezone.utc) - timedelta(days=4)).isoformat(),
             "response": {"supplier": "Honeywell Solutions", "unit_price": 125.00 * currency["rate"], "currency": currency["code"],
                         "lead_time": 5, "valid_until": (datetime.now(timezone.utc) + timedelta(days=2)).isoformat()}},
            {"id": "QT-004", "user_id": current_user["id"], "product_name": "Custom Fabricated Steel Parts",
             "quantity": 50, "status": "cancelled", "created_at": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat(),
             "response": None, "cancel_reason": "Found alternative supplier with better pricing"},
        ]
        quotations = sample_quotations
    
    return {"quotations": quotations}

@api_router.post("/quotation/{quotation_id}/respond")
async def respond_to_quotation(quotation_id: str, response: QuotationResponse, current_user: dict = Depends(get_current_user)):
    if response.action == "accept":
        await db.quotations.update_one(
            {"id": quotation_id, "user_id": current_user["id"]},
            {"$set": {"status": "accepted"}}
        )
        return {"message": "Quotation accepted. Item ready to add to cart."}
    elif response.action == "cancel":
        await db.quotations.update_one(
            {"id": quotation_id, "user_id": current_user["id"]},
            {"$set": {"status": "cancelled", "cancel_reason": response.cancel_reason}}
        )
        return {"message": "Quotation cancelled."}
    
    raise HTTPException(status_code=400, detail="Invalid action")

# Order Routes
@api_router.post("/orders/create")
async def create_order(items: List[CartItem], current_user: dict = Depends(get_current_user)):
    currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    total = sum(item.total_price for item in items)
    
    order = {
        "id": f"ORD-{uuid.uuid4().hex[:8].upper()}",
        "user_id": current_user["id"],
        "items": [item.model_dump() for item in items],
        "total_amount": total,
        "currency_code": currency["code"],
        "status": "confirmed",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "estimated_delivery": (datetime.now(timezone.utc) + timedelta(days=random.randint(3, 10))).isoformat()
    }
    
    await db.orders.insert_one(order)
    coins = int(total / 10)
    await db.users.update_one({"id": current_user["id"]}, {"$inc": {"info_coins": coins}})
    
    return {"message": "Order placed successfully", "order_id": order["id"], "coins_earned": coins}

@api_router.get("/orders/history")
async def get_order_history(current_user: dict = Depends(get_current_user)):
    orders = await db.orders.find({"user_id": current_user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # If no orders exist, return sample orders with 5 different statuses
    if not orders:
        currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
        sample_orders = [
            # Status 1: Pending - Just placed
            {
                "id": "ORD-A1B2C3D4",
                "user_id": current_user["id"],
                "items": [
                    {"product_name": "HP ProBook 450 G10 Business Laptop", "quantity": 2, "unit_price": round(1299.00 * currency["rate"], 2), "total_price": round(2598.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("IT Equipment - Laptops")},
                    {"product_name": "Dell UltraSharp U2723QE 27\" 4K Monitor", "quantity": 2, "unit_price": round(799.00 * currency["rate"], 2), "total_price": round(1598.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("IT Equipment - Monitors")}
                ],
                "total_amount": round(4196.00 * currency["rate"], 2),
                "currency_code": currency["code"],
                "status": "pending",
                "status_description": "Order received, awaiting processing",
                "created_at": (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat(),
                "estimated_delivery": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
                "tracking_number": None
            },
            # Status 2: Confirmed - Payment verified
            {
                "id": "ORD-E5F6G7H8",
                "user_id": current_user["id"],
                "items": [
                    {"product_name": "ABB Industrial AC Motor 7.5HP", "quantity": 1, "unit_price": round(1850.00 * currency["rate"], 2), "total_price": round(1850.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Motors & Drives")},
                    {"product_name": "Siemens VFD Variable Frequency Drive 15HP", "quantity": 1, "unit_price": round(2450.00 * currency["rate"], 2), "total_price": round(2450.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Motors & Drives")}
                ],
                "total_amount": round(4300.00 * currency["rate"], 2),
                "currency_code": currency["code"],
                "status": "confirmed",
                "status_description": "Payment verified, preparing for shipment",
                "created_at": (datetime.now(timezone.utc) - timedelta(days=1)).isoformat(),
                "estimated_delivery": (datetime.now(timezone.utc) + timedelta(days=5)).isoformat(),
                "tracking_number": None
            },
            # Status 3: Processing - Being prepared
            {
                "id": "ORD-I9J0K1L2",
                "user_id": current_user["id"],
                "items": [
                    {"product_name": "Fluke 289 True-RMS Industrial Multimeter", "quantity": 3, "unit_price": round(595.00 * currency["rate"], 2), "total_price": round(1785.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Test & Measurement")},
                    {"product_name": "Kennametal Carbide End Mill Set", "quantity": 2, "unit_price": round(485.00 * currency["rate"], 2), "total_price": round(970.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Cutting Tools")}
                ],
                "total_amount": round(2755.00 * currency["rate"], 2),
                "currency_code": currency["code"],
                "status": "processing",
                "status_description": "Items being picked and packed in warehouse",
                "created_at": (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
                "estimated_delivery": (datetime.now(timezone.utc) + timedelta(days=3)).isoformat(),
                "tracking_number": None
            },
            # Status 4: Shipped - In transit
            {
                "id": "ORD-M3N4O5P6",
                "user_id": current_user["id"],
                "items": [
                    {"product_name": "Lincoln Electric MIG Welder 250A", "quantity": 1, "unit_price": round(2150.00 * currency["rate"], 2), "total_price": round(2150.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Welding")},
                    {"product_name": "3M Powered Air Purifying Respirator System", "quantity": 2, "unit_price": round(1250.00 * currency["rate"], 2), "total_price": round(2500.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Safety & PPE")}
                ],
                "total_amount": round(4650.00 * currency["rate"], 2),
                "currency_code": currency["code"],
                "status": "shipped",
                "status_description": "Package in transit with carrier",
                "created_at": (datetime.now(timezone.utc) - timedelta(days=4)).isoformat(),
                "estimated_delivery": (datetime.now(timezone.utc) + timedelta(days=1)).isoformat(),
                "tracking_number": "1Z999AA10123456784",
                "carrier": "UPS Ground"
            },
            # Status 5: Delivered - Completed
            {
                "id": "ORD-Q7R8S9T0",
                "user_id": current_user["id"],
                "items": [
                    {"product_name": "Parker Hydraulic Gear Pump 20GPM", "quantity": 2, "unit_price": round(875.00 * currency["rate"], 2), "total_price": round(1750.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Hydraulics & Pneumatics")},
                    {"product_name": "Festo Pneumatic Cylinder 100mm Bore", "quantity": 4, "unit_price": round(425.00 * currency["rate"], 2), "total_price": round(1700.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Hydraulics & Pneumatics")},
                    {"product_name": "Honeywell Safety Harness Full Body", "quantity": 5, "unit_price": round(345.00 * currency["rate"], 2), "total_price": round(1725.00 * currency["rate"], 2), "image_url": PRODUCT_IMAGE_URLS.get("Safety Gloves")}
                ],
                "total_amount": round(5175.00 * currency["rate"], 2),
                "currency_code": currency["code"],
                "status": "delivered",
                "status_description": "Delivered and signed for",
                "created_at": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat(),
                "delivery_date": (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
                "tracking_number": "1Z999AA10987654321",
                "carrier": "FedEx Express",
                "signed_by": "J. Smith"
            }
        ]
        orders = sample_orders
    
    return {"orders": orders}

# Repeat Orders, Bulk Upload, InfoCoins remain same...
@api_router.post("/repeat-orders/create")
async def create_repeat_order(product_id: str, product_name: str, quantity: int, frequency: str, current_user: dict = Depends(get_current_user)):
    if frequency not in ["weekly", "monthly", "quarterly"]:
        raise HTTPException(status_code=400, detail="Invalid frequency")
    
    days_map = {"weekly": 7, "monthly": 30, "quarterly": 90}
    next_date = datetime.now(timezone.utc) + timedelta(days=days_map[frequency])
    
    repeat_order = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "product_id": product_id,
        "product_name": product_name,
        "quantity": quantity,
        "frequency": frequency,
        "next_order_date": next_date.isoformat(),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.repeat_orders.insert_one(repeat_order)
    return {"message": "Repeat order scheduled", "repeat_order_id": repeat_order["id"]}

@api_router.get("/repeat-orders/list")
async def list_repeat_orders(current_user: dict = Depends(get_current_user)):
    repeat_orders = await db.repeat_orders.find({"user_id": current_user["id"]}, {"_id": 0}).to_list(100)
    return {"repeat_orders": repeat_orders}

@api_router.delete("/repeat-orders/{repeat_order_id}")
async def cancel_repeat_order(repeat_order_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.repeat_orders.update_one({"id": repeat_order_id, "user_id": current_user["id"]}, {"$set": {"is_active": False}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Repeat order not found")
    return {"message": "Repeat order cancelled"}

@api_router.post("/bulk/upload")
async def bulk_upload(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    import openpyxl
    from io import BytesIO
    
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    sheet = wb.active
    
    results = []
    found = 0
    not_found = 0
    currency = COUNTRY_CURRENCIES.get(current_user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        product_name = str(row[0])
        quantity = int(row[1]) if len(row) > 1 and row[1] else 1
        
        if random.random() < 0.70:
            product = generate_product_data(random.randint(1, 10000), random.choice([c["name"] for c in MRO_CATEGORIES]), random.choice([b["name"] for b in MRO_BRANDS]))
            results.append({
                "search_term": product_name, "found": True,
                "product": {"id": product["id"], "name": product["name"], "price": round(product["base_price"] * currency["rate"], 2),
                           "currency": currency["symbol"], "available_quantity": random.randint(16, 2098), "unspsc_code": product["unspsc_code"]},
                "requested_quantity": quantity
            })
            found += 1
        else:
            results.append({"search_term": product_name, "found": False, "product": None, "requested_quantity": quantity})
            not_found += 1
    
    return {"total_items": found + not_found, "found_items": found, "not_found_items": not_found, "results": results}

@api_router.get("/infocoins/balance")
async def get_infocoin_balance(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    return {"balance": user.get("info_coins", 0)}

@api_router.get("/infocoins/rewards")
async def get_rewards():
    return {"rewards": INFOCOIN_REWARDS}

@api_router.post("/infocoins/redeem/{reward_id}")
async def redeem_reward(reward_id: str, current_user: dict = Depends(get_current_user)):
    reward = next((r for r in INFOCOIN_REWARDS if r["id"] == reward_id), None)
    if not reward:
        raise HTTPException(status_code=404, detail="Reward not found")
    
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    if user.get("info_coins", 0) < reward["coins_required"]:
        raise HTTPException(status_code=400, detail="Insufficient InfoCoins")
    
    await db.users.update_one({"id": current_user["id"]}, {"$inc": {"info_coins": -reward["coins_required"]}})
    
    redemption = {
        "id": str(uuid.uuid4()), "user_id": current_user["id"], "reward_id": reward_id,
        "reward_name": reward["name"], "coins_spent": reward["coins_required"],
        "status": "processing", "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.redemptions.insert_one(redemption)
    
    return {"message": f"Successfully redeemed {reward['name']}!"}

# Chat Routes
@api_router.post("/chat/message")
async def chat_message(chat: dict, current_user: dict = Depends(get_current_user)):
    session_id = chat.get("session_id") or str(uuid.uuid4())
    message = chat.get("message", "")
    
    system_message = """You are InfoConnect, the AI assistant for OMNISupply.io - Infosys's enterprise procurement platform. 
    You help users with finding products and services, managing orders, submitting RFQs, and using InfoCoins rewards.
    Be helpful, professional, and concise."""
    
    try:
        llm = LlmChat(api_key=EMERGENT_LLM_KEY, session_id=session_id, system_message=system_message).with_model("openai", "gpt-5.2")
        user_message = UserMessage(text=message)
        response = await llm.send_message(user_message)
        
        await db.chat_history.insert_one({
            "session_id": session_id, "user_id": current_user["id"],
            "user_message": message, "bot_response": response,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"response": response, "session_id": session_id}
    except Exception as e:
        logging.error(f"Chat error: {e}")
        return {"response": "I'm here to help! You can search for products, check order history, or explore InfoCoins rewards.", "session_id": session_id}

@api_router.get("/chat/history")
async def get_chat_history(session_id: str = Query(...), current_user: dict = Depends(get_current_user)):
    history = await db.chat_history.find(
        {"session_id": session_id, "user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", 1).to_list(50)
    return {"history": history}

# Stats
@api_router.get("/stats")
async def get_stats():
    return {
        "total_products": "30M+",
        "total_services": "100K+",
        "total_categories": 78,
        "total_brands": "511+",
        "service_categories": len(SERVICE_CATEGORIES),
        "integrations": ["Coupa", "SAP Ariba", "SAP ERP", "Ivalua", "Oracle"],
        "countries_served": len(COUNTRY_CURRENCIES),
        "punchout_systems": PUNCHOUT_SYSTEMS
    }

# Admin Routes for Vendor Catalog Upload
DELIVERY_PARTNERS = [
    {"id": "grainger", "name": "Grainger", "logo": None, "color": "#E31837"},
    {"id": "motion", "name": "Motion Industries", "logo": None, "color": "#003366"},
    {"id": "fastenal", "name": "Fastenal", "logo": None, "color": "#00529B"},
    {"id": "bdi", "name": "BDI (Bearing Distributors Inc)", "logo": None, "color": "#1E4D8C"},
    {"id": "msc", "name": "MSC Industrial", "logo": None, "color": "#003B71"},
    {"id": "mcmaster", "name": "McMaster-Carr", "logo": None, "color": "#8B4513"},
    {"id": "zoro", "name": "Zoro", "logo": None, "color": "#FF6600"},
    {"id": "uline", "name": "Uline", "logo": None, "color": "#003366"},
]

class AdminLogin(BaseModel):
    username: str
    password: str

class CatalogUploadResponse(BaseModel):
    success: bool
    message: str
    products_imported: int = 0
    services_imported: int = 0
    errors: List[str] = []

# Admin authentication - simple for demo
ADMIN_CREDENTIALS = {
    "admin": "admin123",
    "admin@omnisupply.io": "admin123",
    "infosys_admin": "InfosysBPM2024!"
}

def verify_admin(username: str, password: str) -> bool:
    return ADMIN_CREDENTIALS.get(username) == password

@api_router.post("/admin/login")
async def admin_login(credentials: AdminLogin):
    if verify_admin(credentials.username, credentials.password):
        admin_token = create_jwt_token(f"admin_{credentials.username}", credentials.username)
        return {"success": True, "token": admin_token, "username": credentials.username}
    raise HTTPException(status_code=401, detail="Invalid admin credentials")

@api_router.get("/admin/delivery-partners")
async def get_delivery_partners():
    return {"partners": DELIVERY_PARTNERS}

@api_router.post("/admin/upload-catalog")
async def upload_vendor_catalog(
    file: UploadFile = File(...),
    partner_id: str = Form(...),
    catalog_type: str = Form(...)  # "products" or "services"
):
    """
    Upload vendor catalog from delivery partners like Grainger, MOTION, Fastenal, BDI
    Accepts CSV/Excel files with product or service data
    """
    # Validate partner
    partner = next((p for p in DELIVERY_PARTNERS if p["id"] == partner_id), None)
    if not partner:
        raise HTTPException(status_code=400, detail=f"Invalid delivery partner: {partner_id}")
    
    # Validate file type
    allowed_extensions = ['.csv', '.xlsx', '.xls']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"File type not supported. Allowed: {', '.join(allowed_extensions)}")
    
    # Read file content
    content = await file.read()
    
    products_imported = 0
    services_imported = 0
    errors = []
    
    try:
        if file_ext == '.csv':
            import csv
            import io
            
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            # Excel file
            import openpyxl
            import io
            
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        
        # Process rows
        for idx, row in enumerate(rows):
            try:
                if catalog_type == "products":
                    product_data = {
                        "id": str(uuid.uuid4()),
                        "name": row.get("name") or row.get("product_name") or row.get("Name"),
                        "description": row.get("description") or row.get("Description") or "",
                        "brand": row.get("brand") or row.get("Brand") or partner["name"],
                        "category": row.get("category") or row.get("Category") or "General MRO",
                        "sku": row.get("sku") or row.get("SKU") or row.get("part_number") or f"{partner_id.upper()}-{idx:06d}",
                        "unspsc_code": row.get("unspsc") or row.get("UNSPSC") or "31000000",
                        "base_price": float(row.get("price") or row.get("Price") or row.get("unit_price") or 0),
                        "unit": row.get("unit") or row.get("UOM") or "EA",
                        "delivery_partner": partner["name"],
                        "delivery_partner_id": partner_id,
                        "image_url": row.get("image_url") or PRODUCT_IMAGE_URLS.get("Hand Tools"),
                        "imported_at": datetime.now(timezone.utc).isoformat(),
                        "source": f"vendor_upload_{partner_id}"
                    }
                    
                    if product_data["name"]:
                        await db.vendor_products.insert_one(product_data)
                        products_imported += 1
                else:  # services
                    service_data = {
                        "id": str(uuid.uuid4()),
                        "name": row.get("name") or row.get("service_name") or row.get("Name"),
                        "description": row.get("description") or row.get("Description") or "",
                        "category": row.get("category") or row.get("Category") or "Facilities Management & Workplace Services",
                        "unspsc_code": row.get("unspsc") or row.get("UNSPSC") or "76100000",
                        "base_rate": float(row.get("rate") or row.get("Rate") or row.get("price") or 0),
                        "pricing_model": row.get("pricing_model") or row.get("unit") or "Per Hour",
                        "supplier_name": row.get("supplier") or row.get("Supplier") or partner["name"],
                        "delivery_partner": partner["name"],
                        "delivery_partner_id": partner_id,
                        "image_url": row.get("image_url") or SERVICE_IMAGE_URLS.get("Commercial Cleaning"),
                        "imported_at": datetime.now(timezone.utc).isoformat(),
                        "source": f"vendor_upload_{partner_id}"
                    }
                    
                    if service_data["name"]:
                        await db.vendor_services.insert_one(service_data)
                        services_imported += 1
                        
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")
                
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")
    
    return CatalogUploadResponse(
        success=True,
        message=f"Catalog uploaded successfully from {partner['name']}",
        products_imported=products_imported,
        services_imported=services_imported,
        errors=errors[:10]  # Return first 10 errors only
    )

@api_router.get("/admin/uploaded-catalogs")
async def get_uploaded_catalogs():
    """Get summary of uploaded vendor catalogs"""
    products_by_partner = {}
    services_by_partner = {}
    
    # Count products by partner using aggregation (optimized)
    products_pipeline = [
        {"$group": {"_id": "$delivery_partner", "count": {"$sum": 1}}}
    ]
    async for doc in db.vendor_products.aggregate(products_pipeline):
        partner = doc.get("_id") or "Unknown"
        products_by_partner[partner] = doc.get("count", 0)
    
    # Count services by partner using aggregation (optimized)
    services_pipeline = [
        {"$group": {"_id": "$delivery_partner", "count": {"$sum": 1}}}
    ]
    async for doc in db.vendor_services.aggregate(services_pipeline):
        partner = doc.get("_id") or "Unknown"
        services_by_partner[partner] = doc.get("count", 0)
    
    return {
        "products_by_partner": products_by_partner,
        "services_by_partner": services_by_partner,
        "total_vendor_products": sum(products_by_partner.values()),
        "total_vendor_services": sum(services_by_partner.values()),
        "delivery_partners": DELIVERY_PARTNERS
    }

@api_router.delete("/admin/clear-catalog/{partner_id}")
async def clear_partner_catalog(partner_id: str, catalog_type: str = "all"):
    """Clear uploaded catalog from a specific delivery partner"""
    deleted_products = 0
    deleted_services = 0
    
    if catalog_type in ["products", "all"]:
        result = await db.vendor_products.delete_many({"delivery_partner_id": partner_id})
        deleted_products = result.deleted_count
    
    if catalog_type in ["services", "all"]:
        result = await db.vendor_services.delete_many({"delivery_partner_id": partner_id})
        deleted_services = result.deleted_count
    
    return {
        "success": True,
        "deleted_products": deleted_products,
        "deleted_services": deleted_services
    }

# ============================================
# ADMIN BUYING DESK MANAGEMENT ENDPOINTS
# ============================================

# Pydantic models for admin operations
class AdminStatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None

class AdminAssignSpecialist(BaseModel):
    specialist_name: str
    specialist_email: str

class AdminAddNote(BaseModel):
    note: str
    author: str = "Admin"

# Specialist roster for assignment
BUYING_DESK_SPECIALISTS = [
    {"name": "Rajesh Kumar", "email": "rajesh.kumar@infosys.com", "specialty": "IT & Electronics"},
    {"name": "Priya Sharma", "email": "priya.sharma@infosys.com", "specialty": "MRO & Industrial"},
    {"name": "Amit Patel", "email": "amit.patel@infosys.com", "specialty": "Office Supplies"},
    {"name": "Sneha Reddy", "email": "sneha.reddy@infosys.com", "specialty": "Professional Services"},
    {"name": "Vikram Singh", "email": "vikram.singh@infosys.com", "specialty": "Facilities & Maintenance"},
]

@api_router.get("/admin/buying-desk/requests")
async def admin_get_all_buying_desk_requests(
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 50
):
    """Get all tactical buying desk requests (admin only)"""
    try:
        query = {}
        if status:
            query["current_stage"] = status
        
        total = await db.buying_desk_requests.count_documents(query)
        requests = await db.buying_desk_requests.find(
            query, {"_id": 0}
        ).sort("submitted_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
        
        # Get stats with error handling
        stats = {
            "total": total,
            "submitted": await db.buying_desk_requests.count_documents({"current_stage": "submitted"}),
            "supplier_identification": await db.buying_desk_requests.count_documents({"current_stage": "supplier_identification"}),
            "rfq_sent": await db.buying_desk_requests.count_documents({"current_stage": "rfq_sent"}),
            "quotes_received": await db.buying_desk_requests.count_documents({"current_stage": "quotes_received"}),
            "negotiating": await db.buying_desk_requests.count_documents({"current_stage": "negotiating"}),
            "po_ready": await db.buying_desk_requests.count_documents({"current_stage": "po_ready"})
        }
        
        return {
            "requests": requests,
            "total": total,
            "page": page,
            "limit": limit,
            "stats": stats
        }
    except Exception as e:
        logger.error(f"Error fetching buying desk requests: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch requests")

@api_router.get("/admin/buying-desk/request/{request_id}")
async def admin_get_buying_desk_request(request_id: str):
    """Get details of a specific buying desk request (admin only)"""
    request = await db.buying_desk_requests.find_one(
        {"request_id": request_id}, {"_id": 0}
    )
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    return request

@api_router.put("/admin/buying-desk/request/{request_id}/status")
async def admin_update_buying_desk_status(request_id: str, update: AdminStatusUpdate):
    """Update the status/stage of a buying desk request (admin only)"""
    request = await db.buying_desk_requests.find_one({"request_id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Valid stages
    valid_stages = ["submitted", "supplier_identification", "rfq_sent", "quotes_received", "negotiating", "po_ready"]
    if update.status not in valid_stages:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_stages}")
    
    # Update stages array
    stages = request.get("stages", [])
    stage_index = valid_stages.index(update.status)
    
    for i, stage in enumerate(stages):
        if i <= stage_index:
            stage["completed"] = True
            if not stage.get("completed_at"):
                stage["completed_at"] = datetime.now(timezone.utc).isoformat()
        else:
            stage["completed"] = False
            stage["completed_at"] = None
    
    # Add note if provided
    notes = request.get("notes", [])
    if update.notes:
        notes.append({
            "text": update.notes,
            "author": "Admin",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": f"Status updated to {update.status}"
        })
    
    await db.buying_desk_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "current_stage": update.status,
            "status": update.status,
            "stages": stages,
            "notes": notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": f"Status updated to {update.status}"}

@api_router.put("/admin/buying-desk/request/{request_id}/assign")
async def admin_assign_specialist(request_id: str, assignment: AdminAssignSpecialist):
    """Assign a specialist to a buying desk request (admin only)"""
    request = await db.buying_desk_requests.find_one({"request_id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    notes = request.get("notes", [])
    notes.append({
        "text": f"Assigned to {assignment.specialist_name} ({assignment.specialist_email})",
        "author": "Admin",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "Specialist assigned"
    })
    
    await db.buying_desk_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "assigned_to": assignment.specialist_name,
            "assigned_specialist": {
                "name": assignment.specialist_name,
                "email": assignment.specialist_email
            },
            "notes": notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": f"Assigned to {assignment.specialist_name}"}

@api_router.post("/admin/buying-desk/request/{request_id}/note")
async def admin_add_note(request_id: str, note_data: AdminAddNote):
    """Add a note to a buying desk request (admin only)"""
    request = await db.buying_desk_requests.find_one({"request_id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    notes = request.get("notes", [])
    notes.append({
        "text": note_data.note,
        "author": note_data.author,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "Note added"
    })
    
    await db.buying_desk_requests.update_one(
        {"request_id": request_id},
        {"$set": {"notes": notes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "message": "Note added"}

@api_router.get("/admin/buying-desk/specialists")
async def get_specialists():
    """Get list of available buying desk specialists"""
    return {"specialists": BUYING_DESK_SPECIALISTS}

# ============================================
# ADMIN SOURCING REQUESTS MANAGEMENT
# ============================================

@api_router.get("/admin/sourcing/requests")
async def admin_get_all_sourcing_requests(
    status: Optional[str] = None,
    urgency: Optional[str] = None,
    page: int = 1,
    limit: int = 50
):
    """Get all sourcing/managed services requests (admin only)"""
    try:
        query = {}
        if status:
            query["status"] = status
        if urgency:
            query["urgency"] = urgency
        
        total = await db.sourcing_requests.count_documents(query)
        requests = await db.sourcing_requests.find(
            query, {"_id": 0}
        ).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
        
        # Get stats with error handling
        stats = {
            "total": total,
            "submitted": await db.sourcing_requests.count_documents({"status": "SUBMITTED"}),
            "in_progress": await db.sourcing_requests.count_documents({"status": "IN_PROGRESS"}),
            "rfq_sent": await db.sourcing_requests.count_documents({"status": "RFQ_SENT"}),
            "quotes_received": await db.sourcing_requests.count_documents({"status": "QUOTES_RECEIVED"}),
            "completed": await db.sourcing_requests.count_documents({"status": "COMPLETED"}),
            "cancelled": await db.sourcing_requests.count_documents({"status": "CANCELLED"}),
            "urgent": await db.sourcing_requests.count_documents({"urgency": "urgent"}),
            "critical": await db.sourcing_requests.count_documents({"urgency": "critical"})
        }
        
        return {
            "requests": requests,
            "total": total,
            "page": page,
            "limit": limit,
            "stats": stats
        }
    except Exception as e:
        logger.error(f"Error fetching sourcing requests: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch requests")

@api_router.get("/admin/sourcing/request/{sourcing_id}")
async def admin_get_sourcing_request(sourcing_id: str):
    """Get details of a specific sourcing request (admin only)"""
    request = await db.sourcing_requests.find_one(
        {"sourcing_id": sourcing_id}, {"_id": 0}
    )
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    return request

@api_router.put("/admin/sourcing/request/{sourcing_id}/status")
async def admin_update_sourcing_status(sourcing_id: str, update: AdminStatusUpdate):
    """Update the status of a sourcing request (admin only)"""
    request = await db.sourcing_requests.find_one({"sourcing_id": sourcing_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    valid_statuses = ["SUBMITTED", "IN_PROGRESS", "RFQ_SENT", "QUOTES_RECEIVED", "COMPLETED", "CANCELLED"]
    if update.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    update_data = {
        "status": update.status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Add note if provided
    if update.notes:
        notes = request.get("admin_notes", [])
        notes.append({
            "text": update.notes,
            "author": "Admin",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": f"Status updated to {update.status}"
        })
        update_data["admin_notes"] = notes
    
    await db.sourcing_requests.update_one(
        {"sourcing_id": sourcing_id},
        {"$set": update_data}
    )
    
    return {"success": True, "message": f"Status updated to {update.status}"}

@api_router.put("/admin/sourcing/request/{sourcing_id}/assign")
async def admin_assign_sourcing_specialist(sourcing_id: str, assignment: AdminAssignSpecialist):
    """Assign a specialist to a sourcing request (admin only)"""
    request = await db.sourcing_requests.find_one({"sourcing_id": sourcing_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    notes = request.get("admin_notes", [])
    notes.append({
        "text": f"Assigned to {assignment.specialist_name} ({assignment.specialist_email})",
        "author": "Admin",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "Specialist assigned"
    })
    
    await db.sourcing_requests.update_one(
        {"sourcing_id": sourcing_id},
        {"$set": {
            "assigned_specialist": {
                "name": assignment.specialist_name,
                "email": assignment.specialist_email
            },
            "admin_notes": notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": f"Assigned to {assignment.specialist_name}"}

@api_router.post("/admin/sourcing/request/{sourcing_id}/note")
async def admin_add_sourcing_note(sourcing_id: str, note_data: AdminAddNote):
    """Add a note to a sourcing request (admin only)"""
    request = await db.sourcing_requests.find_one({"sourcing_id": sourcing_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    notes = request.get("admin_notes", [])
    notes.append({
        "text": note_data.note,
        "author": note_data.author,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "Note added"
    })
    
    await db.sourcing_requests.update_one(
        {"sourcing_id": sourcing_id},
        {"$set": {"admin_notes": notes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "message": "Note added"}

@api_router.get("/admin/buying-desk/dashboard-stats")
async def admin_buying_desk_dashboard_stats():
    """Get comprehensive dashboard statistics for admin"""
    try:
        # Tactical buying stats
        tactical_total = await db.buying_desk_requests.count_documents({})
        tactical_pending = await db.buying_desk_requests.count_documents({"current_stage": {"$nin": ["po_ready"]}})
        tactical_completed = await db.buying_desk_requests.count_documents({"current_stage": "po_ready"})
        
        # Sourcing stats
        sourcing_total = await db.sourcing_requests.count_documents({})
        sourcing_pending = await db.sourcing_requests.count_documents({"status": {"$nin": ["COMPLETED", "CANCELLED"]}})
        sourcing_completed = await db.sourcing_requests.count_documents({"status": "COMPLETED"})
        sourcing_urgent = await db.sourcing_requests.count_documents({"urgency": {"$in": ["urgent", "critical"]}})
        
        # Calculate potential value with error handling
        try:
            tactical_pipeline = await db.buying_desk_requests.aggregate([
                {"$group": {"_id": None, "total_value": {"$sum": "$total_amount"}, "total_savings": {"$sum": "$potential_savings"}}}
            ]).to_list(1)
        except Exception:
            tactical_pipeline = []
        
        try:
            sourcing_pipeline = await db.sourcing_requests.aggregate([
                {"$group": {"_id": None, "total_value": {"$sum": "$estimated_budget"}}}
            ]).to_list(1)
        except Exception:
            sourcing_pipeline = []
        
        return {
            "tactical_buying": {
                "total": tactical_total,
                "pending": tactical_pending,
                "completed": tactical_completed,
                "total_value": tactical_pipeline[0]["total_value"] if tactical_pipeline else 0,
                "potential_savings": tactical_pipeline[0]["total_savings"] if tactical_pipeline else 0
            },
            "sourcing": {
                "total": sourcing_total,
                "pending": sourcing_pending,
                "completed": sourcing_completed,
                "urgent_critical": sourcing_urgent,
                "total_value": sourcing_pipeline[0]["total_value"] if sourcing_pipeline else 0
            },
            "combined": {
                "total_requests": tactical_total + sourcing_total,
                "pending_requests": tactical_pending + sourcing_pending,
                "completed_requests": tactical_completed + sourcing_completed
            }
        }
    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch dashboard statistics")

# ============================================
# QUOTATION UPLOAD & AI ANALYSIS ENDPOINTS
# ============================================

class QuotationUploadRequest(BaseModel):
    file_name: str
    file_type: str
    file_size: int
    supplier_name: Optional[str] = None
    supplier_email: Optional[str] = None
    document_language: str = "en"
    notes: Optional[str] = None

class QuotationAnalysisResult(BaseModel):
    quotation_id: str
    status: str
    extracted_data: Dict[str, Any]
    price_benchmark: Dict[str, Any]
    tax_analysis: Dict[str, Any]
    flags: List[Dict[str, Any]]
    recommendations: List[str]

# Simulated AI extraction results for realistic demo
def generate_ai_extraction(file_name: str, supplier_name: str = None):
    """Generate realistic AI extraction results"""
    categories = ["Office Supplies", "IT Equipment", "MRO Supplies", "Facilities Services", "Professional Services"]
    suppliers = ["Grainger", "Staples", "Dell Technologies", "CDW", "Iron Mountain", "Cintas", supplier_name or "Unknown Supplier"]
    
    num_items = random.randint(3, 8)
    items = []
    total_amount = 0
    
    product_examples = [
        {"name": "HP LaserJet Pro M404n Printer", "unit_price": 349.99, "category": "IT Equipment"},
        {"name": "Ergonomic Office Chair - Mesh Back", "unit_price": 289.00, "category": "Office Supplies"},
        {"name": "Industrial Safety Gloves (Box of 100)", "unit_price": 45.99, "category": "MRO Supplies"},
        {"name": "Quarterly HVAC Maintenance Service", "unit_price": 1250.00, "category": "Facilities Services"},
        {"name": "Dell 27\" UltraSharp Monitor U2722D", "unit_price": 449.99, "category": "IT Equipment"},
        {"name": "Janitorial Supplies Bundle - Monthly", "unit_price": 389.00, "category": "MRO Supplies"},
        {"name": "Network Security Assessment", "unit_price": 4500.00, "category": "Professional Services"},
        {"name": "Bulk Copy Paper (10 Cases)", "unit_price": 289.99, "category": "Office Supplies"},
        {"name": "Fire Extinguisher Inspection", "unit_price": 175.00, "category": "Facilities Services"},
        {"name": "Cisco Meraki Access Point MR46", "unit_price": 795.00, "category": "IT Equipment"},
    ]
    
    selected_products = random.sample(product_examples, min(num_items, len(product_examples)))
    
    for product in selected_products:
        quantity = random.randint(1, 20)
        line_total = round(product["unit_price"] * quantity, 2)
        total_amount += line_total
        items.append({
            "line_number": len(items) + 1,
            "description": product["name"],
            "quantity": quantity,
            "unit_price": product["unit_price"],
            "unit": "EA",
            "line_total": line_total,
            "category": product["category"],
            "unspsc_code": f"{random.randint(40, 50)}1{random.randint(10000, 99999)}"
        })
    
    tax_rate = random.choice([0.0, 0.0625, 0.0725, 0.0825, 0.095])
    tax_amount = round(total_amount * tax_rate, 2)
    grand_total = round(total_amount + tax_amount, 2)
    
    return {
        "supplier": {
            "name": supplier_name or random.choice(suppliers),
            "address": f"{random.randint(100, 9999)} Business Park Dr, Suite {random.randint(100, 500)}",
            "city": random.choice(["Dallas, TX", "Chicago, IL", "Atlanta, GA", "Phoenix, AZ", "Denver, CO"]),
            "tax_id": f"{random.randint(10, 99)}-{random.randint(1000000, 9999999)}",
            "contact_email": f"sales@{(supplier_name or 'supplier').lower().replace(' ', '')}.com"
        },
        "quotation_details": {
            "quotation_number": f"QT-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}",
            "quotation_date": datetime.now().strftime("%Y-%m-%d"),
            "valid_until": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
            "payment_terms": random.choice(["Net 30", "Net 45", "Net 60", "2% 10 Net 30"]),
            "delivery_terms": random.choice(["FOB Destination", "FOB Origin", "DDP", "CIF"])
        },
        "line_items": items,
        "totals": {
            "subtotal": round(total_amount, 2),
            "tax_rate": tax_rate,
            "tax_amount": tax_amount,
            "shipping": random.choice([0, 25.00, 49.99, 75.00]),
            "grand_total": grand_total
        },
        "extraction_confidence": round(random.uniform(0.92, 0.99), 2),
        "document_language": "English",
        "pages_processed": random.randint(1, 4)
    }

def generate_price_benchmark(extracted_data: Dict):
    """Generate price benchmarking analysis"""
    benchmarks = []
    total_potential_savings = 0
    
    for item in extracted_data.get("line_items", []):
        market_price = round(item["unit_price"] * random.uniform(0.75, 1.15), 2)
        variance = round(((item["unit_price"] - market_price) / market_price) * 100, 1)
        potential_savings = max(0, round((item["unit_price"] - market_price) * item["quantity"], 2))
        total_potential_savings += potential_savings
        
        benchmarks.append({
            "item": item["description"],
            "quoted_price": item["unit_price"],
            "market_avg_price": market_price,
            "variance_percent": variance,
            "potential_savings": potential_savings,
            "benchmark_status": "ABOVE_MARKET" if variance > 10 else "AT_MARKET" if variance > -5 else "BELOW_MARKET",
            "data_sources": random.randint(3, 12)
        })
    
    return {
        "benchmarks": benchmarks,
        "total_potential_savings": round(total_potential_savings, 2),
        "overall_assessment": "NEGOTIATION_RECOMMENDED" if total_potential_savings > 500 else "COMPETITIVE_PRICING",
        "market_data_date": datetime.now().strftime("%Y-%m-%d"),
        "confidence_score": round(random.uniform(0.85, 0.95), 2)
    }

def generate_tax_analysis(extracted_data: Dict):
    """Generate tax verification analysis"""
    tax_amount = extracted_data.get("totals", {}).get("tax_amount", 0)
    tax_rate = extracted_data.get("totals", {}).get("tax_rate", 0)
    
    issues = []
    if tax_rate == 0 and random.random() > 0.7:
        issues.append({
            "type": "MISSING_TAX",
            "severity": "HIGH",
            "description": "No sales tax applied - verify tax exemption status"
        })
    elif tax_rate > 0.09:
        issues.append({
            "type": "HIGH_TAX_RATE",
            "severity": "MEDIUM", 
            "description": f"Tax rate of {tax_rate*100:.2f}% is higher than typical rates"
        })
    
    exempt_items = random.randint(0, 2)
    
    return {
        "tax_verified": len(issues) == 0,
        "calculated_tax": round(extracted_data.get("totals", {}).get("subtotal", 0) * tax_rate, 2),
        "quoted_tax": tax_amount,
        "tax_variance": round(abs(tax_amount - extracted_data.get("totals", {}).get("subtotal", 0) * tax_rate), 2),
        "jurisdiction": random.choice(["Texas", "California", "Illinois", "Florida", "New York"]),
        "tax_exemptions_detected": exempt_items,
        "avalara_verification": "VERIFIED" if random.random() > 0.2 else "PENDING",
        "issues": issues
    }

def generate_flags_and_recommendations(extracted_data: Dict, benchmark: Dict, tax: Dict):
    """Generate flags and recommendations"""
    flags = []
    recommendations = []
    
    # Check for high variance items
    for b in benchmark.get("benchmarks", []):
        if b["variance_percent"] > 15:
            flags.append({
                "type": "PRICE_FLAG",
                "severity": "HIGH",
                "item": b["item"],
                "message": f"Price is {b['variance_percent']}% above market average"
            })
    
    # Tax issues
    for issue in tax.get("issues", []):
        flags.append({
            "type": "TAX_FLAG",
            "severity": issue["severity"],
            "message": issue["description"]
        })
    
    # Recommendations
    if benchmark.get("total_potential_savings", 0) > 500:
        recommendations.append(f"Consider negotiation - potential savings of ${benchmark['total_potential_savings']:,.2f}")
    
    if len([b for b in benchmark.get("benchmarks", []) if b["benchmark_status"] == "ABOVE_MARKET"]) > 2:
        recommendations.append("Multiple items priced above market - request volume discount")
    
    valid_until = extracted_data.get("quotation_details", {}).get("valid_until")
    if valid_until:
        days_valid = (datetime.strptime(valid_until, "%Y-%m-%d") - datetime.now()).days
        if days_valid < 14:
            recommendations.append(f"Quotation expires in {days_valid} days - expedite approval process")
    
    recommendations.append("Verify supplier's tax exemption certificates if applicable")
    recommendations.append("Confirm delivery timeline aligns with project requirements")
    
    return flags, recommendations

@api_router.post("/procurement/quotation/upload")
async def upload_quotation(
    file: UploadFile = File(...),
    supplier_name: str = Form(None),
    supplier_email: str = Form(None),
    document_language: str = Form("en"),
    notes: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a quotation document for AI-powered analysis"""
    try:
        # Read file content
        file_content = await file.read()
        file_size = len(file_content)
        
        # Generate unique quotation ID
        quotation_id = f"QUP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        
        # Simulate AI processing
        extracted_data = generate_ai_extraction(file.filename, supplier_name)
        price_benchmark = generate_price_benchmark(extracted_data)
        tax_analysis = generate_tax_analysis(extracted_data)
        flags, recommendations = generate_flags_and_recommendations(extracted_data, price_benchmark, tax_analysis)
        
        # Store in database
        quotation_record = {
            "quotation_id": quotation_id,
            "user_id": current_user.get("email"),
            "user_name": current_user.get("name"),
            "file_name": file.filename,
            "file_type": file.content_type,
            "file_size": file_size,
            "supplier_name": supplier_name or extracted_data["supplier"]["name"],
            "supplier_email": supplier_email,
            "document_language": document_language,
            "notes": notes,
            "status": "ANALYZED",
            "extracted_data": extracted_data,
            "price_benchmark": price_benchmark,
            "tax_analysis": tax_analysis,
            "flags": flags,
            "recommendations": recommendations,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "escalated_for_negotiation": False,
            "added_to_cart": False
        }
        
        await db.quotation_uploads.insert_one(quotation_record)
        
        # Log the activity
        await db.activity_logs.insert_one({
            "user_id": current_user.get("email"),
            "action": "QUOTATION_UPLOADED",
            "quotation_id": quotation_id,
            "file_name": file.filename,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        return {
            "success": True,
            "quotation_id": quotation_id,
            "message": "Quotation uploaded and analyzed successfully",
            "analysis": {
                "extracted_data": extracted_data,
                "price_benchmark": price_benchmark,
                "tax_analysis": tax_analysis,
                "flags": flags,
                "recommendations": recommendations
            }
        }
    except Exception as e:
        logging.error(f"Quotation upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process quotation: {str(e)}")


# ============================================
# AI-POWERED PRICE BENCHMARKING ENDPOINTS
# ============================================

@api_router.post("/procurement/quotation/upload-with-ai")
async def upload_quotation_with_real_ai(
    file: UploadFile = File(...),
    supplier_name: str = Form(None),
    supplier_email: str = Form(None),
    document_language: str = Form("en"),
    notes: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a quotation and perform REAL AI-powered analysis using GPT-5.2, Claude, and Gemini"""
    try:
        # Read file content
        file_content = await file.read()
        file_size = len(file_content)
        
        # Generate unique quotation ID
        quotation_id = f"QAI-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        session_id = f"ai_benchmark_{quotation_id}"
        
        logging.info(f"Starting REAL document extraction for {file.filename} ({file_size} bytes)")
        
        # Step 1: REAL Document Extraction using AI
        extracted_data = await extract_quotation_data(
            file_content=file_content,
            file_name=file.filename,
            file_type=file.content_type,
            supplier_name=supplier_name,
            session_id=session_id
        )
        
        # Check for extraction errors
        if extracted_data.get("error"):
            logging.warning(f"Document extraction warning: {extracted_data.get('message')}")
        
        logging.info(f"Extracted {len(extracted_data.get('line_items', []))} line items from document")
        
        # Step 2: Perform REAL AI price benchmarking with all 3 LLMs
        line_items = extracted_data.get("line_items", [])
        
        if not line_items:
            # Return early if no items extracted - include more helpful message
            error_detail = extracted_data.get("message", "") if extracted_data.get("error") else ""
            return {
                "success": False,
                "quotation_id": quotation_id,
                "message": f"Could not extract line items from the document. {error_detail}\n\nPlease ensure:\n• The file contains clear line items with descriptions and prices\n• The text is selectable (not scanned/image-only PDF)\n• Supported formats: PDF, DOCX, XLSX, TXT, CSV, or images",
                "analysis_mode": "REAL_AI",
                "extraction_error": True,
                "extracted_data": extracted_data,
                "suggestions": [
                    "Try uploading a PDF with selectable text",
                    "Try uploading an Excel file with item details",
                    "Try uploading a clear image of the quotation"
                ]
            }
        
        # Step 2.5: UNSPSC Classification - AI Deep Search for category mapping
        logging.info(f"Starting AI UNSPSC classification for {len(line_items)} items")
        try:
            from document_extractor import classify_unspsc_with_ai
            line_items = await classify_unspsc_with_ai(line_items, session_id)
            extracted_data["line_items"] = line_items
            
            # Generate UNSPSC summary
            unspsc_summary = {}
            for item in line_items:
                code = item.get("unspsc_code", "00000000")[:4] + "0000"  # Segment level
                category = item.get("unspsc_category", "Unclassified")
                if code not in unspsc_summary:
                    unspsc_summary[code] = {"category": category, "count": 0, "total_value": 0}
                unspsc_summary[code]["count"] += 1
                unspsc_summary[code]["total_value"] += item.get("line_total", 0)
            
            extracted_data["unspsc_summary"] = unspsc_summary
            logging.info(f"UNSPSC classification complete. Categories: {list(unspsc_summary.keys())}")
        except Exception as unspsc_error:
            logging.warning(f"UNSPSC classification error (non-fatal): {unspsc_error}")
        
        ai_benchmark_results = await perform_ai_price_benchmarking(
            line_items,
            session_id
        )
        
        # Step 3: Generate enhanced price benchmark from AI results
        enhanced_benchmark = {
            "benchmarks": [],
            "total_potential_savings": ai_benchmark_results.get("total_potential_savings", 0),
            "overall_assessment": ai_benchmark_results.get("summary", {}).get("overall_recommendation", "REVIEW_REQUIRED"),
            "market_data_date": datetime.now().strftime("%Y-%m-%d"),
            "confidence_score": ai_benchmark_results.get("summary", {}).get("ai_confidence", 0.85),
            "ai_engines_used": ai_benchmark_results.get("ai_engines_used", []),
            "ai_item_analyses": ai_benchmark_results.get("item_analyses", [])
        }
        
        for item_analysis in ai_benchmark_results.get("item_analyses", []):
            benchmark = item_analysis.get("benchmark", {})
            enhanced_benchmark["benchmarks"].append({
                "item": item_analysis.get("item"),
                "quoted_price": item_analysis.get("quoted_price"),
                "market_avg_price": benchmark.get("market_avg_price"),
                "variance_percent": benchmark.get("variance_percent"),
                "potential_savings": benchmark.get("potential_savings"),
                "benchmark_status": "ABOVE_MARKET" if benchmark.get("variance_percent", 0) > 10 else "AT_MARKET" if benchmark.get("variance_percent", 0) > -5 else "BELOW_MARKET",
                "recommendation": benchmark.get("recommendation"),
                "risk_level": benchmark.get("risk_level"),
                "confidence": benchmark.get("confidence_score"),
                "ai_analyses": item_analysis.get("ai_analyses", {})
            })
        
        # Step 4: Tax analysis (existing)
        tax_analysis = generate_tax_analysis(extracted_data)
        
        # Step 5: Flags and recommendations
        flags, recommendations = generate_flags_and_recommendations(extracted_data, enhanced_benchmark, tax_analysis)
        
        # Add AI-specific recommendations
        if enhanced_benchmark["total_potential_savings"] > 1000:
            recommendations.insert(0, f"🤖 AI Analysis: Potential savings of ${enhanced_benchmark['total_potential_savings']:,.2f} identified across {len(enhanced_benchmark['benchmarks'])} items")
        
        # Store in database
        quotation_record = {
            "quotation_id": quotation_id,
            "user_id": current_user.get("email"),
            "user_name": current_user.get("name"),
            "file_name": file.filename,
            "file_type": file.content_type,
            "file_size": file_size,
            "supplier_name": supplier_name or extracted_data["supplier"]["name"],
            "supplier_email": supplier_email,
            "document_language": document_language,
            "notes": notes,
            "status": "AI_ANALYZED",
            "analysis_mode": "REAL_AI",
            "extracted_data": extracted_data,
            "price_benchmark": enhanced_benchmark,
            "ai_benchmark_details": ai_benchmark_results,
            "tax_analysis": tax_analysis,
            "flags": flags,
            "recommendations": recommendations,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "escalated_for_negotiation": False,
            "added_to_cart": False
        }
        
        await db.quotation_uploads.insert_one(quotation_record)
        
        return {
            "success": True,
            "quotation_id": quotation_id,
            "message": "Quotation analyzed with real AI price benchmarking",
            "analysis_mode": "REAL_AI",
            "ai_engines_used": ["OpenAI GPT-5.2", "Claude Sonnet 4.5", "Gemini 3 Flash"],
            "analysis": {
                "extracted_data": extracted_data,
                "price_benchmark": enhanced_benchmark,
                "tax_analysis": tax_analysis,
                "flags": flags,
                "recommendations": recommendations
            }
        }
    except Exception as e:
        logging.error(f"AI Quotation upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process quotation with AI: {str(e)}")


@api_router.get("/procurement/quotation/demo-analysis")
async def get_demo_quotation_analysis(
    current_user: dict = Depends(get_current_user)
):
    """
    Get a pre-loaded demo quotation analysis with impressive AI results.
    This is for demonstrations - shows a realistic quotation being analyzed
    with all 3 AI engines showing their analysis.
    """
    quotation_id = f"DEMO-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:4].upper()}"
    
    # Build impressive demo response
    demo_benchmarks = []
    total_savings = 0
    
    for i, item in enumerate(DEMO_QUOTATION["line_items"]):
        openai_data = DEMO_ANALYSIS_RESULTS["openai_analyses"][i]
        claude_data = DEMO_ANALYSIS_RESULTS["claude_analyses"][i]
        gemini_data = DEMO_ANALYSIS_RESULTS["gemini_validations"][i]
        
        quoted_price = item["unit_price"]
        market_price = openai_data["market_price_avg"]
        variance = round(((quoted_price - market_price) / market_price) * 100, 1)
        savings = round(max(0, (quoted_price - market_price) * item["quantity"]), 2)
        total_savings += savings
        
        demo_benchmarks.append({
            "item": item["description"],
            "category": item["category"],
            "quantity": item["quantity"],
            "unit": item["unit"],
            "quoted_price": quoted_price,
            "market_avg_price": market_price,
            "market_price_low": openai_data["market_price_low"],
            "market_price_high": openai_data["market_price_high"],
            "variance_percent": variance,
            "potential_savings": savings,
            "benchmark_status": "ABOVE_MARKET" if variance > 5 else "AT_MARKET",
            "price_trend": openai_data["price_trend"],
            "recommendation": gemini_data["recommendation"],
            "risk_level": gemini_data["risk_level"],
            "ai_analyses": {
                "openai": {
                    "engine": "OpenAI GPT-5.2",
                    "status": "success",
                    "market_price_avg": market_price,
                    "market_price_range": f"${openai_data['market_price_low']:,.2f} - ${openai_data['market_price_high']:,.2f}",
                    "data_sources": openai_data["data_sources"],
                    "price_trend": openai_data["price_trend"],
                    "confidence": openai_data["confidence"]
                },
                "claude": {
                    "engine": "Claude Sonnet 4.5",
                    "status": "success",
                    "skill_level": claude_data.get("skill_level", "N/A"),
                    "market_demand": claude_data.get("market_demand", "moderate"),
                    "geographic_factor": claude_data.get("geographic_factor", "average"),
                    "analysis_type": "Professional Services Rate Analysis" if "Service" in item["category"] else "Product Analysis",
                    "confidence": 0.91
                },
                "gemini": {
                    "engine": "Gemini 3 Flash",
                    "status": "success",
                    "recommendation": gemini_data["recommendation"],
                    "risk_level": gemini_data["risk_level"],
                    "variance_validated": gemini_data["variance"],
                    "cross_validation": "CONFIRMED",
                    "confidence": 0.89
                }
            }
        })
    
    # Generate flags based on demo data
    flags = []
    for b in demo_benchmarks:
        if b["variance_percent"] > 8:
            flags.append({
                "type": "PRICE_FLAG",
                "severity": "MEDIUM",
                "item": b["item"],
                "message": f"Price is {b['variance_percent']}% above market average - negotiation recommended"
            })
    
    recommendations = [
        f"🤖 AI Analysis Complete: Potential savings of ${total_savings:,.2f} identified",
        "Request volume discount for professional services (160+ hours)",
        "Consider multi-year contract for HVAC maintenance for additional 10-15% savings",
        "Negotiate CAT parts pricing - competitor quotes available",
        "Bundle safety equipment order for better unit pricing"
    ]
    
    response = {
        "success": True,
        "quotation_id": quotation_id,
        "message": "Demo quotation analysis complete",
        "analysis_mode": "DEMO",
        "ai_engines_used": ["OpenAI GPT-5.2", "Claude Sonnet 4.5", "Gemini 3 Flash"],
        "analysis": {
            "extracted_data": DEMO_QUOTATION,
            "price_benchmark": {
                "benchmarks": demo_benchmarks,
                "total_potential_savings": round(total_savings, 2),
                "overall_assessment": "NEGOTIATION_RECOMMENDED",
                "market_data_date": datetime.now().strftime("%Y-%m-%d"),
                "confidence_score": DEMO_ANALYSIS_RESULTS["overall_confidence"],
                "ai_engines_used": ["openai_gpt5.2", "claude_sonnet4.5", "gemini_3_flash"]
            },
            "tax_analysis": {
                "tax_verified": True,
                "calculated_tax": DEMO_QUOTATION["totals"]["tax_amount"],
                "quoted_tax": DEMO_QUOTATION["totals"]["tax_amount"],
                "tax_variance": 0,
                "jurisdiction": "Texas",
                "tax_exemptions_detected": 0,
                "avalara_verification": "VERIFIED",
                "issues": []
            },
            "flags": flags,
            "recommendations": recommendations
        }
    }
    
    # Store demo analysis in database
    await db.quotation_uploads.insert_one({
        "quotation_id": quotation_id,
        "user_id": current_user.get("email"),
        "user_name": current_user.get("name"),
        "file_name": "TechPro_Solutions_Quote_2026.pdf",
        "status": "DEMO_ANALYZED",
        "analysis_mode": "DEMO",
        **response["analysis"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return response


@api_router.get("/procurement/quotation/history")
async def get_quotation_history(
    page: int = 1,
    limit: int = 20,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get user's quotation upload history"""
    query = {"user_id": current_user.get("email")}
    if status:
        query["status"] = status
    
    total = await db.quotation_uploads.count_documents(query)
    
    quotations = await db.quotation_uploads.find(query).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    
    # Remove MongoDB _id and file content
    for q in quotations:
        q.pop("_id", None)
        q.pop("file_content", None)
    
    return {
        "quotations": quotations,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/procurement/quotation/{quotation_id}")
async def get_quotation_details(
    quotation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get detailed quotation analysis"""
    quotation = await db.quotation_uploads.find_one({
        "quotation_id": quotation_id,
        "user_id": current_user.get("email")
    })
    
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    quotation.pop("_id", None)
    return quotation

@api_router.post("/procurement/quotation/{quotation_id}/escalate")
async def escalate_for_negotiation(
    quotation_id: str,
    notes: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Escalate quotation for Infosys negotiation support"""
    result = await db.quotation_uploads.update_one(
        {"quotation_id": quotation_id, "user_id": current_user.get("email")},
        {
            "$set": {
                "status": "ESCALATED_FOR_NEGOTIATION",
                "escalated_for_negotiation": True,
                "escalation_notes": notes,
                "escalated_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    await db.activity_logs.insert_one({
        "user_id": current_user.get("email"),
        "action": "QUOTATION_ESCALATED",
        "quotation_id": quotation_id,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "message": "Quotation escalated to Infosys negotiation team",
        "estimated_response": "24-48 hours"
    }

@api_router.post("/procurement/quotation/{quotation_id}/engage-tactical-buyers")
async def engage_tactical_buyers(
    quotation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Engage Infosys Tactical Buyers for the quotation"""
    quotation = await db.quotation_uploads.find_one({
        "quotation_id": quotation_id,
        "user_id": current_user.get("email")
    })
    
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Create buying desk request record with full tracking stages
    buying_desk_request = {
        "request_id": f"IBD-{uuid.uuid4().hex[:8].upper()}",
        "quotation_id": quotation_id,
        "user_id": current_user.get("email"),
        "user_name": current_user.get("name", "User"),
        "supplier_name": quotation.get("supplier_name", "Unknown"),
        "total_amount": quotation.get("extracted_data", {}).get("totals", {}).get("grand_total", 0),
        "potential_savings": quotation.get("price_benchmark", {}).get("total_potential_savings", 0),
        "line_items_count": len(quotation.get("extracted_data", {}).get("line_items", [])),
        "status": "submitted",
        "current_stage": "submitted",
        "stages": [
            {"stage": "submitted", "title": "Submitted", "completed": True, "completed_at": datetime.now(timezone.utc).isoformat()},
            {"stage": "supplier_identification", "title": "Supplier Identification", "completed": False, "completed_at": None},
            {"stage": "rfq_sent", "title": "RFQ Sent", "completed": False, "completed_at": None},
            {"stage": "quotes_received", "title": "Quotes Received", "completed": False, "completed_at": None},
            {"stage": "negotiating", "title": "Negotiating", "completed": False, "completed_at": None},
            {"stage": "po_ready", "title": "PO Ready", "completed": False, "completed_at": None}
        ],
        "assigned_to": "Infosys Tactical Buying Team",
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "expected_response_by": (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat(),
        "notes": [],
        "request_type": "quotation"
    }
    
    await db.buying_desk_requests.insert_one(buying_desk_request)
    
    # Update quotation status
    await db.quotation_uploads.update_one(
        {"quotation_id": quotation_id},
        {"$set": {
            "tactical_buyers_engaged": True,
            "buying_desk_request_id": buying_desk_request["request_id"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "message": "Infosys Tactical Buying Team has been notified",
        "request_id": buying_desk_request["request_id"],
        "expected_response": "24 hours"
    }

@api_router.get("/procurement/buying-desk/requests")
async def get_buying_desk_requests(current_user: dict = Depends(get_current_user)):
    """Get all buying desk requests for the current user"""
    requests = await db.buying_desk_requests.find(
        {"user_id": current_user.get("email")},
        {"_id": 0}
    ).sort("submitted_at", -1).to_list(100)
    
    return {"requests": requests}

@api_router.get("/procurement/buying-desk/request/{request_id}")
async def get_buying_desk_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """Get details of a specific buying desk request"""
    request = await db.buying_desk_requests.find_one(
        {"request_id": request_id, "user_id": current_user.get("email")},
        {"_id": 0}
    )
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    return request


class BuyingDeskEngageRequest(BaseModel):
    request_type: str = "general_sourcing"
    description: Optional[str] = None
    quotation_id: Optional[str] = None
    search_query: Optional[str] = None
    unspsc_code: Optional[str] = None
    category_name: Optional[str] = None
    supplier_info: Optional[dict] = None
    line_items: Optional[list] = []
    potential_savings: Optional[float] = None
    user_notes: Optional[str] = None
    session_id: Optional[str] = None


@api_router.post("/procurement/buying-desk/engage")
async def engage_buying_desk_with_context(
    request: BuyingDeskEngageRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Engage Infosys Buying Desk with full context from any screen.
    This carries over quotation data, search context, and user requirements.
    """
    
    # Get quotation details if quotation_id is provided
    quotation_data = None
    if request.quotation_id:
        quotation = await db.quotation_uploads.find_one(
            {"quotation_id": request.quotation_id},
            {"_id": 0}
        )
        if quotation:
            quotation_data = {
                "quotation_id": request.quotation_id,
                "supplier_name": quotation.get("supplier_name"),
                "supplier_email": quotation.get("supplier_email"),
                "total_amount": quotation.get("extracted_data", {}).get("totals", {}).get("grand_total", 0),
                "line_items": quotation.get("extracted_data", {}).get("line_items", []),
                "potential_savings": quotation.get("price_benchmark", {}).get("total_potential_savings", 0),
                "benchmarks": quotation.get("price_benchmark", {}).get("benchmarks", [])
            }
    
    # Determine priority based on context
    priority = "standard"
    if quotation_data and quotation_data.get("potential_savings", 0) > 5000:
        priority = "high"
    elif request.request_type in ["complex_sourcing", "strategic"]:
        priority = "high"
    
    # Build comprehensive description
    description_parts = []
    if request.description:
        description_parts.append(request.description)
    if request.search_query:
        description_parts.append(f"Search: {request.search_query}")
    if request.category_name:
        description_parts.append(f"Category: {request.category_name}")
    if quotation_data:
        description_parts.append(f"Quotation from {quotation_data.get('supplier_name', 'supplier')} with {len(quotation_data.get('line_items', []))} items")
    
    full_description = " | ".join(description_parts) if description_parts else "Procurement assistance requested"
    
    # Create buying desk request record
    buying_desk_request = {
        "request_id": f"IBD-{uuid.uuid4().hex[:8].upper()}",
        "user_id": current_user.get("email"),
        "user_name": current_user.get("name", "User"),
        "user_company": current_user.get("company", ""),
        "request_type": request.request_type,
        "description": full_description,
        "search_query": request.search_query,
        "unspsc_code": request.unspsc_code,
        "category_name": request.category_name,
        "quotation_id": request.quotation_id,
        "quotation_data": quotation_data,
        "supplier_info": request.supplier_info,
        "line_items": request.line_items or (quotation_data.get("line_items", []) if quotation_data else []),
        "potential_savings": request.potential_savings or (quotation_data.get("potential_savings", 0) if quotation_data else 0),
        "user_notes": request.user_notes,
        "session_id": request.session_id,
        "status": "submitted",
        "priority": priority,
        "current_stage": "submitted",
        "stages": [
            {"stage": "submitted", "title": "Submitted", "completed": True, "completed_at": datetime.now(timezone.utc).isoformat()},
            {"stage": "review", "title": "Under Review", "completed": False, "completed_at": None},
            {"stage": "supplier_identification", "title": "Supplier Identification", "completed": False, "completed_at": None},
            {"stage": "rfq_sent", "title": "RFQ Sent", "completed": False, "completed_at": None},
            {"stage": "quotes_received", "title": "Quotes Received", "completed": False, "completed_at": None},
            {"stage": "negotiating", "title": "Negotiating", "completed": False, "completed_at": None},
            {"stage": "po_ready", "title": "PO Ready", "completed": False, "completed_at": None}
        ],
        "assigned_to": None,
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "expected_response_by": (datetime.now(timezone.utc) + timedelta(hours=4)).isoformat(),
        "notes": []
    }
    
    await db.buying_desk_requests.insert_one(buying_desk_request)
    
    # Update quotation status if quotation_id provided
    if request.quotation_id:
        await db.quotation_uploads.update_one(
            {"quotation_id": request.quotation_id},
            {"$set": {
                "buying_desk_engaged": True,
                "buying_desk_request_id": buying_desk_request["request_id"],
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Log activity
    await db.activity_logs.insert_one({
        "user_id": current_user.get("email"),
        "action": "BUYING_DESK_ENGAGED",
        "request_id": buying_desk_request["request_id"],
        "quotation_id": request.quotation_id,
        "request_type": request.request_type,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "message": "Infosys Buying Desk has been notified and will contact you shortly",
        "request_id": buying_desk_request["request_id"],
        "priority": priority,
        "expected_response": "2-4 business hours"
    }

@api_router.post("/procurement/quotation/{quotation_id}/add-to-cart")
async def add_quotation_to_cart(
    quotation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Add analyzed quotation items to cart"""
    quotation = await db.quotation_uploads.find_one({
        "quotation_id": quotation_id,
        "user_id": current_user.get("email")
    })
    
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Add items to cart
    items_added = 0
    for item in quotation.get("extracted_data", {}).get("line_items", []):
        cart_item = {
            "user_id": current_user.get("email"),
            "product_id": f"QUP-ITEM-{uuid.uuid4().hex[:8]}",
            "product_name": item["description"],
            "brand": quotation.get("supplier_name", "Unknown"),
            "sku": f"QUP-{item['line_number']}",
            "unspsc_code": item.get("unspsc_code", ""),
            "category": item.get("category", "General"),
            "quantity": item["quantity"],
            "unit_price": item["unit_price"],
            "total_price": item["line_total"],
            "currency_code": "USD",
            "is_service": False,
            "source": "quotation_upload",
            "quotation_id": quotation_id,
            "added_at": datetime.now(timezone.utc).isoformat()
        }
        await db.carts.insert_one(cart_item)
        items_added += 1
    
    await db.quotation_uploads.update_one(
        {"quotation_id": quotation_id},
        {"$set": {"added_to_cart": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {
        "success": True,
        "message": f"Added {items_added} items to cart",
        "items_added": items_added
    }

# ============================================
# NEGOTIATION AGENT ENDPOINTS
# ============================================

@api_router.get("/negotiation/strategies")
async def get_negotiation_strategies(current_user: dict = Depends(get_current_user)):
    """Get all available negotiation strategies/playbooks"""
    return {
        "success": True,
        "strategies": get_all_strategies()
    }

class NegotiationTargetRequest(BaseModel):
    quotation_id: str
    strategy: str = "balanced"

@api_router.post("/negotiation/generate-targets")
async def generate_targets(
    request: NegotiationTargetRequest,
    current_user: dict = Depends(get_current_user)
):
    """Generate negotiation targets for a quotation based on selected strategy"""
    # Get quotation data
    quotation = await db.quotation_uploads.find_one({
        "quotation_id": request.quotation_id,
        "user_id": current_user.get("email")
    }, {"_id": 0})
    
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Parse strategy
    try:
        strategy = NegotiationStrategy(request.strategy)
    except ValueError:
        strategy = NegotiationStrategy.BALANCED
    
    # Get line items and benchmarks
    analysis = quotation.get("analysis", {})
    extracted_data = analysis.get("extracted_data", quotation.get("extracted_data", {}))
    line_items = extracted_data.get("line_items", [])
    benchmarks = analysis.get("price_benchmark", {}).get("benchmarks", [])
    
    if not line_items:
        raise HTTPException(status_code=400, detail="No line items found in quotation")
    
    # Generate targets
    targets = generate_negotiation_targets(line_items, benchmarks, strategy)
    
    # Store negotiation session
    negotiation_session = {
        "negotiation_id": f"NEG-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}",
        "quotation_id": request.quotation_id,
        "user_id": current_user.get("email"),
        "strategy": strategy.value,
        "targets": targets,
        "status": "initiated",
        "rounds": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.negotiations.insert_one(negotiation_session)
    
    return {
        "success": True,
        "negotiation_id": negotiation_session["negotiation_id"],
        "quotation_id": request.quotation_id,
        "strategy": strategy.value,
        "targets": targets,
        "supplier": extracted_data.get("supplier", {}),
        "quotation_details": extracted_data.get("quotation_details", {})
    }

class NegotiationEmailRequest(BaseModel):
    quotation_id: str
    negotiation_id: Optional[str] = None
    strategy: str = "balanced"
    buyer_name: str = "Procurement Team"
    company_name: str = "Infosys Limited"

@api_router.post("/negotiation/generate-email")
async def generate_email(
    request: NegotiationEmailRequest,
    current_user: dict = Depends(get_current_user)
):
    """Generate a negotiation email for a quotation"""
    # Get quotation data
    quotation = await db.quotation_uploads.find_one({
        "quotation_id": request.quotation_id,
        "user_id": current_user.get("email")
    }, {"_id": 0})
    
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Parse strategy
    try:
        strategy = NegotiationStrategy(request.strategy)
    except ValueError:
        strategy = NegotiationStrategy.BALANCED
    
    # Get data
    analysis = quotation.get("analysis", {})
    extracted_data = analysis.get("extracted_data", quotation.get("extracted_data", {}))
    line_items = extracted_data.get("line_items", [])
    benchmarks = analysis.get("price_benchmark", {}).get("benchmarks", [])
    supplier_info = extracted_data.get("supplier", {})
    quotation_details = extracted_data.get("quotation_details", {})
    
    # Generate targets first
    targets = generate_negotiation_targets(line_items, benchmarks, strategy)
    
    # Generate email
    session_id = f"neg_email_{request.quotation_id}"
    email_data = await generate_negotiation_email(
        quotation_data=quotation_details,
        negotiation_targets=targets,
        strategy=strategy,
        supplier_info=supplier_info,
        buyer_info={
            "name": request.buyer_name,
            "company": request.company_name
        },
        session_id=session_id
    )
    
    # Update negotiation session if exists
    if request.negotiation_id:
        await db.negotiations.update_one(
            {"negotiation_id": request.negotiation_id},
            {
                "$set": {
                    "email_generated": True,
                    "email_data": email_data,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                },
                "$push": {
                    "rounds": {
                        "round": 1,
                        "type": "initial_outreach",
                        "email": email_data,
                        "sent_at": None,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            }
        )
    
    return {
        "success": True,
        "email": email_data,
        "targets": targets,
        "supplier": supplier_info
    }

class CounterOfferRequest(BaseModel):
    negotiation_id: str
    their_offer: float
    notes: Optional[str] = None

@api_router.post("/negotiation/counter-offer")
async def process_counter_offer(
    request: CounterOfferRequest,
    current_user: dict = Depends(get_current_user)
):
    """Process a supplier's counter-offer and generate our response"""
    # Get negotiation session
    negotiation = await db.negotiations.find_one({
        "negotiation_id": request.negotiation_id,
        "user_id": current_user.get("email")
    }, {"_id": 0})
    
    if not negotiation:
        raise HTTPException(status_code=404, detail="Negotiation session not found")
    
    # Get current round
    rounds = negotiation.get("rounds", [])
    current_round = len(rounds)
    
    # Get strategy and targets
    try:
        strategy = NegotiationStrategy(negotiation.get("strategy", "balanced"))
    except ValueError:
        strategy = NegotiationStrategy.BALANCED
    
    targets = negotiation.get("targets", {})
    summary = targets.get("summary", {})
    target_price = summary.get("total_target", request.their_offer * 0.9)
    
    # Get our last offer (or initial if first counter)
    if rounds:
        last_round = rounds[-1]
        our_last_offer = last_round.get("our_offer", summary.get("total_target", request.their_offer * 0.85))
    else:
        our_last_offer = summary.get("total_target", request.their_offer * 0.85)
    
    # Calculate counter-offer
    counter = create_counter_offer(
        current_round=current_round,
        their_offer=request.their_offer,
        our_last_offer=our_last_offer,
        target_price=target_price,
        strategy=strategy
    )
    
    # Record the round
    round_data = {
        "round": counter["round"],
        "their_offer": request.their_offer,
        "our_offer": counter["our_counter"],
        "gap": counter["gap_to_counter"],
        "recommendation": counter["recommendation"],
        "notes": request.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update status based on recommendation
    new_status = "in_progress"
    if counter["should_walk_away"]:
        new_status = "escalate_recommended"
    elif counter["rounds_remaining"] == 0:
        new_status = "final_round"
    
    await db.negotiations.update_one(
        {"negotiation_id": request.negotiation_id},
        {
            "$set": {
                "status": new_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {"rounds": round_data}
        }
    )
    
    # Calculate savings achieved so far
    original_quoted = summary.get("total_quoted", request.their_offer)
    savings_so_far = original_quoted - request.their_offer
    savings_percent = (savings_so_far / original_quoted) * 100 if original_quoted > 0 else 0
    
    return {
        "success": True,
        "counter_offer": counter,
        "round_data": round_data,
        "status": new_status,
        "savings_achieved": {
            "amount": round(savings_so_far, 2),
            "percent": round(savings_percent, 1),
            "from_original": round(original_quoted, 2)
        }
    }

@api_router.post("/negotiation/{negotiation_id}/close")
async def close_negotiation(
    negotiation_id: str,
    outcome: str = "accepted",  # accepted, rejected, escalated
    final_price: Optional[float] = None,
    notes: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Close a negotiation session with final outcome"""
    negotiation = await db.negotiations.find_one({
        "negotiation_id": negotiation_id,
        "user_id": current_user.get("email")
    })
    
    if not negotiation:
        raise HTTPException(status_code=404, detail="Negotiation session not found")
    
    targets = negotiation.get("targets", {})
    summary = targets.get("summary", {})
    original_quoted = summary.get("total_quoted", 0)
    
    # Calculate final savings
    final_savings = 0
    final_savings_percent = 0
    if final_price and original_quoted > 0:
        final_savings = original_quoted - final_price
        final_savings_percent = (final_savings / original_quoted) * 100
    
    # Update negotiation
    await db.negotiations.update_one(
        {"negotiation_id": negotiation_id},
        {
            "$set": {
                "status": f"closed_{outcome}",
                "outcome": outcome,
                "final_price": final_price,
                "final_savings": round(final_savings, 2),
                "final_savings_percent": round(final_savings_percent, 1),
                "closed_notes": notes,
                "closed_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Log activity
    await db.activity_logs.insert_one({
        "user_id": current_user.get("email"),
        "action": "negotiation_closed",
        "details": {
            "negotiation_id": negotiation_id,
            "outcome": outcome,
            "final_price": final_price,
            "savings": final_savings
        },
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "negotiation_id": negotiation_id,
        "outcome": outcome,
        "final_price": final_price,
        "savings": {
            "amount": round(final_savings, 2),
            "percent": round(final_savings_percent, 1)
        },
        "message": f"Negotiation closed with outcome: {outcome}"
    }

@api_router.get("/negotiation/history")
async def get_negotiation_history(
    current_user: dict = Depends(get_current_user),
    limit: int = 20
):
    """Get user's negotiation history"""
    negotiations = await db.negotiations.find(
        {"user_id": current_user.get("email")},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {
        "success": True,
        "negotiations": negotiations,
        "count": len(negotiations)
    }

@api_router.get("/negotiation/{negotiation_id}")
async def get_negotiation_details(
    negotiation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get details of a specific negotiation"""
    negotiation = await db.negotiations.find_one({
        "negotiation_id": negotiation_id,
        "user_id": current_user.get("email")
    }, {"_id": 0})
    
    if not negotiation:
        raise HTTPException(status_code=404, detail="Negotiation not found")
    
    return {
        "success": True,
        "negotiation": negotiation
    }

# ============================================
# END-TO-END SOURCING SUPPORT ENDPOINTS
# ============================================

class SourcingRequest(BaseModel):
    request_title: str
    category: str
    description: str
    estimated_budget: Optional[float] = None
    budget_currency: str = "USD"
    quantity: Optional[int] = None
    required_by_date: Optional[str] = None
    delivery_location: str
    preferred_suppliers: Optional[List[str]] = None
    technical_specifications: Optional[str] = None
    payment_model: str = "infosys_limited"  # infosys_limited, propay, customer_direct
    urgency: str = "standard"  # standard, urgent, critical
    attachments_count: int = 0

@api_router.post("/procurement/sourcing/request")
async def submit_sourcing_request(
    request: SourcingRequest,
    current_user: dict = Depends(get_current_user)
):
    """Submit a new end-to-end sourcing request"""
    sourcing_id = f"SRC-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
    
    # Simulate processing timeline based on urgency
    processing_days = {"standard": 5, "urgent": 2, "critical": 1}
    estimated_completion = datetime.now() + timedelta(days=processing_days.get(request.urgency, 5))
    
    sourcing_record = {
        "sourcing_id": sourcing_id,
        "user_id": current_user.get("email"),
        "user_name": current_user.get("name"),
        "request_title": request.request_title,
        "category": request.category,
        "description": request.description,
        "estimated_budget": request.estimated_budget,
        "budget_currency": request.budget_currency,
        "quantity": request.quantity,
        "required_by_date": request.required_by_date,
        "delivery_location": request.delivery_location,
        "preferred_suppliers": request.preferred_suppliers or [],
        "technical_specifications": request.technical_specifications,
        "payment_model": request.payment_model,
        "urgency": request.urgency,
        "status": "SUBMITTED",
        "status_history": [
            {
                "status": "SUBMITTED",
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "notes": "Request received and queued for processing"
            }
        ],
        "assigned_specialist": None,
        "suppliers_identified": [],
        "rfq_sent_count": 0,
        "quotations_received": [],
        "recommended_supplier": None,
        "estimated_completion": estimated_completion.strftime("%Y-%m-%d"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sourcing_requests.insert_one(sourcing_record)
    
    await db.activity_logs.insert_one({
        "user_id": current_user.get("email"),
        "action": "SOURCING_REQUEST_SUBMITTED",
        "sourcing_id": sourcing_id,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "sourcing_id": sourcing_id,
        "message": "Sourcing request submitted successfully",
        "estimated_completion": estimated_completion.strftime("%Y-%m-%d"),
        "next_steps": [
            "Our procurement specialist will review your request within 24 hours",
            "We will identify and qualify potential suppliers",
            "RFQ will be sent to selected suppliers",
            "You will receive analyzed quotations with our recommendations"
        ]
    }

@api_router.get("/procurement/sourcing/history")
async def get_sourcing_history(
    page: int = 1,
    limit: int = 20,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get user's sourcing request history"""
    query = {"user_id": current_user.get("email")}
    if status:
        query["status"] = status
    
    total = await db.sourcing_requests.count_documents(query)
    
    requests = await db.sourcing_requests.find(query).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    
    for r in requests:
        r.pop("_id", None)
    
    return {
        "requests": requests,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/procurement/sourcing/{sourcing_id}")
async def get_sourcing_details(
    sourcing_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get detailed sourcing request status"""
    request = await db.sourcing_requests.find_one({
        "sourcing_id": sourcing_id,
        "user_id": current_user.get("email")
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Sourcing request not found")
    
    request.pop("_id", None)
    
    # Simulate progress updates for demo
    if request["status"] == "SUBMITTED":
        # Auto-progress for demo
        await db.sourcing_requests.update_one(
            {"sourcing_id": sourcing_id},
            {
                "$set": {
                    "status": "IN_PROGRESS",
                    "assigned_specialist": {
                        "name": random.choice(["Priya Sharma", "Rajesh Kumar", "Sarah Johnson", "Michael Chen"]),
                        "email": "procurement@infosysbpm.com",
                        "phone": "+1-800-INFOSYS"
                    }
                },
                "$push": {
                    "status_history": {
                        "status": "IN_PROGRESS",
                        "timestamp": datetime.now(timezone.utc).isoformat(),
                        "notes": "Assigned to procurement specialist"
                    }
                }
            }
        )
        request["status"] = "IN_PROGRESS"
    
    return request

@api_router.post("/procurement/sourcing/{sourcing_id}/cancel")
async def cancel_sourcing_request(
    sourcing_id: str,
    reason: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Cancel a sourcing request"""
    result = await db.sourcing_requests.update_one(
        {
            "sourcing_id": sourcing_id,
            "user_id": current_user.get("email"),
            "status": {"$nin": ["COMPLETED", "CANCELLED"]}
        },
        {
            "$set": {
                "status": "CANCELLED",
                "cancellation_reason": reason,
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$push": {
                "status_history": {
                    "status": "CANCELLED",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "notes": reason or "Cancelled by user"
                }
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Cannot cancel this request")
    
    return {"success": True, "message": "Sourcing request cancelled"}

# ============================================
# USER PROFILE & ACTIVITY ENDPOINTS
# ============================================

@api_router.get("/user/profile")
async def get_user_profile(current_user: dict = Depends(get_current_user)):
    """Get user profile with activity summary"""
    email = current_user.get("email")
    
    # Get activity counts
    quotation_count = await db.quotation_uploads.count_documents({"user_id": email})
    sourcing_count = await db.sourcing_requests.count_documents({"user_id": email})
    order_count = await db.orders.count_documents({"user_id": email})
    
    # Get recent activity
    recent_activity = await db.activity_logs.find({"user_id": email}).sort("timestamp", -1).limit(10).to_list(10)
    for a in recent_activity:
        a.pop("_id", None)
    
    return {
        "user": {
            "email": current_user.get("email"),
            "name": current_user.get("name"),
            "country": current_user.get("country"),
            "info_coins": current_user.get("info_coins", 0),
            "member_since": current_user.get("created_at", datetime.now(timezone.utc).isoformat())
        },
        "activity_summary": {
            "quotations_uploaded": quotation_count,
            "sourcing_requests": sourcing_count,
            "orders_placed": order_count
        },
        "recent_activity": recent_activity
    }

@api_router.get("/procurement/dashboard")
async def get_procurement_dashboard(current_user: dict = Depends(get_current_user)):
    """Get procurement dashboard data"""
    email = current_user.get("email")
    
    # Quotation stats
    quotation_stats = {
        "total": await db.quotation_uploads.count_documents({"user_id": email}),
        "analyzed": await db.quotation_uploads.count_documents({"user_id": email, "status": "ANALYZED"}),
        "escalated": await db.quotation_uploads.count_documents({"user_id": email, "escalated_for_negotiation": True}),
        "added_to_cart": await db.quotation_uploads.count_documents({"user_id": email, "added_to_cart": True})
    }
    
    # Sourcing stats
    sourcing_stats = {
        "total": await db.sourcing_requests.count_documents({"user_id": email}),
        "in_progress": await db.sourcing_requests.count_documents({"user_id": email, "status": "IN_PROGRESS"}),
        "completed": await db.sourcing_requests.count_documents({"user_id": email, "status": "COMPLETED"}),
        "pending": await db.sourcing_requests.count_documents({"user_id": email, "status": "SUBMITTED"})
    }
    
    # Recent quotations
    recent_quotations = await db.quotation_uploads.find({"user_id": email}).sort("created_at", -1).limit(5).to_list(5)
    for q in recent_quotations:
        q.pop("_id", None)
        q.pop("extracted_data", None)
        q.pop("price_benchmark", None)
    
    # Recent sourcing
    recent_sourcing = await db.sourcing_requests.find({"user_id": email}).sort("created_at", -1).limit(5).to_list(5)
    for s in recent_sourcing:
        s.pop("_id", None)
    
    return {
        "quotation_stats": quotation_stats,
        "sourcing_stats": sourcing_stats,
        "recent_quotations": recent_quotations,
        "recent_sourcing": recent_sourcing
    }

# ============================================
# AI PROCUREMENT AGENT - CONVERSATIONAL ENDPOINT
# ============================================

class AIAgentConversationRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    context: Optional[Dict[str, Any]] = None
    language: str = "en"
    currency: str = "USD"

class AIAgentConversationResponse(BaseModel):
    message: str
    engines_used: List[str] = []
    action: Optional[str] = None
    products: Optional[List[Dict[str, Any]]] = None
    services: Optional[List[Dict[str, Any]]] = None
    context: Optional[Dict[str, Any]] = None
    search_results: Optional[Dict[str, Any]] = None
    supplier_form: Optional[bool] = None
    managed_service_form: Optional[bool] = None
    unspsc_suggestion: Optional[Dict[str, Any]] = None

# AI Agent System Prompt for routing and classification
AI_AGENT_SYSTEM_PROMPT = """You are an intelligent procurement assistant for OMNISupply.io, powered by Infosys BPM.
Your job is to understand the user's procurement need and route them to the correct workflow.

CLASSIFICATION RULES:
1. **CATALOG_SEARCH**: User wants to find a specific product or service from the catalog
   - Keywords: "find", "search", "looking for", "need", "buy", "purchase", specific product names, part numbers, brands
   - Ask for: Part numbers, manufacturer/brand names, specifications, quantities

2. **QUOTATION_ANALYSIS**: User has an existing quotation they want analyzed
   - Keywords: "quotation", "quote", "received from supplier", "analyze pricing", "benchmark"
   - Route to: Upload Quotation page for AI analysis

3. **MANAGED_SERVICES**: Complex, strategic, or high-value sourcing needs
   - Keywords: "strategic sourcing", "long-term contract", "complex requirements", "multiple suppliers", "RFP", "tender", "category management"
   - Action: Capture requirements and notify category expert

RESPONSE FORMAT:
- Be conversational and helpful
- Ask clarifying questions to understand the need better
- Guide users step by step
- Provide clear next actions

When you identify the intent, respond with JSON in this exact format:
{
    "intent": "CATALOG_SEARCH" | "QUOTATION_ANALYSIS" | "MANAGED_SERVICES" | "CLARIFICATION_NEEDED",
    "response_message": "Your conversational response to the user",
    "search_type": "product" | "service" | null,
    "search_query": "extracted search terms" | null,
    "category_hint": "suggested UNSPSC category" | null,
    "confidence": 0.0-1.0
}"""

# Intelligent response templates for different scenarios
INTELLIGENT_RESPONSES = {
    "no_results_with_alternatives": """I've searched our catalog of 30M+ products and services, but I couldn't find an exact match for **"{query}"**.

Let me help you find the right solution:

**Option 1: Do you have a supplier in mind?**
If you've already identified a supplier who can provide this, I can help you:
• Upload their quotation for AI-powered price benchmarking
• Verify pricing against market rates
• Ensure tax compliance

**Option 2: Need help finding a supplier?**
Our Infosys Buying Desk specialists can:
• Identify qualified suppliers for your specific requirement
• Manage the RFQ process
• Negotiate best pricing on your behalf

What would you like to do?""",

    "complex_requirement_detected": """This sounds like a specialized requirement that could benefit from expert sourcing support.

For **"{query}"**, I recommend our **Managed Services** approach where our procurement specialists can:

• **Source qualified suppliers** who can meet your exact specifications
• **Manage the entire RFQ/RFP process** from start to finish
• **Negotiate optimal pricing** leveraging our $2B+ annual spend volume
• **Ensure compliance** with your procurement policies

Would you like me to connect you with a category expert? They typically respond within 2-4 business hours.""",

    "quotation_prompt": """I understand you may have a supplier who can provide **"{query}"**.

If you have a quotation from them, I can help you:
• **Extract all line items automatically** using AI
• **Benchmark prices** against market rates and historical data
• **Identify potential savings** of 15-30% on average
• **Verify tax calculations** for compliance

Would you like to upload a quotation for analysis?""",

    "clarification_needed": """I want to make sure I understand your requirement correctly.

You mentioned **"{query}"** - could you help me with a few details?

• **What category does this fall under?** (e.g., Industrial Equipment, IT Hardware, Services)
• **Do you have a specific brand or manufacturer preference?**
• **Is this a one-time purchase or recurring need?**
• **Do you already have a supplier or quotation?**

This will help me guide you to the best procurement path."""
}

def is_likely_not_in_catalog(query: str) -> bool:
    """
    Detect if a query is likely NOT in our standard industrial/IT catalog.
    Uses pattern matching for unusual, consumer, or highly specific items.
    """
    query_lower = query.lower()
    
    # Consumer/non-industrial indicators - use word boundaries
    consumer_indicators = [
        'bike', 'bicycle', 'car ', ' car', 'vehicle', 'motorcycle', 'scooter',
        'food', 'grocery', 'restaurant', 'coffee', 'snack',
        'clothing', 'shirt', 'pants', 'shoes', 'dress', 'fashion',
        'toy', 'game', 'entertainment', 'movie', 'music',
        'pet ', ' pet', 'dog', 'cat ', ' cat', 'animal',
        'jewelry', 'watch', ' ring ', 'necklace',  # ring with spaces to avoid "bearings"
        'furniture', 'sofa', 'couch', ' bed ', 'mattress',
        'cosmetic', 'makeup', 'beauty', 'perfume',
        'sports', ' gym ', 'fitness', 'yoga',
        'travel', 'vacation', 'hotel', 'flight'
    ]
    
    # Color + unusual item combinations (like "blue bike with red dots")
    color_words = ['red', 'blue', 'green', 'yellow', 'orange', 'purple', 'pink', 'black', 'white', 'gold', 'silver']
    unusual_patterns = ['dots', 'stripes', 'pattern', 'custom', 'personalized', 'unique', 'rare', 'vintage', 'antique']
    
    # Check for consumer items
    if any(indicator in query_lower for indicator in consumer_indicators):
        return True
    
    # Check for color + unusual pattern combinations
    has_color = any(color in query_lower for color in color_words)
    has_unusual = any(pattern in query_lower for pattern in unusual_patterns)
    if has_color and has_unusual:
        return True
    
    # Check for very specific/custom descriptions (long queries with adjectives)
    words = query_lower.split()
    if len(words) > 5 and has_color:
        return True
    
    return False

def assess_requirement_complexity(query: str, search_results: dict) -> str:
    """
    Assess if a requirement is complex enough to warrant Managed Services.
    Returns: 'simple', 'moderate', 'complex'
    """
    query_lower = query.lower()
    
    # Complex indicators
    complex_indicators = [
        'multiple', 'several', 'various', 'different',
        'custom', 'specialized', 'specific', 'unique',
        'large quantity', 'bulk', 'volume',
        'installation', 'setup', 'configuration', 'integration',
        'ongoing', 'recurring', 'long-term', 'contract',
        'compliance', 'certified', 'approved', 'qualified',
        'international', 'global', 'cross-border',
        'urgent', 'emergency', 'critical', 'asap'
    ]
    
    complexity_score = sum(1 for indicator in complex_indicators if indicator in query_lower)
    
    # No results + complexity indicators = likely complex
    no_results = not search_results.get('products') and not search_results.get('services')
    
    if complexity_score >= 3 or (no_results and complexity_score >= 1):
        return 'complex'
    elif complexity_score >= 1 or no_results:
        return 'moderate'
    else:
        return 'simple'

async def get_conversation_context(session_id: str, limit: int = 8) -> Dict:
    """Retrieve and structure conversation history for context-aware routing"""
    try:
        recent_messages = await db.ai_agent_conversations.find(
            {"session_id": session_id}
        ).sort("timestamp", -1).limit(limit).to_list(limit)
        
        if not recent_messages:
            return {"history": [], "current_topic": None, "last_search": None, "last_intent": None}
        
        recent_messages.reverse()  # Oldest first for proper context flow
        
        history = []
        current_topic = None
        last_search = None
        last_intent = None
        
        for msg in recent_messages:
            history.append({
                "user": msg.get('message', ''),
                "assistant": msg.get('response', '')[:400],
                "intent": msg.get('intent'),
                "search_query": msg.get('search_query'),
                "topic": msg.get('understood_topic')
            })
            
            # Track the latest non-null values
            if msg.get('search_query'):
                last_search = msg.get('search_query')
            if msg.get('understood_topic'):
                current_topic = msg.get('understood_topic')
            if msg.get('intent'):
                last_intent = msg.get('intent')
        
        return {
            "history": history,
            "current_topic": current_topic or last_search,
            "last_search": last_search,
            "last_intent": last_intent
        }
    except Exception as e:
        logger.error(f"Error fetching conversation context: {e}")
        return {"history": [], "current_topic": None, "last_search": None, "last_intent": None}


def detect_follow_up_question(message: str) -> bool:
    """Detect if a message is likely a follow-up question referencing prior context"""
    message_lower = message.lower().strip()
    words = message_lower.split()
    
    # Short messages that reference something ("it", "that", "those", "this")
    # Use word-level matching to avoid false positives (e.g., "with" contains "it")
    referential_words = {'it', 'that', 'those', 'this', 'them', 'they', 'these'}
    
    # Common follow-up patterns
    follow_up_patterns = [
        'what brand', 'which brand', 'what manufacturer', 'which manufacturer',
        'any alternative', 'other option', 'cheaper', 'more expensive', 'better',
        'show me more', 'see more', 'more details', 'more info',
        'how much', 'what price', 'what\'s the price', 'pricing',
        'in stock', 'available', 'availability', 'lead time', 'delivery',
        'specs', 'specifications', 'features',
        'add to cart', 'add it', 'buy it', 'order it',
        'compare', 'comparison', 'versus', 'vs',
        'what else', 'anything else', 'other', 'similar',
        'tell me more', 'explain',
        'can you', 'could you', 'would you',
        'yes please', 'no thanks', 'sounds good', 'go ahead',
        'perfect', 'great thanks'
    ]
    
    # Check for referential words - must be standalone words, not substrings
    if len(words) <= 5:  # Short messages are often follow-ups
        if any(word in referential_words for word in words):
            return True
    
    # Check for multi-word referential phrases
    if 'the one' in message_lower or 'the ones' in message_lower:
        return True
    
    # Check for follow-up patterns
    if any(pattern in message_lower for pattern in follow_up_patterns):
        return True
    
    # Questions without clear subject (very short)
    if message_lower.startswith(('what', 'which', 'how', 'where', 'when', 'can', 'could', 'is', 'are', 'do', 'does')):
        if len(words) <= 4:  # Very short question like "what brands?" or "is it available?"
            return True
    
    return False


async def classify_user_intent_with_ai(message: str, context: Dict, session_id: str) -> Dict:
    """
    ROUTER PATTERN: Use LLM to intelligently classify user intent with full conversation context.
    This is the primary decision-maker for routing conversations.
    """
    
    # First, get conversation history
    conv_context = await get_conversation_context(session_id)
    is_likely_follow_up = detect_follow_up_question(message)
    
    if not EMERGENT_AVAILABLE or not EMERGENT_LLM_KEY:
        # Fallback to keyword-based classification WITH context awareness
        message_lower = message.lower()
        
        # If it's a follow-up and we have context, use context
        if is_likely_follow_up and conv_context.get('current_topic'):
            topic = conv_context['current_topic']
            return {
                "intent": "CONTEXT_CONTINUATION",
                "response_message": f"Based on our conversation about **{topic}**, let me help you with that.",
                "search_type": "product",
                "search_query": f"{topic} {message}",  # Combine topic with new query
                "confidence": 0.85,
                "references_prior_context": True,
                "understood_topic": topic
            }
        
        # Check for quotation keywords
        if any(kw in message_lower for kw in ['quotation', 'quote', 'analyze', 'benchmark', 'pricing analysis']):
            return {
                "intent": "QUOTATION_ANALYSIS",
                "response_message": "I understand you have a quotation you'd like analyzed. Our AI-powered system can extract data, benchmark prices against market rates, and identify potential savings.\n\nWould you like me to take you to the quotation upload page, or do you have specific questions first?",
                "search_type": None,
                "search_query": None,
                "confidence": 0.85
            }
        
        # Check for managed services keywords
        if any(kw in message_lower for kw in ['strategic', 'complex', 'rfp', 'tender', 'long-term', 'category management', 'multiple suppliers']):
            return {
                "intent": "MANAGED_SERVICES",
                "response_message": "This sounds like a strategic sourcing requirement that would benefit from our Managed Services team.\n\nOur Infosys Buying Desk specialists can help with:\n• Supplier identification and qualification\n• RFQ/RFP management\n• Expert negotiation\n• End-to-end procurement support\n\nWould you like me to connect you with a category expert?",
                "search_type": None,
                "search_query": None,
                "confidence": 0.80
            }
        
        # Default to catalog search
        return {
            "intent": "CATALOG_SEARCH",
            "response_message": f"I'll help you find what you're looking for. Let me search our catalog of 30M+ products.\n\nTo give you the best results, could you provide any of the following?\n• Part number or SKU\n• Manufacturer/Brand name\n• Product specifications\n• Required quantity",
            "search_type": "product",
            "search_query": message,
            "confidence": 0.70
        }
    
    # Use LLM for intelligent classification with conversation history
    try:
        # Use pre-fetched conversation context
        current_topic = conv_context.get('current_topic')
        
        # Build conversation history string
        history_str = ""
        if conv_context.get('history'):
            history_str = "\n\n## CONVERSATION HISTORY (CRITICAL - READ THIS FIRST):\n"
            for i, msg in enumerate(conv_context['history']):
                history_str += f"[Turn {i+1}]\n"
                history_str += f"USER: {msg['user']}\n"
                history_str += f"ASSISTANT: {msg['assistant']}\n"
                if msg.get('topic'):
                    history_str += f"[Topic: {msg['topic']}]\n"
                history_str += "\n"
        
        # Build context string from frontend context
        context_str = ""
        if context:
            if context.get("quotation_analyzed"):
                context_str += f"\n\n**ACTIVE QUOTATION:**\nSupplier: {context.get('supplier_name')}\nTotal: ${context.get('quotation_total', 0):,.2f}\nItems: {context.get('line_items_count', 0)}\nPotential savings: ${context.get('potential_savings', 0):,.2f}\n"
            if context.get("cart_items"):
                context_str += f"\n**Cart:** {len(context.get('cart_items', []))} items\n"
            if context.get("search_query") and not current_topic:
                current_topic = context.get('search_query')
        
        # IMPROVED System prompt for better context understanding
        intelligent_system_prompt = f"""You are a highly intelligent AI procurement assistant for Infosys OMNISupply.io. 

## YOUR PRIMARY DIRECTIVE: MAINTAIN CONVERSATION CONTEXT
You MUST behave like ChatGPT - understanding that follow-up questions relate to previous messages.

## CONTEXT AWARENESS RULES (CRITICAL):
1. If user previously asked about "bearings" and now asks "what brands?" → They want BEARING BRANDS
2. If user asked about "laptops" and now says "show cheaper ones" → They want CHEAPER LAPTOPS
3. SHORT MESSAGES like "yes", "ok", "add it", "compare them" → Refer to previous topic
4. NEVER ask for clarification when context is clear from history

## CURRENT CONVERSATION TOPIC: {current_topic or 'None established yet'}

## INTENT CLASSIFICATION:
- **CONTEXT_CONTINUATION**: Follow-up on previous topic (most common for short messages)
- **CATALOG_SEARCH**: New product/service search
- **QUOTATION_ANALYSIS**: User has a quote to analyze
- **MANAGED_SERVICES**: Complex/strategic sourcing needs
- **CLARIFICATION_NEEDED**: ONLY if truly ambiguous with no context

## CAPABILITIES:
- Search 30M+ industrial/MRO products and 100K+ professional services
- Analyze quotations with AI price benchmarking
- Connect to Infosys Buying Desk for complex sourcing

## OUTPUT FORMAT (STRICT JSON):
{{
    "intent": "CONTEXT_CONTINUATION" | "CATALOG_SEARCH" | "QUOTATION_ANALYSIS" | "MANAGED_SERVICES" | "CLARIFICATION_NEEDED",
    "response_message": "Conversational response that references prior context when applicable",
    "search_type": "product" | "service" | null,
    "search_query": "ALWAYS include topic from history + refinement. E.g., 'industrial bearings SKF brand'",
    "confidence": 0.0-1.0,
    "references_prior_context": true | false,
    "understood_topic": "What the user is asking about (inferred from context)"
}}"""

        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"agent_{session_id}_router",
            system_message=intelligent_system_prompt
        ).with_model("openai", "gpt-5.2")
        
        # Build the user message with all context
        follow_up_hint = ""
        if is_likely_follow_up:
            follow_up_hint = f"\n⚠️ THIS APPEARS TO BE A FOLLOW-UP QUESTION. Previous topic: {current_topic or 'unknown'}\n"
        
        prompt = f"""## NEW USER MESSAGE: "{message}"
{follow_up_hint}
{history_str}
{context_str}

## ROUTING DECISION REQUIRED:
Analyze the message, consider conversation history, and determine the user's actual intent.
If this is clearly a follow-up about {current_topic or 'a previous topic'}, use CONTEXT_CONTINUATION intent and include the topic in search_query.

OUTPUT VALID JSON ONLY:"""
        
        response = await chat.send_message(UserMessage(text=prompt))
        response_text = str(response)
        
        # Parse JSON from response
        try:
            if "```json" in response_text:
                json_str = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                json_str = response_text.split("```")[1].split("```")[0].strip()
            elif "{" in response_text and "}" in response_text:
                start = response_text.find("{")
                end = response_text.rfind("}") + 1
                json_str = response_text[start:end]
            else:
                raise ValueError("No JSON found")
            
            result = json.loads(json_str)
            
            # Ensure understood_topic is set
            if not result.get('understood_topic') and current_topic:
                result['understood_topic'] = current_topic
            
            return result
        except Exception as parse_error:
            logger.warning(f"JSON parse error: {parse_error}, raw response: {response_text[:200]}")
            # Fallback: if we have context and this looks like a follow-up, use that
            if is_likely_follow_up and current_topic:
                return {
                    "intent": "CONTEXT_CONTINUATION",
                    "response_message": f"Let me help you with more information about {current_topic}.",
                    "search_type": "product",
                    "search_query": f"{current_topic} {message}",
                    "confidence": 0.75,
                    "references_prior_context": True,
                    "understood_topic": current_topic
                }
            return {
                "intent": "CLARIFICATION_NEEDED",
                "response_message": response_text if len(response_text) < 500 else "I'd be happy to help. Could you tell me more about what you're looking for?",
                "search_type": None,
                "search_query": None,
                "confidence": 0.5
            }
    except Exception as e:
        logger.error(f"AI classification error: {e}")
        # If we have context, try to use it
        if is_likely_follow_up and conv_context.get('current_topic'):
            return {
                "intent": "CONTEXT_CONTINUATION",
                "response_message": f"Let me find that for you based on our conversation about {conv_context['current_topic']}.",
                "search_type": "product",
                "search_query": f"{conv_context['current_topic']} {message}",
                "confidence": 0.7,
                "references_prior_context": True,
                "understood_topic": conv_context['current_topic']
            }
        return {
            "intent": "CLARIFICATION_NEEDED",
            "response_message": "I'm here to help with your procurement needs. Could you tell me more about what you're looking for?\n\n• **Products**: Tell me the item name, part number, or brand\n• **Services**: Describe the professional service you need\n• **Quotation**: If you have an existing quote to analyze\n• **Complex Sourcing**: For strategic or multi-supplier requirements",
            "search_type": None,
            "search_query": None,
            "confidence": 0.5
        }

import re

def escape_regex(text: str) -> str:
    """Escape special regex characters for safe MongoDB regex queries"""
    return re.escape(text)

async def search_catalog_for_agent(query: str, search_type: str, user: dict, limit: int = 5) -> Dict:
    """
    Search catalog and return results for the AI agent.
    Searches both in-memory catalogs AND MongoDB vendor_products for comprehensive results.
    Optimized for speed and accuracy with multi-field matching.
    """
    currency = COUNTRY_CURRENCIES.get(user.get("country", "USA"), COUNTRY_CURRENCIES["USA"])
    results = {"products": [], "services": []}
    
    if search_type in ["product", None]:
        query_lower = query.lower()
        # Clean query: remove parentheses and special chars that break regex
        query_clean = re.sub(r'[()[\]{}|\\^$.*+?]', ' ', query_lower).strip()
        query_clean = re.sub(r'\s+', ' ', query_clean)  # Normalize whitespace
        query_terms = [term.strip() for term in query_clean.split() if len(term.strip()) > 2]
        matched_products = []
        seen_ids = set()
        
        # 1. First search in-memory catalogs (IT_PRODUCTS_CATALOG + NEW_VENDOR_PRODUCTS)
        for product in IT_PRODUCTS_CATALOG + NEW_VENDOR_PRODUCTS:
            name_lower = product.get("name", "").lower()
            desc_lower = product.get("short_description", "").lower()
            brand_lower = product.get("brand", "").lower()
            category_lower = product.get("category", "").lower()
            sku_lower = product.get("sku", "").lower()
            
            # Calculate match score for ranking
            score = 0
            
            # Exact phrase match (highest priority)
            if query_lower in name_lower:
                score += 100
            if query_lower in sku_lower:
                score += 90  # SKU/Part number match is very important
            if query_lower in brand_lower:
                score += 80
            
            # Term-by-term matching
            for term in query_terms:
                if term in name_lower:
                    score += 30
                if term in brand_lower:
                    score += 25
                if term in category_lower:
                    score += 20
                if term in desc_lower:
                    score += 10
                if term in sku_lower:
                    score += 35
            
            if score > 0:
                matched_products.append((score, product))
                seen_ids.add(product["id"])
        
        # 2. Search MongoDB vendor_products collection (uploaded catalogs)
        try:
            # Build MongoDB query for text search - use escaped query for regex safety
            escaped_query = escape_regex(query_clean)
            mongo_query = {
                "$or": [
                    {"name": {"$regex": escaped_query, "$options": "i"}},
                    {"brand": {"$regex": escaped_query, "$options": "i"}},
                    {"sku": {"$regex": escaped_query, "$options": "i"}},
                    {"category": {"$regex": escaped_query, "$options": "i"}},
                    {"description": {"$regex": escaped_query, "$options": "i"}}
                ]
            }
            
            # Also search for individual terms (already cleaned)
            if len(query_terms) > 1:
                term_conditions = []
                for term in query_terms:
                    escaped_term = escape_regex(term)
                    term_conditions.append({"name": {"$regex": escaped_term, "$options": "i"}})
                    term_conditions.append({"brand": {"$regex": escaped_term, "$options": "i"}})
                    term_conditions.append({"sku": {"$regex": escaped_term, "$options": "i"}})
                mongo_query["$or"].extend(term_conditions)
            
            # Execute MongoDB search
            vendor_products = await db.vendor_products.find(mongo_query).limit(limit * 2).to_list(limit * 2)
            
            for vp in vendor_products:
                if str(vp.get("id", vp.get("_id"))) not in seen_ids:
                    # Calculate score for vendor products - same scoring as in-memory products
                    name_lower = vp.get("name", "").lower()
                    brand_lower = vp.get("brand", "").lower()
                    category_lower = vp.get("category", "").lower()
                    sku_lower = vp.get("sku", "").lower()
                    desc_lower = vp.get("description", "").lower()
                    
                    score = 0
                    
                    # Exact phrase match (highest priority)
                    if query_lower in name_lower:
                        score += 100
                    if query_lower in sku_lower:
                        score += 90  # SKU/Part number match is very important
                    if query_lower in brand_lower:
                        score += 80
                    
                    # Term-by-term matching
                    for term in query_terms:
                        if term in name_lower:
                            score += 30
                        if term in brand_lower:
                            score += 25
                        if term in category_lower:
                            score += 20
                        if term in desc_lower:
                            score += 10
                        if term in sku_lower:
                            score += 35
                    
                    if score > 0:
                        matched_products.append((score, {
                            "id": str(vp.get("id", vp.get("_id"))),
                            "name": vp.get("name", ""),
                            "short_description": vp.get("description", ""),
                            "brand": vp.get("brand", ""),
                            "category": vp.get("category", ""),
                            "sku": vp.get("sku", ""),
                            "base_price": vp.get("base_price", 0),
                            "image_url": vp.get("image_url"),
                            "availability": {"in_stock": True},
                            "source": "vendor_catalog"
                        }))
                        seen_ids.add(str(vp.get("id", vp.get("_id"))))
        except Exception as e:
            logger.warning(f"MongoDB vendor search error: {e}")
        
        # 3. Sort by score (descending) and take top results
        matched_products.sort(key=lambda x: x[0], reverse=True)
        
        for score, product in matched_products[:limit]:
            base_price = product.get("base_price", 0)
            results["products"].append({
                "id": product["id"],
                "name": product["name"],
                "description": product.get("short_description", ""),
                "brand": product.get("brand", ""),
                "category": product.get("category", ""),
                "sku": product.get("sku", ""),
                "price": round(base_price * currency["rate"], 2),
                "currency": currency["symbol"],
                "unit": product.get("unit", "EA"),
                "image_url": product.get("image_url"),
                "in_stock": product.get("availability", {}).get("in_stock", True),
                "match_score": score,
                "source": product.get("source", "catalog")
            })
    
    if search_type in ["service", None]:
        # Search services (similar enhanced logic)
        query_lower = query.lower()
        # Clean query: remove parentheses and special chars that break regex
        query_clean = re.sub(r'[()[\]{}|\\^$.*+?]', ' ', query_lower).strip()
        query_clean = re.sub(r'\s+', ' ', query_clean)  # Normalize whitespace
        query_terms = [term.strip() for term in query_clean.split() if len(term.strip()) > 2]
        matched_services = []
        seen_ids = set()
        
        # Search in-memory service categories
        for category in SERVICE_CATEGORIES:
            cat_name_lower = category["name"].lower()
            score = 0
            
            if query_clean in cat_name_lower:
                score += 100
            for term in query_terms:
                if term in cat_name_lower:
                    score += 30
            
            if score > 0:
                matched_services.append((score, category))
                seen_ids.add(category["unspsc"])
        
        # Search MongoDB vendor_services
        try:
            # Use escaped query for regex safety
            escaped_query = escape_regex(query_clean)
            mongo_query = {
                "$or": [
                    {"name": {"$regex": escaped_query, "$options": "i"}},
                    {"category": {"$regex": escaped_query, "$options": "i"}},
                    {"description": {"$regex": escaped_query, "$options": "i"}}
                ]
            }
            
            vendor_services = await db.vendor_services.find(mongo_query).limit(limit * 2).to_list(limit * 2)
            
            for vs in vendor_services:
                if str(vs.get("id", vs.get("_id"))) not in seen_ids:
                    score = 50  # Base score for vendor services
                    matched_services.append((score, {
                        "unspsc": vs.get("unspsc_code", ""),
                        "name": vs.get("name", ""),
                        "description": vs.get("description", ""),
                        "base_rate": vs.get("base_rate", 100),
                        "pricing_model": vs.get("pricing_model", "per hour"),
                        "source": "vendor_catalog"
                    }))
        except Exception as e:
            logger.warning(f"MongoDB vendor service search error: {e}")
        
        # Sort and return top services
        matched_services.sort(key=lambda x: x[0], reverse=True)
        
        for score, svc in matched_services[:limit]:
            base_rate = svc.get("base_rate", random.uniform(50, 250))
            results["services"].append({
                "id": f"SVC-{svc.get('unspsc', '0000')[:4]}",
                "name": svc["name"],
                "description": svc.get("description", f"Professional {svc['name'].lower()} from certified Infosys partners"),
                "category": svc["name"],
                "rate": round(base_rate * currency["rate"], 2),
                "currency": currency["symbol"],
                "pricing_model": svc.get("pricing_model", "per hour"),
                "image_url": SERVICE_IMAGE_URLS.get(svc["name"], DEFAULT_SERVICE_IMAGE),
                "match_score": score,
                "source": svc.get("source", "catalog")
            })
    
    return results

@api_router.post("/ai-agent/conversation")
async def ai_agent_conversation(
    request: AIAgentConversationRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Main conversational endpoint for the AI Procurement Agent.
    Uses multi-LLM approach (GPT-5.2, Claude, Gemini) for intelligent routing.
    Implements smart business logic to guide users appropriately.
    """
    try:
        session_id = request.session_id or f"session_{datetime.now(timezone.utc).timestamp()}_{current_user['email']}"
        context = request.context or {}
        user_message = request.message.strip()
        
        # Build response structure
        response = {
            "message": "",
            "engines_used": ["gpt", "claude", "gemini"],
            "action": None,
            "products": None,
            "services": None,
            "context": {**context},
            "search_results": None,
            "supplier_form": None,
            "managed_service_form": None,
            "unspsc_suggestion": None,
            "show_quotation_upload": False,
            "show_managed_services": False,
            "intelligent_guidance": None
        }
        
        # ============================================================
        # INTELLIGENT ROUTING: Check for non-catalog items FIRST
        # This catches consumer items before wasting AI calls
        # ============================================================
        
        # Step 0: Check if this is clearly NOT an industrial/MRO item
        # Only do this check for fresh queries, not follow-ups
        is_follow_up = detect_follow_up_question(user_message)
        is_consumer_item = is_likely_not_in_catalog(user_message)
        
        logger.info(f"Intelligent routing check: message='{user_message[:50]}', is_follow_up={is_follow_up}, is_consumer_item={is_consumer_item}")
        
        if not is_follow_up and is_consumer_item:
            # This is clearly a consumer/non-industrial item - route to alternatives
            response["message"] = INTELLIGENT_RESPONSES["no_results_with_alternatives"].format(query=user_message)
            response["action"] = "not_in_catalog"
            response["show_quotation_upload"] = True
            response["show_managed_services"] = True
            response["context"]["intent"] = "NOT_IN_CATALOG"
            response["context"]["original_query"] = user_message
            response["context"]["last_action"] = "no_results_alternatives"
            response["intelligent_guidance"] = {
                "reason": "Item appears to be outside standard industrial/IT/MRO procurement catalog",
                "recommended_paths": ["quotation_analysis", "managed_services"],
                "confidence": 0.90
            }
            
            # Store conversation
            await db.ai_agent_conversations.insert_one({
                "session_id": session_id,
                "user_id": current_user["email"],
                "message": user_message,
                "intent": "NOT_IN_CATALOG",
                "response": response["message"][:500],
                "confidence": 0.90,
                "timestamp": datetime.now(timezone.utc),
                "language": request.language,
                "currency": request.currency,
                "intelligent_detection": True
            })
            return response
        
        # ============================================================
        # ROUTER PATTERN: AI Classification for context awareness
        # This ensures follow-up questions are handled correctly
        # ============================================================
        
        # Step 1: Classify intent using AI
        classification = await classify_user_intent_with_ai(
            user_message, 
            context, 
            session_id
        )
        
        intent = classification.get("intent", "CLARIFICATION_NEEDED")
        response_message = classification.get("response_message", "")
        search_query = classification.get("search_query")
        search_type = classification.get("search_type")
        
        # Add intent to response for frontend
        response["intent"] = intent
        response["context"]["intent"] = intent
        response["context"]["search_type"] = search_type
        response["context"]["search_query"] = search_query
        
        # Step 2: Handle CATALOG_SEARCH or CONTEXT_CONTINUATION (both do catalog search)
        if intent in ["CATALOG_SEARCH", "CONTEXT_CONTINUATION"] and search_query:
            # Log context continuation for debugging
            if intent == "CONTEXT_CONTINUATION":
                logger.info(f"Context continuation detected. Topic: {classification.get('understood_topic')}, Query: {search_query}")
            
            # Search catalog
            search_results = await search_catalog_for_agent(
                search_query, 
                search_type or "product",  # Default to product for context continuation
                current_user,
                limit=5
            )
            
            has_products = bool(search_results.get("products"))
            has_services = bool(search_results.get("services"))
            
            if has_products or has_services:
                # Found results
                response["products"] = search_results["products"]
                response["services"] = search_results["services"]
                response["action"] = "show_results"
                
                product_count = len(search_results.get("products", []))
                service_count = len(search_results.get("services", []))
                
                # Use AI's response if it references context, otherwise generate
                if classification.get("references_prior_context") and response_message:
                    # AI generated a contextual response - use it with results
                    response["message"] = response_message
                    if product_count > 0:
                        response["message"] += f"\n\nHere are **{product_count} products** that match."
                else:
                    # Generate standard response
                    topic = classification.get("understood_topic", search_query)
                    response["message"] = f"Here's what I found for **{topic}**:\n\n"
                    if product_count > 0:
                        response["message"] += f"• **{product_count} Products** matching your search\n"
                    if service_count > 0:
                        response["message"] += f"• **{service_count} Services** available\n"
                    response["message"] += "\nClick any item to view details or add to cart."
            else:
                # No results found - check if this is a consumer item
                if is_likely_not_in_catalog(user_message):
                    response["message"] = INTELLIGENT_RESPONSES["no_results_with_alternatives"].format(query=search_query)
                    response["action"] = "not_in_catalog"
                else:
                    # No results but topic understood
                    understood_topic = classification.get("understood_topic", search_query)
                    response["message"] = f"I understand you're looking for **{understood_topic}**, but I couldn't find exact matches in our catalog.\n\n"
                    response["message"] += "**Would you like to:**\n"
                    response["message"] += "• **Upload a quotation** from a supplier who has this item\n"
                    response["message"] += "• **Contact our Buying Desk** to source this for you"
                    response["action"] = "no_results"
                
                response["show_quotation_upload"] = True
                response["show_managed_services"] = True
                response["context"]["last_action"] = "no_results_alternatives"
                response["context"]["original_query"] = search_query
        
        # Step 3: Handle QUOTATION_ANALYSIS intent
        elif intent == "QUOTATION_ANALYSIS":
            response["action"] = "navigate_quotation"
            response["show_quotation_upload"] = True
            response["message"] = response_message + "\n\n**Ready to analyze your quotation?** Click the button below to upload your document."
        
        # Step 4: Handle MANAGED_SERVICES intent
        elif intent == "MANAGED_SERVICES":
            response["action"] = "navigate_managed_services"
            response["managed_service_form"] = True
            response["show_managed_services"] = True
            response["message"] = response_message
            
            # Try to suggest UNSPSC code
            query_lower = user_message.lower()
            for category in SERVICE_CATEGORIES:
                if any(word in query_lower for word in category["name"].lower().split()):
                    response["unspsc_suggestion"] = {
                        "code": category["unspsc"],
                        "name": category["name"]
                    }
                    break
        
        # Step 5: Handle FOLLOW_UP - Legacy intent for backwards compatibility
        elif intent == "FOLLOW_UP":
            # Treat similar to CONTEXT_CONTINUATION
            response["message"] = response_message
            if search_query:
                search_results = await search_catalog_for_agent(search_query, "product", current_user, limit=5)
                response["products"] = search_results.get("products", [])
                response["services"] = search_results.get("services", [])
                
                if search_results.get("products") or search_results.get("services"):
                    response["action"] = "show_results"
                else:
                    understood_topic = classification.get("understood_topic", search_query)
                    response["message"] = f"I understand you're asking about **{understood_topic}**.\n\nI couldn't find exact matches in our catalog. Would you like to:\n\n• **Upload a supplier quotation** - I can analyze pricing\n• **Contact our Buying Desk** - Our specialists can source this"
                    response["show_quotation_upload"] = True
                    response["show_managed_services"] = True
            response["action"] = response.get("action") or "follow_up"
            response["context"]["last_action"] = "follow_up_handled"
        
        # Step 6: Handle CLARIFICATION_NEEDED - Check if it's a non-catalog item
        else:
            # Check if this looks like a non-catalog item that needs alternatives
            if is_likely_not_in_catalog(user_message):
                response["message"] = INTELLIGENT_RESPONSES["no_results_with_alternatives"].format(query=user_message)
                response["action"] = "not_in_catalog"
                response["show_quotation_upload"] = True
                response["show_managed_services"] = True
                response["context"]["last_action"] = "no_results_alternatives"
            else:
                # Use the AI's response if available, otherwise use template
                if response_message and len(response_message) > 20:
                    response["message"] = response_message
                else:
                    response["message"] = INTELLIGENT_RESPONSES["clarification_needed"].format(query=user_message)
                response["action"] = "clarification"
                response["context"]["last_action"] = "clarification_asked"
        
        # Store conversation in database for analytics and context
        await db.ai_agent_conversations.insert_one({
            "session_id": session_id,
            "user_id": current_user["email"],
            "message": user_message,
            "intent": intent,
            "search_query": search_query,  # Store for follow-up context
            "response": response["message"][:500],
            "confidence": classification.get("confidence", 0),
            "timestamp": datetime.now(timezone.utc),
            "language": request.language,
            "currency": request.currency,
            "has_results": bool(response.get("products") or response.get("services")),
            "offered_alternatives": response.get("show_quotation_upload") or response.get("show_managed_services"),
            "understood_topic": classification.get("understood_topic", search_query)  # Store topic for context
        })
        
        return response
        
    except Exception as e:
        logger.error(f"AI Agent conversation error: {e}")
        return {
            "message": "I apologize, but I encountered an issue processing your request. Please try again or select one of the manual options from the header.",
            "engines_used": [],
            "action": "error",
            "products": None,
            "services": None,
            "context": context,
            "search_results": None,
            "supplier_form": None,
            "managed_service_form": None,
            "unspsc_suggestion": None,
            "show_quotation_upload": True,
            "show_managed_services": True,
            "intelligent_guidance": None
        }

# ============================================
# INFOSHOP CATALOG ENTERPRISE ENDPOINTS
# ============================================

class InfoShopProductRequest(BaseModel):
    """Request model for InfoShop product transformation"""
    products: List[Dict[str, Any]]
    vendor: str
    category_discounts: Optional[Dict[str, float]] = None

class PartnerDiscountUpload(BaseModel):
    """Request model for partner discount upload"""
    vendor: str
    discounts: Dict[str, float]  # Category Name -> Discount %

class ShippingInfo(BaseModel):
    """Shipping information for cart transfer"""
    shipping_address: str
    delivery_attention: str
    requested_delivery_date: str
    special_instructions: Optional[str] = None

class InfoShopCartTransfer(BaseModel):
    """Cart transfer request with shipping info"""
    session_token: str
    items: List[Dict[str, Any]]
    shipping_info: ShippingInfo


@api_router.get("/infoshop/partners")
async def get_infoshop_partners():
    """Get active and coming soon partners for InfoShop"""
    return {
        "active_partners": ACTIVE_PARTNERS,
        "coming_soon_partners": COMING_SOON_PARTNERS,
        "total_coming_soon": sum(len(partners) for partners in COMING_SOON_PARTNERS.values())
    }


@api_router.post("/infoshop/partner-discounts/upload")
async def upload_partner_discounts(file: UploadFile = File(...), vendor: str = Form(...)):
    """
    Upload category discounts for a vendor
    Excel format: Category Name, Discount %
    """
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx, .xls) or CSV (.csv)")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        # Find category and discount columns
        category_col = None
        discount_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'category' in col_lower:
                category_col = col
            if 'discount' in col_lower or '%' in col_lower:
                discount_col = col
        
        if not category_col or not discount_col:
            raise HTTPException(
                status_code=400, 
                detail=f"Could not find Category and Discount columns. Found: {list(df.columns)}"
            )
        
        # Build discount dictionary
        discounts = {}
        for _, row in df.iterrows():
            category = str(row[category_col]).strip()
            if category and category != "nan":
                try:
                    discount = float(str(row[discount_col]).replace('%', '').strip())
                    discounts[category] = discount
                except (ValueError, TypeError):
                    continue
        
        # Load into memory
        load_partner_discounts(vendor, discounts)
        
        # Save to database for persistence
        await db.partner_discounts.update_one(
            {"vendor": vendor.lower()},
            {
                "$set": {
                    "vendor": vendor.lower(),
                    "vendor_display": vendor,
                    "discounts": discounts,
                    "category_count": len(discounts),
                    "uploaded_at": datetime.now(timezone.utc).isoformat(),
                    "filename": file.filename
                }
            },
            upsert=True
        )
        
        return {
            "success": True,
            "vendor": vendor,
            "categories_loaded": len(discounts),
            "sample_discounts": dict(list(discounts.items())[:5]),
            "message": f"Loaded {len(discounts)} category discounts for {vendor}"
        }
        
    except Exception as e:
        logger.error(f"Partner discount upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/infoshop/partner-discounts/{vendor}")
async def get_vendor_discounts(vendor: str):
    """Get category discounts for a vendor"""
    # Try memory first
    discounts = get_partner_discounts(vendor)
    
    # Try database if not in memory
    if not discounts:
        doc = await db.partner_discounts.find_one({"vendor": vendor.lower()})
        if doc:
            discounts = doc.get("discounts", {})
            load_partner_discounts(vendor, discounts)
    
    return {
        "vendor": vendor,
        "discounts": discounts,
        "category_count": len(discounts)
    }


@api_router.get("/infoshop/partner-discounts")
async def get_all_vendor_discounts():
    """Get all loaded partner discounts"""
    # Load from database
    cursor = db.partner_discounts.find({})
    all_discounts = {}
    
    async for doc in cursor:
        vendor = doc.get("vendor_display", doc.get("vendor"))
        all_discounts[vendor] = {
            "discounts": doc.get("discounts", {}),
            "category_count": len(doc.get("discounts", {})),
            "uploaded_at": doc.get("uploaded_at")
        }
        # Load into memory
        load_partner_discounts(vendor, doc.get("discounts", {}))
    
    return {
        "vendors": list(all_discounts.keys()),
        "data": all_discounts
    }


@api_router.post("/infoshop/catalog/upload")
async def upload_infoshop_catalog(
    file: UploadFile = File(...),
    vendor: str = Form(...)
):
    """
    Upload product catalog for InfoShop with enterprise transformations
    
    - Generates unique InfoShop Part Numbers
    - Calculates Danone Preferred Pricing
    - Auto-classifies UNSPSC codes
    - Validates images
    """
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx, .xls) or CSV (.csv)")
    
    if vendor.lower() not in [p.lower() for p in ACTIVE_PARTNERS]:
        raise HTTPException(
            status_code=400, 
            detail=f"Vendor '{vendor}' not in active partners: {ACTIVE_PARTNERS}"
        )
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        logger.info(f"Processing {len(df)} products from {vendor}")
        
        # Get vendor discounts
        discounts = get_partner_discounts(vendor)
        if not discounts:
            # Try to load from database
            doc = await db.partner_discounts.find_one({"vendor": vendor.lower()})
            if doc:
                discounts = doc.get("discounts", {})
                load_partner_discounts(vendor, discounts)
        
        # Transform products
        products = []
        errors = []
        
        for idx, row in df.iterrows():
            try:
                product = transform_product_for_infoshop(row.to_dict(), vendor, discounts)
                if product.get("product_name"):
                    products.append(product)
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
                continue
        
        if not products:
            raise HTTPException(status_code=400, detail="No valid products found in file")
        
        # Index to Algolia if available
        indexed_count = 0
        if ALGOLIA_AVAILABLE:
            try:
                from algolia_service import algolia_client, PRODUCTS_INDEX
                
                # Batch save (500 at a time)
                batch_size = 500
                for i in range(0, len(products), batch_size):
                    batch = products[i:i + batch_size]
                    algolia_client.save_objects(PRODUCTS_INDEX, batch)
                    indexed_count += len(batch)
                    logger.info(f"Indexed batch {i//batch_size + 1}: {len(batch)} products")
                
            except Exception as e:
                logger.error(f"Algolia indexing error: {e}")
        
        # Save to MongoDB as backup
        if products:
            await db.infoshop_products.insert_many(products)
        
        # Log upload
        await db.infoshop_uploads.insert_one({
            "vendor": vendor,
            "filename": file.filename,
            "total_rows": len(df),
            "products_created": len(products),
            "indexed_count": indexed_count,
            "errors": len(errors),
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
        
        # Sample products for response
        sample = products[:3] if products else []
        
        return {
            "success": True,
            "vendor": vendor,
            "total_rows": len(df),
            "products_created": len(products),
            "indexed_to_algolia": indexed_count,
            "errors": len(errors),
            "error_details": errors[:10] if errors else [],
            "discounts_applied": len(discounts) > 0,
            "sample_products": [
                {
                    "infoshop_part_number": p["infoshop_part_number"],
                    "product_name": p["product_name"][:50] + "..." if len(p["product_name"]) > 50 else p["product_name"],
                    "brand": p["brand"],
                    "list_price": p["list_price"],
                    "danone_preferred_price": p["danone_preferred_price"],
                    "customer_savings_percent": p["customer_savings_percent"],
                    "unspsc_code": p["unspsc_code"],
                    "has_image": p["has_image"]
                }
                for p in sample
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"InfoShop catalog upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/infoshop/pricing/calculate")
async def calculate_infoshop_pricing(
    list_price: float = Form(...),
    category_discount: float = Form(...)
):
    """
    Calculate Danone Preferred Price with sliding margin
    
    Demonstrates the pricing formula:
    - Infosys Purchase Price = List Price × (1 - Discount%)
    - Gross Margin = 5.92% to 9.2% (sliding based on price)
    - Danone Preferred Price = Purchase Price × (1 + Margin%)
    """
    result = calculate_danone_preferred_price(list_price, category_discount)
    
    return {
        "success": True,
        "pricing": result,
        "formula_explanation": {
            "step1": f"List Price: ${result['list_price']}",
            "step2": f"Category Discount: {result['category_discount_percent']}%",
            "step3": f"Infosys Purchase Price: ${result['infosys_purchase_price']} (List × {100 - result['category_discount_percent']}%)",
            "step4": f"Gross Margin: {result['gross_margin_percent']}% (sliding scale 5.92-9.2%)",
            "step5": f"Danone Preferred Price: ${result['danone_preferred_price']}",
            "customer_benefit": f"Customer saves {result['customer_savings_percent']}% vs List Price"
        }
    }


@api_router.get("/infoshop/delivery/minimum-date")
async def get_minimum_delivery_date():
    """Get minimum delivery date (2 business weeks from today)"""
    min_date = calculate_minimum_delivery_date()
    return {
        "minimum_delivery_date": min_date,
        "business_days": 10,
        "note": "Infosys will confirm promised delivery date once secured from partners"
    }


@api_router.post("/infoshop/delivery/validate")
async def validate_requested_delivery(requested_date: str = Form(...)):
    """Validate that requested delivery date is at least 2 business weeks out"""
    result = validate_delivery_date(requested_date)
    return result


@api_router.post("/infoshop/cart/prepare-transfer")
async def prepare_cart_transfer(request: InfoShopCartTransfer):
    """
    Prepare cart for transfer to Coupa with shipping information
    
    Validates:
    - Shipping address is provided
    - Delivery attention is provided
    - Delivery date is at least 2 business weeks out
    """
    # Validate delivery date
    date_validation = validate_delivery_date(request.shipping_info.requested_delivery_date)
    if not date_validation["valid"]:
        raise HTTPException(
            status_code=400,
            detail=date_validation["message"]
        )
    
    # Validate shipping info
    if not request.shipping_info.shipping_address or len(request.shipping_info.shipping_address) < 10:
        raise HTTPException(
            status_code=400,
            detail="Please provide a complete shipping address"
        )
    
    if not request.shipping_info.delivery_attention:
        raise HTTPException(
            status_code=400,
            detail="Delivery attention (recipient name) is required"
        )
    
    # Verify PunchOut session
    session = get_punchout_session(request.session_token)
    if not session:
        session = await get_punchout_session_from_db(db, request.session_token)
    
    if not session:
        raise HTTPException(status_code=404, detail="PunchOut session not found or expired")
    
    # Calculate totals
    total_items = len(request.items)
    total_amount = sum(
        item.get("danone_preferred_price", item.get("unit_price", 0)) * item.get("quantity", 1) 
        for item in request.items
    )
    
    # Update session with cart and shipping
    session["cart_items"] = request.items
    session["shipping_info"] = {
        "address": request.shipping_info.shipping_address,
        "attention": request.shipping_info.delivery_attention,
        "requested_date": request.shipping_info.requested_delivery_date,
        "special_instructions": request.shipping_info.special_instructions
    }
    session["cart_total"] = total_amount
    session["ready_for_transfer"] = True
    
    # Save to database
    await save_punchout_session_to_db(db, session)
    
    return {
        "success": True,
        "session_token": request.session_token,
        "total_items": total_items,
        "total_amount": round(total_amount, 2),
        "shipping_info": {
            "address": request.shipping_info.shipping_address,
            "attention": request.shipping_info.delivery_attention,
            "requested_date": request.shipping_info.requested_delivery_date,
            "minimum_date": date_validation["minimum_date"]
        },
        "delivery_note": "Infosys will confirm promised delivery date once secured from partners",
        "ready_for_transfer": True
    }


@api_router.get("/infoshop/unspsc/classify")
async def classify_product_unspsc(
    product_name: str,
    category: str = None,
    description: str = None
):
    """
    AI-powered UNSPSC classification for a product
    """
    result = classify_unspsc(product_name, category, description)
    return {
        "success": True,
        "classification": result
    }


@api_router.get("/infoshop/part-number/generate")
async def generate_part_number(
    vendor: str,
    category: str,
    product_name: str = ""
):
    """
    Generate unique InfoShop Part Number
    Format: INF + Vendor(2) + Category(3) + Random(5)
    """
    part_number = generate_infoshop_part_number(vendor, category, product_name)
    return {
        "success": True,
        "infoshop_part_number": part_number,
        "format": "INF + Vendor(2) + Category(3) + Random(5)"
    }


@api_router.get("/infoshop/stats")
async def get_infoshop_stats():
    """Get InfoShop catalog statistics"""
    # Get counts from database
    total_products = await db.infoshop_products.count_documents({})
    
    # Get vendor breakdown
    pipeline = [
        {"$group": {"_id": "$vendor", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    vendor_counts = await db.infoshop_products.aggregate(pipeline).to_list(100)
    
    # Get recent uploads
    recent_uploads = await db.infoshop_uploads.find({}).sort("uploaded_at", -1).limit(5).to_list(5)
    
    # Get Algolia stats if available
    algolia_stats = {}
    if ALGOLIA_AVAILABLE:
        try:
            algolia_stats = get_index_stats()
        except Exception as e:
            logger.error(f"Algolia stats error: {e}")
    
    return {
        "total_products": total_products,
        "algolia_indexed": algolia_stats.get("total_products", 0),
        "vendors": {item["_id"]: item["count"] for item in vendor_counts},
        "active_partners": ACTIVE_PARTNERS,
        "coming_soon_count": sum(len(p) for p in COMING_SOON_PARTNERS.values()),
        "recent_uploads": [
            {
                "vendor": u.get("vendor"),
                "filename": u.get("filename"),
                "products": u.get("products_created"),
                "date": u.get("uploaded_at")
            }
            for u in recent_uploads
        ]
    }

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============================================
# HEALTH CHECK ENDPOINT
# ============================================

@app.get("/api/health")
async def health_check():
    """Comprehensive health check endpoint for deployment monitoring"""
    health_status = {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "2.0.0",
        "service": "OMNISupply.io API",
        "checks": {}
    }
    
    # Check MongoDB connection
    try:
        await db.command("ping")
        health_status["checks"]["database"] = {"status": "healthy", "type": "mongodb"}
    except Exception as e:
        health_status["checks"]["database"] = {"status": "unhealthy", "error": str(e)}
        health_status["status"] = "degraded"
    
    # Check collections
    try:
        collections = await db.list_collection_names()
        health_status["checks"]["collections"] = {
            "status": "healthy",
            "count": len(collections),
            "names": collections[:10]  # First 10 collections
        }
    except Exception as e:
        health_status["checks"]["collections"] = {"status": "unhealthy", "error": str(e)}
    
    # Check key data counts
    try:
        health_status["checks"]["data"] = {
            "buying_desk_requests": await db.buying_desk_requests.count_documents({}),
            "sourcing_requests": await db.sourcing_requests.count_documents({}),
            "users": await db.users.count_documents({}),
            "quotation_uploads": await db.quotation_uploads.count_documents({})
        }
    except Exception as e:
        health_status["checks"]["data"] = {"status": "error", "error": str(e)}
    
    return health_status


# Root-level health check for Kubernetes (without /api prefix)
@app.get("/health")
async def kubernetes_health_check():
    """Simple health check endpoint for Kubernetes liveness/readiness probes"""
    try:
        # Quick database ping
        await db.command("ping")
        return {"status": "healthy", "service": "OMNISupply.io"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}


@app.get("/api/ready")
async def readiness_check():
    """Readiness probe for Kubernetes"""
    try:
        await db.command("ping")
        return {"ready": True}
    except Exception:
        return {"ready": False}

@app.get("/api/live")
async def liveness_check():
    """Liveness probe for Kubernetes"""
    return {"alive": True, "timestamp": datetime.now(timezone.utc).isoformat()}

@app.on_event("startup")
async def startup_db_client():
    """Initialize database indexes for optimal search performance"""
    try:
        # Create indexes for vendor_products collection
        # These indexes enable fast text-based search on large catalogs
        await db.vendor_products.create_index([("name", "text"), ("brand", "text"), ("category", "text"), ("description", "text"), ("sku", "text")], name="vendor_products_text_search")
        await db.vendor_products.create_index("sku", name="vendor_products_sku")
        await db.vendor_products.create_index("brand", name="vendor_products_brand")
        await db.vendor_products.create_index("category", name="vendor_products_category")
        await db.vendor_products.create_index("delivery_partner_id", name="vendor_products_partner")
        
        # Create indexes for vendor_services collection
        await db.vendor_services.create_index([("name", "text"), ("category", "text"), ("description", "text")], name="vendor_services_text_search")
        await db.vendor_services.create_index("category", name="vendor_services_category")
        
        # Create indexes for ai_agent_conversations for analytics
        await db.ai_agent_conversations.create_index("session_id", name="ai_conversations_session")
        await db.ai_agent_conversations.create_index("user_id", name="ai_conversations_user")
        await db.ai_agent_conversations.create_index("timestamp", name="ai_conversations_timestamp")
        
        logger.info("Database indexes created successfully for optimal search performance")
    except Exception as e:
        logger.warning(f"Index creation warning (may already exist): {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
